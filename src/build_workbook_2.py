"""Stage 2: Portfolios sheet (holdings + asset-class weight roll-up via SUMIFS) and
Asset Class Returns (raw monthly data)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

from portfolios import PORTFOLIOS, AC, portfolio_summary

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.0%"
PCT3 = "0.000%"

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())

ASSET_CLASSES = list(AC.keys())  # fixed order used everywhere

# ---------------------------------------------------------------------
# Portfolios
# ---------------------------------------------------------------------
ws = wb.create_sheet("Portfolios")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
for col, w in zip("BCDEF", [46, 22, 10, 10, 14]):
    ws.column_dimensions[col].width = w
ws["B2"] = "Portfolio Definitions"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("Holdings and weights sourced from the FNZ 'Growth Passive Plus' holdings file. "
            "'Better' portfolio weights are a judgement-based diversification overlay - see Instructions.")
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)

holding_ranges = {}  # portfolio -> (first_row, last_row) of holdings table
row = 5
for name in PORTFOLIOS:
    ws.cell(row=row, column=2, value=f"{name} portfolio").font = BOLD
    ws.cell(row=row, column=2).fill = SUBHEAD_FILL
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = SUBHEAD_FILL
    row += 1
    headers = ["Holding", "Asset class", "Weight", "OCF (% pa)", "Fee contribution"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h).font = HEADER_FONT
        ws.cell(row=row, column=2 + i).fill = HEADER_FILL
    header_row = row
    row += 1
    first_data_row = row
    df = portfolio_summary(name)
    for _, r_ in df.iterrows():
        ws.cell(row=row, column=2, value=r_["Holding"]).font = BLACK
        ws.cell(row=row, column=3, value=r_["AssetClass"]).font = BLUE
        wcell = ws.cell(row=row, column=4, value=float(r_["Weight"]))
        wcell.font = BLUE
        wcell.number_format = PCT
        ocell = ws.cell(row=row, column=5, value=float(r_["OCF"]))
        ocell.font = BLUE
        ocell.number_format = PCT3
        fcell = ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        fcell.number_format = PCT3
        row += 1
    last_data_row = row - 1
    # totals
    ws.cell(row=row, column=2, value="Total / weighted-avg OCF").font = BOLD
    ws.cell(row=row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})").number_format = PCT
    ws.cell(row=row, column=4).font = BOLD
    ws.cell(row=row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})/D{row}").number_format = PCT3
    ws.cell(row=row, column=6).font = BOLD
    holding_ranges[name] = (first_data_row, last_data_row, row)  # row = total row (has weighted OCF)
    row += 3

Path("holding_ranges.json").write_text(json.dumps(holding_ranges))

# ---------------------------------------------------------------------
# Asset-class net weight roll-up (SUMIFS from holdings tables above) - used by Portfolio Annual
# Returns via SUMPRODUCT. Laid out horizontally: one row per portfolio, one column per asset class,
# in the SAME order as Asset Class Returns / Annual Asset Returns columns.
# ---------------------------------------------------------------------
row += 1
ws.cell(row=row, column=2, value="Asset-class net weights (roll-up, used by return calculations)").font = BOLD
row += 1
weight_header_row = row
ws.cell(row=row, column=2, value="Portfolio").font = HEADER_FONT
ws.cell(row=row, column=2).fill = HEADER_FILL
for i, ac in enumerate(ASSET_CLASSES):
    c = 3 + i
    cell = ws.cell(row=row, column=c, value=ac)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(c)].width = 13
row += 1
weight_rows = {}
for name in PORTFOLIOS:
    ws.cell(row=row, column=2, value=name).font = BOLD
    fr, lr, tot = holding_ranges[name]
    for i, ac in enumerate(ASSET_CLASSES):
        c = 3 + i
        col_letter = get_column_letter(c)
        formula = f'=SUMIFS($D${fr}:$D${lr},$C${fr}:$C${lr},{col_letter}${weight_header_row})'
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = PCT
    weight_rows[name] = row
    row += 1

Path("weight_rows.json").write_text(json.dumps({"header_row": weight_header_row, "rows": weight_rows,
                                                  "asset_classes": ASSET_CLASSES}))

# ---------------------------------------------------------------------
# Asset Class Returns (raw monthly source data)
# ---------------------------------------------------------------------
ac_df = pd.read_csv("/home/claude/mobius_decumulation/data/asset_class_returns.csv", index_col=0, parse_dates=True)
ac_df = ac_df[list(AC.values())]  # order columns to match ASSET_CLASSES
ac_df.columns = ASSET_CLASSES

ws2 = wb.create_sheet("Asset Class Returns")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B3"
ws2["A1"] = "Monthly total returns by asset class - Source: Bloomberg, provided by user, 9 July 2026 (UK CPI is YoY level, not a return)"
ws2["A1"].font = SUBTITLE_FONT
ws2.cell(row=2, column=1, value="Month end").font = HEADER_FONT
ws2.cell(row=2, column=1).fill = HEADER_FILL
ws2.column_dimensions["A"].width = 12
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    cell = ws2.cell(row=2, column=c, value=ac)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws2.column_dimensions[get_column_letter(c)].width = 13

for r_i, (dt, r_) in enumerate(ac_df.iterrows()):
    row_n = 3 + r_i
    dcell = ws2.cell(row=row_n, column=1, value=dt.to_pydatetime())
    dcell.number_format = "mmm-yyyy"
    for i, ac in enumerate(ASSET_CLASSES):
        v = r_[ac]
        c = 2 + i
        cell = ws2.cell(row=row_n, column=c)
        if pd.notna(v):
            cell.value = float(v)
            cell.font = BLUE
            cell.number_format = PCT if ac != "" else PCT
        cell.number_format = "0.00%" if ac != "" else "0.00%"

last_row = 2 + len(ac_df)
Path("asset_returns_range.json").write_text(json.dumps({
    "first_row": 3, "last_row": last_row, "asset_classes": ASSET_CLASSES, "n_months": len(ac_df)
}))

wb.save(OUT_PATH)
print("Saved stage 2. Rows:", holding_ranges)
print("Asset returns rows 3..", last_row, "n=", len(ac_df))
