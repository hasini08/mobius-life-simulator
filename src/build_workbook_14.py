"""Stage 14: Tax & State Pension. Adds the toggle/State Pension inputs to Inputs, and a new 'Tax'
sheet holding the UK income tax bands (editable, rest-of-UK, 2026/27) plus the derived closed-form
breakpoints used everywhere else in the workbook to gross up a desired NET spend into the actual
taxable withdrawal needed from the pot - the Excel-native mirror of src/tax.py's net-to-gross
inversion (same derivation, see that module's docstring for the maths)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent))
import tax as tax_mod

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.0%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())

# ---------------------------------------------------------------------
# Inputs: tax/State Pension toggle block
# ---------------------------------------------------------------------
ws = wb["Inputs"]
r0 = 23
ws.cell(row=r0, column=2, value="Tax & State Pension").font = BOLD
tax_rows = [
    ("apply_tax_on", "Include income tax + State Pension? (Y/N)", "N", None,
     "Y = 'Desired annual spend' above is treated as NET (take-home); the model grosses it up"),
    ("sp_annual", "Full State Pension, today's £ per year", tax_mod.FULL_NEW_STATE_PENSION_ANNUAL, GBP,
     "Defaults to the 2026/27 full new State Pension (£241.30/week). Lower if not entitled to the full amount"),
    ("sp_age", "State Pension age", tax_mod.DEFAULT_STATE_PENSION_AGE, None,
     "Varies by date of birth - confirm the client's own via gov.uk/state-pension-age"),
]
r = r0 + 1
labels = {}
for key, label, val, fmt, note in tax_rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=val)
    cell.font = BLUE
    cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = SUBTITLE_FONT
    labels[key] = r
    r += 1
for key, row in labels.items():
    NR[key] = f"Inputs!$C${row}"
Path("cellrefs.json").write_text(json.dumps(NR, indent=2))
print("Added tax/SP inputs:", labels)

# ---------------------------------------------------------------------
# Tax sheet
# ---------------------------------------------------------------------
if "Tax" in wb.sheetnames:
    del wb["Tax"]
ws = wb.create_sheet("Tax")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 55

ws["B2"] = "Tax & State Pension"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "UK rest-of-UK (England/Wales/Northern Ireland) income tax, 2026/27 rates - Scotland has different "
    "bands. SIMPLIFIED BASIS: the whole pot is treated as a taxable pension wrapper (every pound "
    "withdrawn counts as income) - no 25% pension-commencement lump sum, ISA (tax-free) or GIA (capital "
    "gains) modelling yet, so this slightly OVERSTATES tax if some of a client's money is actually in "
    "tax-free wrappers. Tax bands and the State Pension (Inputs tab) are both held in TODAY'S MONEY "
    "(assumed to rise with inflation) for the whole plan - a standard simplifying convention for "
    "long-horizon retirement modelling, rather than literally freezing today's nominal thresholds for "
    "30 years."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 70

ws.cell(row=5, column=2, value="Tax bands (editable)").font = BOLD
band_rows = [
    ("pa", "Personal Allowance", tax_mod.PERSONAL_ALLOWANCE, GBP),
    ("basic_limit", "Basic rate (20%) upper limit", tax_mod.BASIC_RATE_LIMIT, GBP),
    ("higher_limit", "Higher rate (40%) upper limit / additional rate starts", tax_mod.HIGHER_RATE_LIMIT, GBP),
    ("taper_start", "Personal Allowance taper starts", tax_mod.PA_TAPER_START, GBP),
    ("basic_rate", "Basic rate", tax_mod.BASIC_RATE, PCT),
    ("higher_rate", "Higher rate", tax_mod.HIGHER_RATE, PCT),
    ("add_rate", "Additional rate", tax_mod.ADDITIONAL_RATE, PCT),
]
r = 6
band_cells = {}
for key, label, val, fmt in band_rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=val)
    cell.font = BLUE
    cell.fill = INPUT_FILL
    cell.number_format = fmt
    band_cells[key] = f"$C${r}"
    r += 1

r += 1
ws.cell(row=r, column=2, value="Derived breakpoints (do not edit - formulas)").font = BOLD
r += 1
PA, BL, HL, TS = band_cells["pa"], band_cells["basic_limit"], band_cells["higher_limit"], band_cells["taper_start"]
BR, HR, AR = band_cells["basic_rate"], band_cells["higher_rate"], band_cells["add_rate"]

deriv = {}
def put(key, label, formula, fmt=GBP):
    global r
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=formula)
    cell.number_format = fmt
    deriv[key] = f"Tax!$C${r}"
    r += 1

put("tax_x2", "Tax due at basic-rate limit (X2)", f"={BR}*({BL}-{PA})")
put("tax_x3", "Tax due at PA-taper start (X3)", f"={deriv['tax_x2']}+{HR}*({TS}-{BL})".replace("Tax!$C$", "C"))
put("tax_x4", "Tax due at higher-rate limit (X4, 60% taper zone)", f"={deriv['tax_x3']}+0.6*({HL}-{TS})".replace("Tax!$C$", "C"))
put("n1", "Net income breakpoint N1 (= PA)", f"={PA}")
put("n2", "Net income breakpoint N2", f"={BL}-{deriv['tax_x2']}".replace("Tax!$C$", "C"))
put("n3", "Net income breakpoint N3", f"={TS}-{deriv['tax_x3']}".replace("Tax!$C$", "C"))
put("n4", "Net income breakpoint N4", f"={HL}-{deriv['tax_x4']}".replace("Tax!$C$", "C"))
put("seg2_i", "Segment 2 intercept (net = intercept + slope*gross)", f"={PA}*{BR}")
put("seg2_s", "Segment 2 slope", f"=1-{BR}", PCT)
put("seg3_i", "Segment 3 intercept", f"={HR}*{BL}-{deriv['tax_x2']}".replace("Tax!$C$", "C"))
put("seg3_s", "Segment 3 slope", f"=1-{HR}", PCT)
put("seg4_i", "Segment 4 intercept", f"=0.6*{TS}-{deriv['tax_x3']}".replace("Tax!$C$", "C"))
put("seg4_s", "Segment 4 slope", "=0.4", PCT)
put("seg5_i", "Segment 5 intercept", f"={AR}*{HL}-{deriv['tax_x4']}".replace("Tax!$C$", "C"))
put("seg5_s", "Segment 5 slope", f"=1-{AR}", PCT)

# also expose the raw band cells, sheet-qualified, for tax_due()-style formulas built from OTHER
# sheets (band_cells itself holds LOCAL refs like "$C$6", used above for Tax-sheet-local formulas)
for key, local_ref in band_cells.items():
    deriv[key] = f"Tax!{local_ref}"

Path("tax_range.json").write_text(json.dumps(deriv, indent=2))
print("Tax sheet breakpoints:", deriv)

# --- illustrative net-vs-gross box, mirroring the app ---
r += 2
ws.cell(row=r, column=2, value="Illustrative: what this means for the Inputs-tab client").font = BOLD
r += 1
illus_header = r
r += 1


def gross_for_net_formula(n_expr):
    N1, N2, N3, N4 = deriv["n1"], deriv["n2"], deriv["n3"], deriv["n4"]
    s2i, s2s, s3i, s3s = deriv["seg2_i"], deriv["seg2_s"], deriv["seg3_i"], deriv["seg3_s"]
    s4i, s4s, s5i, s5s = deriv["seg4_i"], deriv["seg4_s"], deriv["seg5_i"], deriv["seg5_s"]
    n = f"MAX({n_expr},0)"
    return (f'IF({n}<={N1},{n},IF({n}<={N2},({n}-{s2i})/{s2s},IF({n}<={N3},({n}-{s3i})/{s3s},'
            f'IF({n}<={N4},({n}-{s4i})/{s4s},({n}-{s5i})/{s5s}))))')


def tax_due_formula(x_expr):
    X1, X2, X3, X4 = PA, BL, TS, HL
    x = f"({x_expr})"
    return (f'IF({x}<={X1},0,IF({x}<={X2},{BR}*({x}-{X1}),IF({x}<={X3},'
            f'{deriv["tax_x2"]}+{HR}*({x}-{X2}),IF({x}<={X4},{deriv["tax_x3"]}+0.6*({x}-{X3}),'
            f'{deriv["tax_x4"]}+{AR}*({x}-{X4})))))')


# gross_up_pot_withdrawal(net, other=SP) = MAX(gross_for_net(net) - other, 0) - NOT
# gross_for_net(net - other), since tax_due() is non-linear (piecewise), so the two are not
# interchangeable. This matches src/tax.py's gross_up_pot_withdrawal() exactly.
gross_after_formula = f'MAX({gross_for_net_formula(NR["spend"])}-{NR["sp_annual"]},0)'

ws.cell(row=illus_header, column=2, value=f"Age <SP age (before State Pension)").font = HEADER_FONT
ws.cell(row=illus_header, column=2).fill = HEADER_FILL
ws.cell(row=illus_header, column=3, value="Gross withdrawal needed from pot").font = HEADER_FONT
ws.cell(row=illus_header, column=3).fill = HEADER_FILL
r = illus_header + 1
ws.cell(row=r, column=2, value=f'="Desired net spend: £"&TEXT({NR["spend"]},"#,##0")')
ws.cell(row=r, column=3, value=f"={gross_for_net_formula(NR['spend'])}").number_format = GBP
r += 1
ws.cell(row=r, column=2, value='="From State Pension age "&'+NR["sp_age"]+'&":"')
ws.cell(row=r, column=3, value=f"={gross_after_formula}").number_format = GBP
r += 1
ws.cell(row=r, column=2, value='="State Pension received (also taxable): £"&TEXT('+NR["sp_annual"]+',"#,##0")')
r += 2
ws.cell(row=r, column=2, value=(
    "These figures are illustrative (today's money, no investment growth or guardrail adjustment) - "
    "they show the tax/State Pension mechanics in isolation, using the SAME formulas the Historical "
    "Projection and MC sheets apply every year to the actual (inflating, guardrail-adjusted) target."
)).font = SUBTITLE_FONT
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
ws.row_dimensions[r].height = 40

wb.save(OUT_PATH)
print("Saved stage 14 (Tax sheet + Inputs).")
