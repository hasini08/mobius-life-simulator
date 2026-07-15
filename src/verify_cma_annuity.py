"""Cross-checks the Excel workbook's CMA blend + annuitization figures against src/cma.py and
src/annuity.py, and against engine.historical_single_path, at several Inputs settings. Run AFTER
recalc.py has recalculated the workbook (reads cached formula VALUES via openpyxl data_only=True)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
import pandas as pd

import cma as cma_mod
import annuity as annuity_mod
from portfolios import AC, PORTFOLIOS
from engine import load_asset_returns, load_cpi, historical_single_path, ClientProfile

OUT_PATH = sys.argv[1] if len(sys.argv) > 1 else \
    "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"

wb = openpyxl.load_workbook(OUT_PATH, data_only=True)
ws_in = wb["Inputs"]
ws_cma = wb["CMA"]
ws_ann = wb["Annuity"]
NR = json.loads(Path("cellrefs.json").read_text())
cma_range = json.loads(Path("cma_range.json").read_text())
ann_range = json.loads(Path("annuity_range.json").read_text())


def cell(ref):
    sheet, addr = ref.split("!")
    return wb[sheet][addr.replace("$", "")].value


print("=" * 70)
print("Reading current Inputs toggle values from the (recalculated) workbook:")
age = cell(NR["age"])
pot = cell(NR["pot"])
cma_blend = cell(NR["cma_blend"])
annuity_pct = cell(NR["annuity_pct"])
annuity_joint_on = cell(NR["annuity_joint_on"])
print(f"  age={age}  pot={pot}  cma_blend={cma_blend}  annuity_pct={annuity_pct}  "
      f"annuity_joint_on={annuity_joint_on}")

print("\n" + "=" * 70)
print("CMA cross-check (per-asset-class monthly shift, Excel vs Python at blend=1):")
py_shifts = cma_mod.cma_shifts(load_asset_returns(), AC)
mismatches = 0
for i, ac in enumerate(cma_range["asset_classes"]):
    col = i + 2  # column B=2 is first asset class
    excel_shift = ws_cma.cell(row=cma_range["shift_row"], column=col).value
    py_shift = py_shifts.get(ac)
    if py_shift is None:
        continue
    diff = abs(excel_shift - py_shift)
    status = "OK" if diff < 1e-6 else "MISMATCH"
    if status == "MISMATCH":
        mismatches += 1
    print(f"  {ac:24s} excel={excel_shift:.6f}  python={py_shift:.6f}  diff={diff:.2e}  [{status}]")
print(f"CMA per-asset-class shift mismatches: {mismatches}")

print("\nPortfolio-level monthly shift (Excel vs SUMPRODUCT-by-hand in Python):")
weight_rows = json.loads(Path("weight_rows.json").read_text())
wb_vals = wb["Portfolios"]
for name in PORTFOLIOS:
    excel_cell = cma_range["portfolio_shift_cells"][name]
    excel_val = cell(excel_cell)
    wrow = weight_rows["rows"][name]
    py_val = 0.0
    for i, ac in enumerate(cma_range["asset_classes"]):
        w = wb_vals.cell(row=wrow, column=3 + i).value or 0.0
        py_val += w * py_shifts.get(ac, 0.0)
    diff = abs(excel_val - py_val)
    print(f"  {name:12s} excel={excel_val:.6f}  recomputed={py_val:.6f}  diff={diff:.2e}  "
          f"[{'OK' if diff < 1e-6 else 'MISMATCH'}]")

print("\n" + "=" * 70)
print("Annuity rate cross-check (Excel interpolation vs annuity.py, at Inputs!age):")
excel_single_rate = cell(ann_range["single_rate_cell"])
py_single_rate = annuity_mod.annuity_rate(int(age), joint=False)
print(f"  age={age}  excel single-life rate={excel_single_rate:.6f}  python={py_single_rate:.6f}  "
      f"diff={abs(excel_single_rate - py_single_rate):.2e}")
excel_used_rate = cell(ann_range["used_rate_cell"])
py_used_rate = annuity_mod.annuity_rate(int(age), joint=(annuity_joint_on == "Y"))
print(f"  used rate (joint={annuity_joint_on}): excel={excel_used_rate:.6f}  python={py_used_rate:.6f}  "
      f"diff={abs(excel_used_rate - py_used_rate):.2e}")
excel_income = cell(ann_range["income_cell"])
py_income = pot * annuity_pct * py_used_rate
print(f"  guaranteed income: excel=£{excel_income:,.2f}  python=£{py_income:,.2f}  "
      f"diff=£{abs(excel_income - py_income):,.2f}")

print("\n" + "=" * 70)
print("Historical Projection full reconciliation (Excel vs Python historical_single_path):")
asset_df = load_asset_returns()
cpi = load_cpi(asset_df)
blocks = json.loads(Path("historical_projection_blocks.json").read_text())
ws_hist = wb["Historical Projection"]

apply_tax_on = cell(NR["apply_tax_on"])
guardrails_on = cell(NR["guardrails_on"])
sp_annual = cell(NR["sp_annual"])
sp_age = cell(NR["sp_age"])
spend = cell(NR["spend"])
horizon = cell(NR["horizon"])
band = cell(NR["band"])
cut = cell(NR["cut"])
raise_ = cell(NR["raise"])

if cma_blend and cma_blend > 0:
    asset_df_used = cma_mod.apply_cma_blend(asset_df, AC, cma_blend)
else:
    asset_df_used = asset_df

profile = ClientProfile(
    starting_age=int(age), horizon_years=int(horizon), starting_pot=float(pot),
    initial_annual_spend=float(spend), guardrails=(guardrails_on == "Y"),
    guardrail_band=band, guardrail_cut=cut, guardrail_raise=raise_,
    apply_tax=(apply_tax_on == "Y"), state_pension_annual=float(sp_annual), state_pension_age=int(sp_age),
)
if annuity_pct and annuity_pct > 0:
    profile, used_rate, income = annuity_mod.annuitize(profile, annuity_pct, int(age),
                                                         joint=(annuity_joint_on == "Y"))

total_mismatches = 0
for name, header_row in blocks.items():
    py_df = historical_single_path(name, asset_df_used, cpi, profile, start_date="2000-01-01")
    first_data_row = header_row + 2
    n_years = len(py_df) - 1  # py_df includes a year-0 row
    mism = 0
    for y_i in range(1, n_years + 1):
        row = first_data_row + y_i - 1
        excel_pot = ws_hist.cell(row=row, column=9).value
        excel_net = ws_hist.cell(row=row, column=10).value
        py_pot = py_df["PortfolioValue"].iloc[y_i]
        py_net = py_df["Spend"].iloc[y_i]
        if excel_pot is None:
            continue
        pot_diff = abs(excel_pot - py_pot)
        net_diff = abs((excel_net or 0) - py_net)
        if pot_diff > 0.5 or net_diff > 0.5:
            mism += 1
            print(f"  [{name}] year {y_i}: pot excel={excel_pot:.2f} py={py_pot:.2f} diff={pot_diff:.2f} | "
                  f"net excel={excel_net} py={py_net:.2f} diff={net_diff:.2f}")
    total_mismatches += mism
    print(f"  {name}: {n_years} years checked, {mism} mismatches (>£0.50)")

print(f"\nTOTAL mismatches across all portfolios: {total_mismatches}")
print("=" * 70)
