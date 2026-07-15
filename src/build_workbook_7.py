"""Stage 7: Summary sheet - headline statistics (probability of ruin, legacy quantiles, shortfall
years) for all three portfolios, pulling from the Excel Monte Carlo sheets, with the guardrails
toggle applied throughout via Inputs."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
PCT = "0.0%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())

if "Summary" in wb.sheetnames:
    del wb["Summary"]
ws = wb.create_sheet("Summary", 0)  # place first
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 34
for col in "CDE":
    ws.column_dimensions[col].width = 18

ws["B2"] = "Mobius Wealth — Decumulation Simulator: Summary"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("Client: age "
            f"={NR['age']} today, £")
ws["B3"] = None
ws["B4"] = ("Key results for the client & guardrail settings on the Inputs tab. Change any yellow "
            "cell on Inputs and everything here recalculates.")
ws["B4"].font = SUBTITLE_FONT
ws["B4"].alignment = Alignment(wrap_text=True)

# Quick client readout
r = 6
for label, ref, fmt in [
    ("Starting pot", NR["pot"], GBP), ("Desired annual spend (today's £)", NR["spend"], GBP),
    ("Initial withdrawal rate", NR["wr0"], PCT), ("Time horizon (years)", NR["horizon"], "0"),
    ("Guardrails applied?", NR["guardrails_on"], None),
]:
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=f"={ref}")
    if fmt:
        cell.number_format = fmt
    cell.font = BOLD
    r += 1

r += 1
ws.cell(row=r, column=2, value="Monte Carlo results (Excel bootstrap engine)").font = BOLD
r += 1
mc_header_row = r
headers = ["Portfolio", "Probability of ruin", "Ruin prob SE", "Ruin prob 95% CI (low)",
           "Ruin prob 95% CI (high)", "Median legacy", "5th pctl legacy", "95th pctl legacy",
           "Avg shortfall years", "% paths with any shortfall"]
for i, h in enumerate(headers):
    cell = ws.cell(row=r, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(2 + i)].width = 17
r += 1
mc_data_first_row = r
for name in PORTFOLIOS:
    mc = json.loads(Path(f"mc_range_{name}.json").read_text())
    fr, lr = mc["first_row"], mc["last_row"]
    n_sims_here = mc["n_sims"]
    legacy_col_letter = get_column_letter(mc["legacy_col"])
    ruin_col_letter = get_column_letter(mc["ruin_col"])
    shortfall_col_letter = get_column_letter(mc["shortfall_col"])
    sheet = f"'MC {name}'"
    ws.cell(row=r, column=2, value=name).font = BOLD
    ruin_p_ref = f"C{r}"
    ws.cell(row=r, column=3, value=f"=AVERAGE({sheet}!{ruin_col_letter}{fr}:{ruin_col_letter}{lr})").number_format = PCT
    # binomial standard error on the ruin-probability estimate: sqrt(p(1-p)/n) - each simulated path
    # is an independent draw, so this is a valid description of the Monte Carlo sampling noise
    se_formula = f"=SQRT(MAX({ruin_p_ref}*(1-{ruin_p_ref}),0)/{n_sims_here})"
    ws.cell(row=r, column=4, value=se_formula).number_format = "0.00%"
    ws.cell(row=r, column=5, value=f"=MAX(0,{ruin_p_ref}-1.96*D{r})").number_format = PCT
    ws.cell(row=r, column=6, value=f"=MIN(1,{ruin_p_ref}+1.96*D{r})").number_format = PCT
    ws.cell(row=r, column=7, value=f"=MEDIAN({sheet}!{legacy_col_letter}{fr}:{legacy_col_letter}{lr})").number_format = GBP
    ws.cell(row=r, column=8, value=f"=PERCENTILE({sheet}!{legacy_col_letter}{fr}:{legacy_col_letter}{lr},0.05)").number_format = GBP
    ws.cell(row=r, column=9, value=f"=PERCENTILE({sheet}!{legacy_col_letter}{fr}:{legacy_col_letter}{lr},0.95)").number_format = GBP
    ws.cell(row=r, column=10, value=f"=AVERAGE({sheet}!{shortfall_col_letter}{fr}:{shortfall_col_letter}{lr})").number_format = "0.00"
    ws.cell(row=r, column=11, value=f"=COUNTIF({sheet}!{shortfall_col_letter}{fr}:{shortfall_col_letter}{lr},\">0\")/({lr}-{fr}+1)").number_format = PCT
    r += 1
mc_data_last_row = r - 1
ws.cell(row=r, column=2, value=f"(each based on {json.loads(Path('mc_range_Original.json').read_text())['n_sims']:,} simulated paths - see the SE/CI columns for sampling noise)").font = SUBTITLE_FONT

r += 1
ws.cell(row=r, column=2, value="Historical (deterministic, non-stochastic) result").font = BOLD
r += 1
hist_header_row = r
for i, h in enumerate(["Portfolio", "Final value (from Inputs!start_year)", "Any shortfall years?"]):
    cell = ws.cell(row=r, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
r += 1
hp_blocks = json.loads(Path("historical_projection_blocks.json").read_text())
MAX_HORIZON = 30
for name in PORTFOLIOS:
    header_row = hp_blocks[name]
    last_row = header_row + 1 + MAX_HORIZON  # header+1(year0)+MAX_HORIZON data rows
    ws.cell(row=r, column=2, value=name).font = BOLD
    ws.cell(row=r, column=3, value=f"='Historical Projection'!I{last_row}").number_format = GBP
    spend_col_range = f"'Historical Projection'!H{header_row+2}:H{last_row}"
    ws.cell(row=r, column=4, value=f'=IF(COUNTIF({spend_col_range},"<"&{NR["spend"]}*0.999)>0,"Yes","No")')
    r += 1

r += 2
ws.cell(row=r, column=2, value="Note: the Python/Streamlit app (see /app) runs the same comparison with "
                                 "an improved monthly stochastic engine (stationary block bootstrap / skew-t) "
                                 "and more simulation paths - see Instructions tab.").font = SUBTITLE_FONT
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

# Simple bar chart: probability of ruin by portfolio
chart = BarChart()
chart.title = "Probability of ruin by portfolio (Excel Monte Carlo)"
chart.y_axis.title = "Probability of ruin"
chart.x_axis.title = "Portfolio"
data = Reference(ws, min_col=3, min_row=mc_header_row, max_row=mc_data_last_row)
cats = Reference(ws, min_col=2, min_row=mc_data_first_row, max_row=mc_data_last_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 14
ws.add_chart(chart, f"B{mc_data_last_row + 3}")

wb.save(OUT_PATH)
print("Saved stage 7 (Summary).")
