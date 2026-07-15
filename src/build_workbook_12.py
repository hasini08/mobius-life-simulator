"""Stage 12: adds mortality columns to each MC sheet - 'Ruin year' (first year, 1-indexed, pot hits
~0), 'P(alive when ruined)' (the EXACT contribution to P(ruin before death), via S(ruin_year-1) off
the Mortality tab - exact given the market-path/mortality independence assumption, no extra sampling
noise), a RAND()-based 'Death year' draw (inverse-transform sampling against the same survival curve -
needed because legacy-at-death is path-dependent, so unlike 'ruin before death' it can't be reduced to
an expectation without an actual per-path death sample), and 'Legacy at death' (pot value at that
sampled death year, or final pot if the path outlives the horizon)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
mort = json.loads(Path("mortality_range.json").read_text())
s_eff_range = f"Mortality!{mort['s_eff_range']}"  # 31 rows: t=0..30, i.e. element #k = S(k-1)

for name in PORTFOLIOS:
    mc = json.loads(Path(f"mc_range_{name}.json").read_text())
    fr, lr = mc["first_row"], mc["last_row"]
    max_horizon = mc["max_horizon"]
    legacy_col, ruin_col, shortfall_col = mc["legacy_col"], mc["ruin_col"], mc["shortfall_col"]
    ws = wb[f"MC {name}"]

    # locate the year-block pot columns: same layout as build_workbook_6.py (5 cols/year, pot = +4)
    cols_per_year = mc.get("cols_per_year", 5)
    first_pot_col = 2 + (cols_per_year - 1)          # year 1's pot column
    pot_cols = [first_pot_col + cols_per_year * (y - 1) for y in range(1, max_horizon + 1)]
    final_pot_col_letter = get_column_letter(pot_cols[-1])
    first_pot_col_letter = get_column_letter(pot_cols[0])
    ruin_col_letter = get_column_letter(ruin_col)

    ruinyear_col = shortfall_col + 1
    palive_col = shortfall_col + 2
    u_col = shortfall_col + 3
    deathyear_col = shortfall_col + 4
    legacyatdeath_col = shortfall_col + 5

    headers = [
        (ruinyear_col, "Ruin year"), (palive_col, "P(alive when ruined)"),
        (u_col, "U (death draw)"), (deathyear_col, "Death year"),
        (legacyatdeath_col, "Legacy at death"),
    ]
    for c, h in headers:
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = 15

    for row in range(fr, lr + 1):
        ruinyear_letter = get_column_letter(ruinyear_col)
        u_letter = get_column_letter(u_col)
        deathyear_letter = get_column_letter(deathyear_col)

        # COUNTIF can't take a comma-joined list of non-contiguous cells as its range, so count
        # positive-pot years explicitly as a sum of boolean terms instead.
        positive_years_expr = "+".join(f"({get_column_letter(c)}{row}>0.01)" for c in pot_cols)
        ruinyear_formula = f'=IF({ruin_col_letter}{row}=1,{positive_years_expr}+1,"")'
        ws.cell(row=row, column=ruinyear_col, value=ruinyear_formula)

        palive_formula = (f'=IF({ruinyear_letter}{row}="",0,'
                           f'INDEX({s_eff_range},{ruinyear_letter}{row}))')
        ws.cell(row=row, column=palive_col, value=palive_formula).number_format = "0.0000"

        ws.cell(row=row, column=u_col, value="=RAND()").number_format = "0.0000"
        # inverse-transform: MATCH(u, S_eff, -1) needs S_eff descending (true by construction) and
        # returns the position of the LAST element >= u, i.e. element #k -> t=k-1 (0-indexed) is the
        # largest t with S(t)>=u -> death occurs in (1-indexed) year k = MATCH(...) itself. If that
        # position is the very last row (t=30=max_horizon), the path survives the full horizon -> blank.
        deathyear_formula = (f'=IF({u_letter}{row}<=INDEX({s_eff_range},{max_horizon+1}),"",'
                              f'MATCH({u_letter}{row},{s_eff_range},-1))')
        ws.cell(row=row, column=deathyear_col, value=deathyear_formula)

        # NOTE: uses INDEX, not OFFSET, deliberately. OFFSET is volatile (recalculates on every pass
        # regardless of dependencies), and at full scale (600 sims x 3 portfolios = 1,800 volatile
        # OFFSET cells, each dynamically re-resolving a cell reference every recalculation) this was
        # confirmed empirically to pathologically deadlock LibreOffice's headless multi-threaded
        # recalculation - hanging indefinitely (30+ min, zero CPU progress) at exactly 600 sim rows,
        # while 599 rows (one fewer volatile OFFSET cell per sheet) recalculated in seconds. Disabling
        # threaded calculation entirely also "fixed" it but made a single recalc pass crawl (still
        # progressing after 20+ minutes) - not viable for a one-shot delivery build. INDEX over the
        # same row's full year-block range achieves an identical dynamic column lookup but is NOT
        # volatile, which resolved the hang outright (confirmed: 600-row file recalculates cleanly in
        # under a minute with default threaded calculation, 0 errors).
        legacy_formula = (f'=IF({deathyear_letter}{row}="",{final_pot_col_letter}{row},'
                           f'INDEX({first_pot_col_letter}{row}:{final_pot_col_letter}{row},1,'
                           f'({deathyear_letter}{row}-1)*{cols_per_year}+1))')
        ws.cell(row=row, column=legacyatdeath_col, value=legacy_formula).number_format = GBP

    Path(f"mc_range_{name}.json").write_text(json.dumps({
        **mc, "ruinyear_col": ruinyear_col, "palive_col": palive_col, "u_col": u_col,
        "deathyear_col": deathyear_col, "legacyatdeath_col": legacyatdeath_col,
    }))
    print(f"Added mortality columns to MC {name}: ruinyear={ruinyear_col}, palive={palive_col}, "
          f"deathyear={deathyear_col}, legacyatdeath={legacyatdeath_col}")

wb.save(OUT_PATH)
print("Saved stage 12.")
