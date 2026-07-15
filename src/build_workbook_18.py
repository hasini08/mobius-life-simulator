"""Stage 18: applies partial annuitization (stage 17's Annuity sheet) to the Historical Projection
sheet - reduces the pot actually put into drawdown at outset, and nets the guaranteed annuity income
off the withdrawal need every year (real-terms equivalent, since the annuity is a LEVEL nominal
amount - see Annuity sheet), on top of whatever the tax/State Pension toggle is doing. Mirrors
src/engine.py's _gross_withdrawal_target() exactly: annuity income offsets the withdrawal need
REGARDLESS of the tax toggle; State Pension only enters via the tax gross-up (unchanged behaviour
when annuity_pct = 0, i.e. fully backward compatible).

NOTE: the MC sheets do NOT need patching here - build_workbook_6.py now builds them with annuity
awareness natively (reads annuity_range.json directly when generating the Target/Withdrawal/Pot
formulas), specifically to avoid re-deriving the tax/annuity gross-up expression 30 times inline
into one Shortfall cell per row (which caused a severe recalculation stall at full 600-sim scale -
see build_workbook_6.py's docstring). If you change the annuity mechanics, edit build_workbook_6.py
and re-run stages 6 -> 7 -> 12 -> 13 (rebuilds MC + Summary base + mortality columns), not this file."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
GBP = '£#,##0;(£#,##0);"-"'
FONT = "Arial"
GREEN = Font(name=FONT, color="008000")

wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
tax_deriv = json.loads(Path("tax_range.json").read_text())
ann = json.loads(Path("annuity_range.json").read_text())
INCOME = ann["income_cell"]
REMAINING_POT = ann["remaining_pot_cell"]


def gross_for_net_formula(n_expr):
    d = tax_deriv
    n = f"MAX({n_expr},0)"
    return (f'IF({n}<={d["n1"]},{n},IF({n}<={d["n2"]},({n}-{d["seg2_i"]})/{d["seg2_s"]},'
            f'IF({n}<={d["n3"]},({n}-{d["seg3_i"]})/{d["seg3_s"]},'
            f'IF({n}<={d["n4"]},({n}-{d["seg4_i"]})/{d["seg4_s"]},({n}-{d["seg5_i"]})/{d["seg5_s"]}))))')


def tax_due_formula(x_expr):
    d = tax_deriv
    x = f"({x_expr})"
    return (f'IF({x}<={d["pa"]},0,IF({x}<={d["basic_limit"]},{d["basic_rate"]}*({x}-{d["pa"]}),'
            f'IF({x}<={d["taper_start"]},{d["tax_x2"]}+{d["higher_rate"]}*({x}-{d["basic_limit"]}),'
            f'IF({x}<={d["higher_limit"]},{d["tax_x3"]}+0.6*({x}-{d["taper_start"]}),'
            f'{d["tax_x4"]}+{d["add_rate"]}*({x}-{d["higher_limit"]})))))')


def target_real_expr(realspend_expr, sp_real_expr, annuity_real_expr):
    """Real-terms (today's-money) withdrawal target still needed from the pot, mirroring
    engine._gross_withdrawal_target(): annuity income offsets the target REGARDLESS of the tax
    toggle; State Pension only enters via the tax gross-up branch."""
    gross_target = f'MAX({gross_for_net_formula(realspend_expr)}-({sp_real_expr})-({annuity_real_expr}),0)'
    no_tax_target = f'MAX(({realspend_expr})-({annuity_real_expr}),0)'
    return f'IF({NR["apply_tax_on"]}="Y",{gross_target},{no_tax_target})'


# =======================================================================
# Historical Projection
# =======================================================================
ws = wb["Historical Projection"]
blocks = json.loads(Path("historical_projection_blocks.json").read_text())
MAX_HORIZON = 30

for name, header_row in blocks.items():
    # year-0 row: pot starts at the REMAINING (post-annuitization) pot, not the full pot - a no-op
    # when annuity_pct = 0 since Annuity!remaining_pot_cell = Inputs!pot * (1-0) = Inputs!pot.
    year0_row = header_row + 1
    ws.cell(row=year0_row, column=9, value=f"={REMAINING_POT}").number_format = GBP

    first_data_row = header_row + 2
    for y_i in range(1, MAX_HORIZON + 1):
        row = first_data_row + y_i - 1
        prev = row - 1
        age_this_year = f"({NR['age']}+{y_i}-1)"
        sp_real = f'IF({age_this_year}>={NR["sp_age"]},{NR["sp_annual"]},0)'
        annuity_real = f'{INCOME}/F{row}'
        target_real = target_real_expr(f"G{row}", sp_real, annuity_real)
        spend_formula = f'=IF(D{row}="","",MIN(({target_real})*F{row},MAX(I{prev},0)))'
        ws.cell(row=row, column=8, value=spend_formula).number_format = GBP

        sp_nominal = f'({sp_real})*F{row}'
        total_gross_nominal = f'(H{row}+{sp_nominal}+{INCOME})'
        net_formula = (f'=IF(D{row}="","",IF({NR["apply_tax_on"]}="Y",'
                        f'{total_gross_nominal}-({tax_due_formula(total_gross_nominal)}),'
                        f'H{row}+{sp_nominal}+{INCOME}))')
        ws.cell(row=row, column=10, value=net_formula).number_format = GBP

print("Patched Historical Projection for annuitization.")

wb.save(OUT_PATH)
print("Saved stage 18 (annuitization applied to Historical Projection). "
      "MC sheets already handled natively by build_workbook_6.py.")
