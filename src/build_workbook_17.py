"""Stage 17: Partial annuitization - Excel-native mirror of src/annuity.py. Adds Inputs toggles
(% of pot to annuitize, joint life Y/N) and a new 'Annuity' sheet holding the real, dated UK
annuity-rate table (linear interpolation between quoted ages, clipped outside 55-75) and the
resulting guaranteed nominal annual income. Stage 18 patches the Historical Projection / MC sheets
to actually apply it (reduce the pot at outset, offset the withdrawal need)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import annuity as annuity_mod

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
PCT1 = "0.0%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())

# ---------------------------------------------------------------------
# Inputs: annuitization toggle block
# ---------------------------------------------------------------------
ws_in = wb["Inputs"]
existing_rows = [c.row for c in ws_in["B"] if c.value]
r0 = max(existing_rows) + 2
ws_in.cell(row=r0, column=2, value="Partial Annuitization").font = BOLD
r = r0 + 1
ws_in.cell(row=r, column=2, value="% of pot to annuitize at outset").font = BLACK
cell = ws_in.cell(row=r, column=3, value=0.0)
cell.font = BLUE
cell.fill = INPUT_FILL
cell.number_format = PCT1
ws_in.cell(row=r, column=4, value=(
    "0% = pure drawdown (unchanged). >0% moves that share of the STARTING pot into a guaranteed "
    "lifetime income (see the Annuity tab) at outset, leaving the rest in drawdown alongside it."
)).font = SUBTITLE_FONT
NR["annuity_pct"] = f"Inputs!$C${r}"
r += 1
ws_in.cell(row=r, column=2, value="Annuity: joint life (50% to survivor)? (Y/N)").font = BLACK
cell = ws_in.cell(row=r, column=3, value="N")
cell.font = BLUE
cell.fill = INPUT_FILL
ws_in.cell(row=r, column=4, value="Y = uses the lower joint-life rate (Annuity tab)").font = SUBTITLE_FONT
NR["annuity_joint_on"] = f"Inputs!$C${r}"
Path("cellrefs.json").write_text(json.dumps(NR, indent=2))
print("Added annuitization inputs:", {k: NR[k] for k in ["annuity_pct", "annuity_joint_on"]})

# ---------------------------------------------------------------------
# Annuity sheet
# ---------------------------------------------------------------------
if "Annuity" in wb.sheetnames:
    del wb["Annuity"]
ws = wb.create_sheet("Annuity")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 55

ws["B2"] = "Should part of the pot be swapped for a guaranteed income? (Partial Annuitization)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: an annuity means handing over part of the pot, ONCE, in exchange for an income "
    "paid for as long as the client lives, no matter how long that is or what happens to markets. "
    "It's a trade: give up growth potential and access to that money, in exchange for certainty.\n\n"
    "Converting part of the pot into a guaranteed LIFETIME income at outset. Rates below: single-"
    "life, LEVEL (does not rise with inflation), no-guarantee-period annuity, annual income per £1 "
    "purchased, scaled examples from Hargreaves Lansdown's published £100,000 best-buy annuity-rate "
    "table, 14 May 2026 (via pensionbible.co.uk), cross-checked directly against HL's own site, 28 "
    "May 2026 (within ~1% - normal week-to-week market movement). Joint-life (50% to survivor) uses "
    "a flat discount factor from HL's own age-65 joint-life quote (28 May 2026): £7,374 vs £7,970 "
    "per £100,000 = a 92.5% factor, applied at every age - a simplification, since no fuller joint-"
    "life age curve was available in the sources checked. Ages outside the quoted 55-75 range are "
    "CLIPPED to the nearest quoted age rather than extrapolated (no cited source covers them)."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 150

# --- rate table (5 anchor ages) ---
ws.cell(row=5, column=2, value="Single-life, level, no-guarantee rate table (editable)").font = BOLD
ages = sorted(annuity_mod.SINGLE_LIFE_RATE_TABLE)
r = 6
age_cells, rate_cells = {}, {}
ws.cell(row=r, column=2, value="Age").font = HEADER_FONT
ws.cell(row=r, column=2).fill = HEADER_FILL
ws.cell(row=r, column=3, value="Annual income per £1").font = HEADER_FONT
ws.cell(row=r, column=3).fill = HEADER_FILL
r += 1
for a in ages:
    ws.cell(row=r, column=2, value=a).font = BLACK
    cell = ws.cell(row=r, column=3, value=annuity_mod.SINGLE_LIFE_RATE_TABLE[a])
    cell.font = BLUE
    cell.fill = INPUT_FILL
    cell.number_format = PCT
    age_cells[a] = f"$B${r}"
    rate_cells[a] = f"$C${r}"
    r += 1
r += 1
ws.cell(row=r, column=2, value="Joint-life (50% survivor) / single-life factor").font = BLACK
cell = ws.cell(row=r, column=3, value=annuity_mod.JOINT_LIFE_50PCT_FACTOR)
cell.font = BLUE
cell.fill = INPUT_FILL
cell.number_format = PCT
joint_factor_cell = f"Annuity!$C${r}"
r += 2

# --- interpolated rate for the Inputs-tab client's own age ---
min_age, max_age = ages[0], ages[-1]
clipped_age_expr = f"MIN(MAX({NR['age']},{min_age}),{max_age})"
# piecewise-linear interpolation between consecutive anchor ages, nested IF (mirrors the Tax sheet's
# nested-IF style for consistency) - innermost segment first
interp_expr = f"C{6 + len(ages)}"  # placeholder never reached (age clipped to <= max_age)
for a0, a1 in reversed(list(zip(ages[:-1], ages[1:]))):
    r0c, r1c = rate_cells[a0], rate_cells[a1]
    seg = f"({r0c}+({r1c}-{r0c})*(({clipped_age_expr})-{a0})/({a1}-{a0}))"
    interp_expr = f'IF(({clipped_age_expr})<={a1},{seg},{interp_expr})' if a1 != max_age else seg
ws.cell(row=r, column=2, value="Interpolated single-life rate for Inputs!age").font = BOLD
single_rate_formula = f"={interp_expr}"
cell = ws.cell(row=r, column=3, value=single_rate_formula)
cell.number_format = PCT
cell.font = BOLD
single_rate_cell = f"Annuity!$C${r}"
r += 1

ws.cell(row=r, column=2, value="Rate actually used (single or joint, per Inputs toggle)").font = BOLD
formula = f'=IF({NR["annuity_joint_on"]}="Y",{single_rate_cell}*{joint_factor_cell},{single_rate_cell})'
cell = ws.cell(row=r, column=3, value=formula)
cell.number_format = PCT
cell.font = BOLD
used_rate_cell = f"Annuity!$C${r}"
r += 2

ws.cell(row=r, column=2, value="Guaranteed nominal annual income (LEVEL, does not inflate)").font = BOLD
formula = f'={NR["pot"]}*{NR["annuity_pct"]}*{used_rate_cell}'
cell = ws.cell(row=r, column=3, value=formula)
cell.number_format = GBP
cell.font = BOLD
income_cell = f"Annuity!$C${r}"
r += 1
ws.cell(row=r, column=2, value="Pot remaining for drawdown after annuitization").font = BLACK
formula = f'={NR["pot"]}*(1-{NR["annuity_pct"]})'
cell = ws.cell(row=r, column=3, value=formula)
cell.number_format = GBP
remaining_pot_cell = f"Annuity!$C${r}"
r += 2

ws.cell(row=r, column=2, value=(
    "This income is a fixed £ amount every year for life - it does NOT rise with inflation (unlike "
    "the State Pension, which is modelled here as growing with inflation), so its real purchasing "
    "power falls over time - roughly halving after ~20 years at ~3% inflation. It counts as taxable "
    "income (like State Pension) when Inputs!apply_tax_on = \"Y\", and directly reduces the withdrawal "
    "needed from the remaining pot either way. See the Historical Projection and MC sheets for how "
    "it's applied."
)).font = SUBTITLE_FONT
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
ws.row_dimensions[r].height = 55

annuity_range = {
    "single_rate_cell": single_rate_cell, "used_rate_cell": used_rate_cell,
    "income_cell": income_cell, "remaining_pot_cell": remaining_pot_cell,
    "joint_factor_cell": joint_factor_cell,
}
Path("annuity_range.json").write_text(json.dumps(annuity_range, indent=2))

# ---------------------------------------------------------------------
# Patch Inputs!wr0 (initial withdrawal rate - the guardrail baseline threshold every MC/Historical
# Projection sheet compares each year's withdrawal rate against) to divide by the REMAINING
# (post-annuitization) pot, not the full starting pot. Originally "=C8/C7" (spend/full pot), built
# before annuitization existed. Guardrails are about sustaining what's actually left in drawdown once
# annuitization has carved its share out at outset, so the baseline needs to shrink accordingly -
# mirrors src/engine.py exactly, where wr0 = profile.initial_annual_spend / profile.starting_pot and
# profile.starting_pot is ALREADY the annuitized (reduced) pot by the time any simulation runs (see
# annuity.annuitize(), which replaces starting_pot via dataclasses.replace). A no-op when
# annuity_pct=0, since remaining_pot_cell = Inputs!pot * (1-0) = Inputs!pot exactly - confirmed this
# was the source of a cross-engine mismatch (Excel vs Python guardrail-adjusted spend/pot diverging
# from the first guardrail-eligible year onward) caught by verify_cma_annuity.py once annuitization
# and guardrails were exercised together.
wr0_sheet, wr0_addr = NR["wr0"].split("!")
wb[wr0_sheet][wr0_addr.replace("$", "")] = f"={NR['spend']}/{remaining_pot_cell}"

wb.save(OUT_PATH)
print("Saved stage 17 (Annuity sheet + Inputs toggle). Patched Inputs!wr0 to use the remaining "
      "(post-annuitization) pot.")
print(annuity_range)
