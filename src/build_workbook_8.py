"""Stage 8: Equity Sweep sheet - a formula-driven (not Monte Carlo) exploration of how total equity
weight (Global + EM equities combined) affects each portfolio's historical risk/return profile.

For each of the three portfolios and each of 9 target equity weights (20%-100%), this sheet:
  1. Rescales the portfolio's asset-class weight vector to hit that target equity weight, preserving
     the relative split within the equity sleeve and within the rest of the portfolio (same method as
     the Python engine's scale_to_equity_weight / equity_sweep - see src/portfolios.py, src/engine.py).
  2. Derives that weight vector's calendar-year historical returns (net of the portfolio's own
     weighted-average fee - held fixed across the sweep, matching the Python engine).
  3. Computes CAGR, annualised volatility and max drawdown from that annual return series.
  4. Runs a deterministic year-by-year projection (same method as 'Historical Projection', guardrails
     included) starting Inputs!start_year, to show the final legacy value and whether a shortfall/ruin
     occurs on the actual historical path at that equity weight.

A full per-point Monte Carlo replication (27 grid points x thousands of paths) is impractical inside a
spreadsheet, so this sheet uses the historical/analytical route above; the Python/Streamlit app's
'Equity allocation sweep' section runs the full stochastic version (see Instructions)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
wr = json.loads(Path("weight_rows.json").read_text())
holding_ranges = json.loads(Path("holding_ranges.json").read_text())
ar = json.loads(Path("annual_returns_range.json").read_text())

ASSET_CLASSES = wr["asset_classes"]
EQUITY_CLASSES = {"Global Equities", "EM Equities"}
n_ac = len(ASSET_CLASSES)
years = ar["years"]
afr, alr = ar["first_row"], ar["last_row"]  # 'Annual Asset Returns' rows
cpi_col_letter = get_column_letter(ar["cpi_annual_col"])

# 'Annual Asset Returns' asset-class block starts at column B (index 2)
first_ac_col_letter = get_column_letter(2)
last_ac_col_letter = get_column_letter(1 + n_ac)
# 'Portfolios' weight roll-up block starts at column C (index 3)
first_w_col_letter = get_column_letter(3)
last_w_col_letter = get_column_letter(2 + n_ac)

GRID = [round(0.20 + 0.10 * i, 2) for i in range(9)]  # 20%, 30%, ..., 100%
MAX_HORIZON = 30

if "Equity Sweep" in wb.sheetnames:
    del wb["Equity Sweep"]
ws = wb.create_sheet("Equity Sweep")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 30
for i in range(len(GRID)):
    ws.column_dimensions[get_column_letter(3 + i)].width = 12

ws["B2"] = "How much should be in shares vs. safer assets? (Equity Allocation Sweep)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: re-tests each portfolio at different overall share exposures, from cautious to "
    "all-in, to show the trade-off between growth potential and smoothness directly. || Technical "
    "detail: rescales each portfolio's asset-class weights to hit a target TOTAL equity weight "
    "(Global + EM equities), keeping the relative split within the equity sleeve and within the rest "
    "of the portfolio fixed, then derives that mix's historical stats. Formula-driven, not Monte Carlo "
    "- see the Python app for the full stochastic version of this sweep."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 42

sweep_return_ranges = {}  # name -> {"first_row": .., "last_row": .., "cols": {grid_idx: col_letter}}

row = 5
for name in PORTFOLIOS:
    wrow = wr["rows"][name]
    tot_row = holding_ranges[name][2]  # row holding weighted-avg OCF, col F

    ws.cell(row=row, column=2, value=f"{name} portfolio").font = BOLD
    for c in range(2, 3 + len(GRID)):
        ws.cell(row=row, column=c).fill = SUBHEAD_FILL
    row += 1

    # Base equity / non-equity totals for this portfolio (from the Portfolios weight roll-up row).
    # NOTE: non-equity total is the SUM of the actual non-equity weight cells, not "1 - equity total" -
    # source holding weights don't always sum to exactly 100% (e.g. Alternative sums to 100.1% due to
    # rounding in the FNZ file), and the Python engine's scale_to_equity_weight() divides by the actual
    # non-equity sum too, so this must match it exactly rather than assuming a clean complement.
    equity_cols = [3 + i for i, ac in enumerate(ASSET_CLASSES) if ac in EQUITY_CLASSES]
    non_equity_cols = [3 + i for i, ac in enumerate(ASSET_CLASSES) if ac not in EQUITY_CLASSES]
    eq_col_letters = [get_column_letter(c) for c in equity_cols]
    noneq_col_letters = [get_column_letter(c) for c in non_equity_cols]
    ws.cell(row=row, column=2, value="Base total equity weight (Global + EM)").font = BLACK
    base_eq_formula = "=" + "+".join(f"Portfolios!${cl}${wrow}" for cl in eq_col_letters)
    ws.cell(row=row, column=3, value=base_eq_formula).number_format = PCT
    base_eq_row = row
    row += 1
    ws.cell(row=row, column=2, value="Base total non-equity weight").font = BLACK
    base_noneq_formula = "=" + "+".join(f"Portfolios!${cl}${wrow}" for cl in noneq_col_letters)
    ws.cell(row=row, column=3, value=base_noneq_formula).number_format = PCT
    base_noneq_row = row
    row += 1

    # ---- Target equity weight header row ----
    ws.cell(row=row, column=2, value="Target total equity weight ->").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    grid_header_row = row
    for i, g in enumerate(GRID):
        c = 3 + i
        cell = ws.cell(row=row, column=c, value=g)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.number_format = "0%"
    row += 1

    # ---- Scaled weight vectors: one row per asset class ----
    weight_block_first = row
    for j, ac in enumerate(ASSET_CLASSES):
        ws.cell(row=row, column=2, value=ac).font = BLACK
        src_col_letter = get_column_letter(3 + j)  # matches Portfolios weight roll-up column order
        for i, g in enumerate(GRID):
            c = 3 + i
            grid_letter = get_column_letter(c)
            if ac in EQUITY_CLASSES:
                formula = (f"=Portfolios!${src_col_letter}${wrow}*"
                           f"({grid_letter}${grid_header_row}/$C${base_eq_row})")
            else:
                formula = (f"=Portfolios!${src_col_letter}${wrow}*"
                           f"((1-{grid_letter}${grid_header_row})/$C${base_noneq_row})")
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = PCT
            cell.font = GREEN
        row += 1
    weight_block_last = row - 1
    row += 1

    # ---- Annual returns for each swept weight vector (fee held fixed at the portfolio's own OCF) ----
    ws.cell(row=row, column=2, value="Calendar-year return at swept weights (net of portfolio's own fee)").font = BOLD
    row += 1
    ret_header_row = row
    ws.cell(row=row, column=2, value="Year").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    for i in range(len(GRID)):
        c = 3 + i
        cell = ws.cell(row=row, column=c, value=f"={get_column_letter(c)}{grid_header_row}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.number_format = "0%"
    row += 1
    ret_first_row = row
    for k, y in enumerate(years):
        yr_row = afr + k  # matching row in 'Annual Asset Returns'
        ws.cell(row=row, column=2, value=y).font = BOLD
        for i in range(len(GRID)):
            c = 3 + i
            w_col_letter = get_column_letter(c)
            # NOTE: the weight vector here is laid out VERTICALLY (one row per asset class, in
            # weight_block_first..weight_block_last) while 'Annual Asset Returns' asset-class returns
            # for a given year are laid out HORIZONTALLY (one column per asset class). SUMPRODUCT
            # requires matching array shapes, so this is built as an explicit sum of the 11
            # weight x return products (safer than TRANSPOSE(), which needs array/CSE entry and is
            # not reliably supported across Excel/LibreOffice).
            terms = []
            for j, ac2 in enumerate(ASSET_CLASSES):
                ac_col_letter = get_column_letter(2 + j)
                w_row = weight_block_first + j
                terms.append(f"${w_col_letter}${w_row}*'Annual Asset Returns'!${ac_col_letter}${yr_row}")
            formula = "=" + "+".join(terms) + f"-Portfolios!$F${tot_row}"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = PCT
            cell.font = BLACK
        row += 1
    ret_last_row = row - 1
    sweep_return_ranges[name] = {
        "first_row": ret_first_row, "last_row": ret_last_row,
        "cols": {i: get_column_letter(3 + i) for i in range(len(GRID))},
    }
    row += 1

    # ---- Wealth index / running max / drawdown helper blocks (pure market path, no withdrawals) ----
    ws.cell(row=row, column=2, value="Wealth index (£1 invested, no withdrawals) - for max drawdown").font = SUBTITLE_FONT
    row += 1
    wealth_first = row
    for k, y in enumerate(years):
        ws.cell(row=row, column=2, value=y).font = BLACK
        for i in range(len(GRID)):
            c = 3 + i
            col_letter = get_column_letter(c)
            prior = "1" if k == 0 else f"{col_letter}{row-1}"
            formula = f"={prior}*(1+{col_letter}{ret_first_row+k})"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = "0.0000"
        row += 1
    wealth_last = row - 1
    row += 1

    ws.cell(row=row, column=2, value="Running max of wealth index").font = SUBTITLE_FONT
    row += 1
    runmax_first = row
    for k, y in enumerate(years):
        ws.cell(row=row, column=2, value=y).font = BLACK
        for i in range(len(GRID)):
            c = 3 + i
            col_letter = get_column_letter(c)
            formula = f"=MAX(${col_letter}${wealth_first}:{col_letter}{wealth_first+k})"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = "0.0000"
        row += 1
    runmax_last = row - 1
    row += 1

    ws.cell(row=row, column=2, value="Drawdown from peak").font = SUBTITLE_FONT
    row += 1
    dd_first = row
    for k, y in enumerate(years):
        ws.cell(row=row, column=2, value=y).font = BLACK
        for i in range(len(GRID)):
            c = 3 + i
            col_letter = get_column_letter(c)
            wrow_ = wealth_first + k
            rrow_ = runmax_first + k
            formula = f"=({col_letter}{wrow_}-{col_letter}{rrow_})/{col_letter}{rrow_}"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = PCT
        row += 1
    dd_last = row - 1
    row += 2

    # ---- Summary stats table: one row per grid point ----
    ws.cell(row=row, column=2, value="Summary by target equity weight").font = BOLD
    row += 1
    summary_header_row = row
    for i, h in enumerate(["Target equity weight", "CAGR (annualised)", "Annualised volatility",
                            "Max drawdown (market only)", "Historical final legacy", "Any shortfall on hist. path?"]):
        cell = ws.cell(row=row, column=2 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(2 + i)].width = max(ws.column_dimensions[get_column_letter(2 + i)].width or 10, 17)
    row += 1
    summary_first_row = row
    n_years = len(years)
    hist_final_refs = {}  # grid_idx -> (final_pot_cell, any_shortfall_cell) filled in after projection blocks built
    for i, g in enumerate(GRID):
        col_letter = get_column_letter(3 + i)
        ws.cell(row=row, column=2, value=f"={col_letter}{grid_header_row}").number_format = "0%"
        cagr_formula = f"=EXP(SUMPRODUCT(LN(1+{col_letter}{ret_first_row}:{col_letter}{ret_last_row})))^(1/{n_years})-1"
        ws.cell(row=row, column=3, value=cagr_formula).number_format = PCT
        vol_formula = f"=STDEV({col_letter}{ret_first_row}:{col_letter}{ret_last_row})"
        ws.cell(row=row, column=4, value=vol_formula).number_format = PCT
        dd_formula = f"=MIN({col_letter}{dd_first}:{col_letter}{dd_last})"
        ws.cell(row=row, column=5, value=dd_formula).number_format = PCT
        hist_final_refs[i] = row  # placeholder; columns 6 & 7 filled after the projection blocks below
        row += 1
    summary_last_row = row - 1
    row += 2

    # ---- Deterministic historical projection per grid point (mirrors 'Historical Projection') ----
    ws.cell(row=row, column=2, value="Historical single-path projection at each swept equity weight "
                                       "(same method as 'Historical Projection'; guardrails apply if "
                                       "Inputs!guardrails_on = \"Y\")").font = BOLD
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 2

    for i, g in enumerate(GRID):
        ret_col_letter = get_column_letter(3 + i)
        ws.cell(row=row, column=2, value=f"Equity weight = {g:.0%}").font = BOLD
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = SUBHEAD_FILL
        row += 1
        headers = ["Year #", "Calendar year", "Return used", "Inflation used", "Cumulative inflation index",
                   "Real spend level", "Spend taken", "End-of-year value"]
        for hi, h in enumerate(headers):
            ws.cell(row=row, column=2 + hi, value=h).font = HEADER_FONT
            ws.cell(row=row, column=2 + hi).fill = HEADER_FILL
        header_row = row
        row += 1
        c_year_num, c_cal_year, c_ret, c_infl, c_cuminfl, c_realspend, c_spend, c_pot = range(2, 10)

        ws.cell(row=row, column=c_year_num, value=0)
        ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}-1")
        ws.cell(row=row, column=c_cuminfl, value=1.0).number_format = "0.0000"
        ws.cell(row=row, column=c_realspend, value=f"={NR['spend']}").number_format = GBP
        ws.cell(row=row, column=c_pot, value=f"={NR['pot']}").number_format = GBP
        row += 1

        for y_i in range(1, MAX_HORIZON + 1):
            prev = row - 1
            ws.cell(row=row, column=c_year_num, value=y_i)
            ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}+{y_i}-1")
            avail_check = (f"AND(B{row}<={NR['horizon']},"
                            f"COUNTIF($B${ret_first_row}:$B${ret_last_row},C{row})>0)")
            ret_formula = (f'=IF({avail_check},'
                            f"INDEX(${ret_col_letter}${ret_first_row}:${ret_col_letter}${ret_last_row},"
                            f"MATCH(C{row},$B${ret_first_row}:$B${ret_last_row},0)),\"\")")
            ws.cell(row=row, column=c_ret, value=ret_formula).number_format = PCT
            infl_formula = (f'=IF(D{row}="","",IF({avail_check},'
                             f"INDEX('Annual Asset Returns'!${cpi_col_letter}${afr}:${cpi_col_letter}${alr},"
                             f"MATCH(C{row},'Annual Asset Returns'!$A${afr}:$A${alr},0)),\"\"))")
            ws.cell(row=row, column=c_infl, value=infl_formula).number_format = PCT

            cuminfl_formula = f'=IF(D{row}="",F{prev},F{prev}*(1+E{row}))'
            ws.cell(row=row, column=c_cuminfl, value=cuminfl_formula).number_format = "0.0000"

            if y_i == 1:
                real_spend_formula = f"={NR['spend']}"
            else:
                wr0 = NR["wr0"]
                prior_realspend = f"G{prev}"
                target_nominal = f"({prior_realspend}*F{row})"
                wr_now = f"IF(I{prev}>0,{target_nominal}/I{prev},9^9)"
                raw_adj = (
                    f'IF({wr_now}>{wr0}*(1+{NR["band"]}),{prior_realspend}*(1-{NR["cut"]}),'
                    f'IF({wr_now}<{wr0}*(1-{NR["band"]}),{prior_realspend}*(1+{NR["raise"]}),{prior_realspend}))'
                )
                real_spend_formula = (
                    f'=IF(D{row}="",{prior_realspend},'
                    f'IF({NR["guardrails_on"]}<>"Y",{prior_realspend},'
                    f'MIN(MAX({raw_adj},0.5*{NR["spend"]}),1.5*{NR["spend"]})))'
                )
            ws.cell(row=row, column=c_realspend, value=real_spend_formula).number_format = GBP

            spend_formula = f'=IF(D{row}="","",MIN(G{row}*F{row},MAX(I{prev},0)))'
            ws.cell(row=row, column=c_spend, value=spend_formula).number_format = GBP

            pot_formula = f'=IF(D{row}="",I{prev},MAX(I{prev}-H{row},0)*(1+D{row}))'
            ws.cell(row=row, column=c_pot, value=pot_formula).number_format = GBP
            row += 1
        last_data_row = row - 1
        first_data_row = header_row + 2  # first actual year-1 row (skips header + year-0 row)

        # wire this block's final value / shortfall flag back into the summary table
        srow = hist_final_refs[i]
        ws.cell(row=srow, column=6, value=f"=I{last_data_row}").number_format = GBP
        spend_range = f"H{first_data_row}:H{last_data_row}"
        ws.cell(row=srow, column=7,
                value=f'=IF(COUNTIF({spend_range},"<"&{NR["spend"]}*0.999)>0,"Yes","No")')
        row += 2

wb.save(OUT_PATH)
Path("equity_sweep_ranges.json").write_text(json.dumps(sweep_return_ranges))
print("Saved stage 8 (Equity Sweep). Grid:", GRID)
