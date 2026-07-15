"""Stage 4: Portfolio Returns (MONTHLY, weighted asset returns net of fee - matches the Python
engine's method exactly), then Portfolio Annual Returns compounds those monthly returns per
calendar year (so Excel and the Python app agree numerically)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
wr = json.loads(Path("weight_rows.json").read_text())
ar = json.loads(Path("asset_returns_range.json").read_text())
ASSET_CLASSES = ar["asset_classes"]
afr, alr = ar["first_row"], ar["last_row"]
holding_ranges = json.loads(Path("holding_ranges.json").read_text())

n_ac = len(ASSET_CLASSES)
# 'Asset Class Returns' / 'Annual Asset Returns' asset-class block starts at column B (index 2).
first_ac_col_letter = get_column_letter(2)
last_ac_col_letter = get_column_letter(1 + n_ac)
# 'Portfolios' asset-class weight roll-up block starts at column C (index 3) - one column over,
# because column B there holds the portfolio name label.
first_w_col_letter = get_column_letter(3)
last_w_col_letter = get_column_letter(2 + n_ac)

# ---------------------------------------------------------------------
# Portfolio Returns (MONTHLY)
# ---------------------------------------------------------------------
ws = wb.create_sheet("Portfolio Returns")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B3"
ws["A1"] = ("Monthly net-of-fee portfolio return = SUMPRODUCT(monthly asset-class returns, portfolio "
            "weight vector) - (weighted-avg annual fee)/12. This matches the Python engine's method "
            "exactly (monthly compounding, not annual weighted-average) so Excel and the app agree.")
ws["A1"].font = SUBTITLE_FONT
ws.cell(row=2, column=1, value="Month end").font = HEADER_FONT
ws.cell(row=2, column=1).fill = HEADER_FILL
ws.column_dimensions["A"].width = 12
port_cols = {}
c = 2
for name in PORTFOLIOS:
    cell = ws.cell(row=2, column=c, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(c)].width = 14
    port_cols[name] = c
    c += 1

n_months = alr - afr + 1
for i in range(n_months):
    row_n = 3 + i
    src_row = afr + i
    ws.cell(row=row_n, column=1, value=f"='Asset Class Returns'!A{src_row}").number_format = "mmm-yyyy"
    for name in PORTFOLIOS:
        c = port_cols[name]
        wrow = wr["rows"][name]
        tot_row = holding_ranges[name][2]
        formula = (f"=SUMPRODUCT('Asset Class Returns'!${first_ac_col_letter}${src_row}:${last_ac_col_letter}${src_row},"
                   f"Portfolios!${first_w_col_letter}${wrow}:${last_w_col_letter}${wrow})"
                   f"-Portfolios!$F${tot_row}/12")
        cell = ws.cell(row=row_n, column=c, value=formula)
        cell.number_format = PCT
        cell.font = GREEN

pr_first_row, pr_last_row = 3, 3 + n_months - 1
Path("portfolio_returns_range.json").write_text(json.dumps({
    "first_row": pr_first_row, "last_row": pr_last_row, "port_cols": port_cols, "n_months": n_months
}))
print(f"Portfolio Returns (monthly): rows {pr_first_row}..{pr_last_row}")

# ---------------------------------------------------------------------
# Portfolio Annual Returns - compounds the MONTHLY portfolio returns above per calendar year
# ---------------------------------------------------------------------
import pandas as pd
ac_raw = pd.read_csv("/home/claude/mobius_decumulation/data/asset_class_returns.csv", index_col=0, parse_dates=True)
cpi_col_name = [c_ for c_ in ac_raw.columns if "CPI" in c_][0]
dates = ac_raw.index
cpi_series = ac_raw[cpi_col_name]

year_month_rows = {}  # year -> (first_pr_row, last_pr_row) within Portfolio Returns sheet
for i, dt in enumerate(dates):
    y = dt.year
    pr_row = pr_first_row + i
    year_month_rows.setdefault(y, [pr_row, pr_row])
    year_month_rows[y][1] = pr_row
years = sorted(y for y in year_month_rows if y >= 2000)

cpi_last_row_by_year = {}
for i, dt in enumerate(dates):
    if pd.notna(cpi_series.iloc[i]):
        cpi_last_row_by_year[dt.year] = afr + i  # row in 'Asset Class Returns' (has CPI column added in stage 3)

annr_prev = json.loads(Path("annual_returns_range.json").read_text())
cpi_col_excel = annr_prev["cpi_annual_col"] - (2 + n_ac) + (2 + n_ac)  # same column index used in Asset Class Returns
# CPI column in 'Asset Class Returns' is at index 2+n_ac (0-based col count) -> compute directly:
cpi_col_in_acreturns = 2 + n_ac

ws2 = wb.create_sheet("Portfolio Annual Returns")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B3"
ws2["A1"] = "Calendar-year compounded portfolio returns (from the monthly 'Portfolio Returns' sheet) and that year's inflation (UK CPI YoY, latest available month)."
ws2["A1"].font = SUBTITLE_FONT
ws2.cell(row=2, column=1, value="Year").font = HEADER_FONT
ws2.cell(row=2, column=1).fill = HEADER_FILL
ws2.column_dimensions["A"].width = 10
pcols2 = {}
c = 2
for name in PORTFOLIOS:
    cell = ws2.cell(row=2, column=c, value=name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws2.column_dimensions[get_column_letter(c)].width = 14
    pcols2[name] = c
    c += 1
infl_col = c
cell = ws2.cell(row=2, column=infl_col, value="Inflation (UK CPI YoY)")
cell.font = HEADER_FONT
cell.fill = HEADER_FILL
ws2.column_dimensions[get_column_letter(infl_col)].width = 18

for i, y in enumerate(years):
    row_n = 3 + i
    ws2.cell(row=row_n, column=1, value=y).font = BOLD
    fr, lr = year_month_rows[y]
    for name in PORTFOLIOS:
        c = pcols2[name]
        col_letter = get_column_letter(port_cols[name])
        formula = f"=EXP(SUMPRODUCT(LN(1+'Portfolio Returns'!{col_letter}{fr}:{col_letter}{lr})))-1"
        cell = ws2.cell(row=row_n, column=c, value=formula)
        cell.number_format = PCT
        cell.font = GREEN
    cpi_row = cpi_last_row_by_year.get(y, lr)
    cpi_col_letter = get_column_letter(cpi_col_in_acreturns)
    formula = f"='Asset Class Returns'!{cpi_col_letter}{cpi_row}"
    cell = ws2.cell(row=row_n, column=infl_col, value=formula)
    cell.number_format = PCT
    cell.font = GREEN

par_first_row, par_last_row = 3, 3 + len(years) - 1
Path("portfolio_annual_returns_range.json").write_text(json.dumps({
    "first_row": par_first_row, "last_row": par_last_row, "years": years,
    "port_cols": pcols2, "infl_col": infl_col
}))

wb.save(OUT_PATH)
print("Saved stage 4 (Portfolio Returns monthly + Portfolio Annual Returns).")
