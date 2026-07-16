"""
Standalone summary sheet for a NEW candidate "Mobius Better v4" construction (weights supplied
15 July 2026), compared against Aspen Four Seasons - in the same style as the app's PDF export,
extended with downside-risk stats (Max DD, Average DD, CVaR). NO changes made to app.py / engine.py
/ portfolios.py - this is fully standalone, per instruction.

BERENBERG / "PROTECTED EQUITIES" (20%) = the previous model's "Eq Gbl DM Novum Mgd Vol" series
(a managed-volatility equity strategy) - confirmed as the correct data source for this holding
(not a placeholder).

DATA SOURCE - read before using: all 11 holdings' asset-class return series come from the PREVIOUS
Mobius model workbook's own 'Asset Returns' tab (Copy of Mobius decumulation model v2 (004).xlsm) -
a DIFFERENT, OLDER data source than the current app's Bloomberg 14 July 2026 pull. The common window
where ALL 11 series have data simultaneously is 2001-01 to 2025-01 (~24 years, shorter than this
project's usual 1999/2000-2026 window - the Novum Mgd Vol series only starts 2001-01). For a fair
comparison, Aspen Four Seasons is ALSO evaluated over this same window in this file's own tables
(not its full window through 2026), which trims away its most recent ~18 months - noted explicitly
in the sheet.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from fpdf.enums import XPos, YPos

from engine import load_asset_returns, load_cpi, weighted_monthly_returns, _draw_stationary_block
from portfolios import asset_class_weights, weighted_avg_fee
from build_recent_window_summary import compound_ret, vol, cvar, rolling_12m, max_drawdown, compute_irr

V2_MODEL_PATH = r"C:\Users\YahampathH\OneDrive - Mobius Life Limited\Documents\Copy of Mobius decumulation model v2 (004).xlsm"
OUT = "../output/Mobius_Wealth_Better_v4_vs_FourSeasons_Summary.xlsx"
OUT_PDF = "../output/Mobius_Wealth_Better_v4_vs_FourSeasons_Summary.pdf"

# name in old model -> (column index in its 'Asset Returns' sheet, requested weight)
OLD_MODEL_COLS = {
    "Global Agg Bonds":                              (25, 0.325),
    "Eq Gbl DM Quality Gross":                        (8,  0.125),
    "Berenberg / Protected Equities (Eq Gbl DM Novum Mgd Vol)": (14, 0.200),
    "Eq EM Net":                                      (6,  0.025),
    "US HY Corp Bond":                                (22, 0.025),
    "US ABS":                                         (21, 0.050),
    "EM Corp Bond":                                   (23, 0.025),
    "US Prop REITS":                                  (19, 0.025),
    "Commod":                                         (18, 0.050),
    "Hedge Fund Credit Suisse":                       (15, 0.075),
    "HF Trend":                                       (16, 0.075),
}

FLAT_FEE = 0.0007  # same flat 7bps convention used for Mobius Alternative/Better elsewhere this session

BLUE = "E9F2FB"
WARN = "FCE8E6"
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="C9C9C9")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def blue_fill():
    return PatternFill("solid", fgColor=BLUE)


def warn_fill():
    return PatternFill("solid", fgColor=WARN)


def extract_old_model_returns() -> pd.DataFrame:
    """Pulls the 10 needed asset-class monthly return series from the previous model's 'Asset
    Returns' tab, aligned on the common window where all 10 have real (numeric) data."""
    import openpyxl
    import numbers
    import os
    try:
        wb = openpyxl.load_workbook(V2_MODEL_PATH, data_only=True, keep_vba=True)
    except PermissionError:
        # Source file is open in Excel and even a Win32-level copy is denied while it's locked -
        # fall back to a pre-made scratch copy if one exists (made via a POSIX-style `cp`, which
        # Windows permits even while Excel holds the file open).
        fallback = os.environ.get("V2_MODEL_FALLBACK_COPY")
        if not fallback or not os.path.exists(fallback):
            raise SystemExit(
                f"Cannot open {V2_MODEL_PATH} - it's open in Excel, and no fallback copy was set "
                "via the V2_MODEL_FALLBACK_COPY environment variable. Please close it and re-run."
            )
        wb = openpyxl.load_workbook(fallback, data_only=True, keep_vba=True)
    ws = wb["Asset Returns"]
    dates = [ws.cell(row=r, column=3).value for r in range(9, ws.max_row + 1)]
    series = {}
    for label, (col, _w) in OLD_MODEL_COLS.items():
        vals = [ws.cell(row=r, column=col).value for r in range(9, ws.max_row + 1)]
        pairs = {d: v for d, v in zip(dates, vals) if d is not None and isinstance(v, numbers.Number)}
        series[label] = pd.Series(pairs)
    df = pd.DataFrame(series)
    df.index = pd.DatetimeIndex(df.index).to_period("M").to_timestamp("M")
    df = df.sort_index().dropna(how="any")  # keep only months where ALL 10 series have real data
    return df


def build_better_v4_weights() -> pd.Series:
    """Rescales the 10 available holdings' originally-requested weights to fill 100%, excluding
    the unavailable Berenberg slot."""
    raw = pd.Series({label: w for label, (_col, w) in OLD_MODEL_COLS.items()})
    scale = 1.0 / raw.sum()  # raw sums to 0.80 (1.0 - 0.20 Berenberg)
    return raw * scale


def historical_walk(monthly_ret: pd.Series, cpi: pd.Series, starting_pot, spend, horizon_years,
                     starting_age=65):
    """Reimplements engine.py's historical_single_path walk (guardrails off, tax off, no
    mortality/annuity - a clean baseline) directly on an arbitrary monthly-return series, so it
    works without a name registered in portfolios.py. Matches that function's logic exactly."""
    idx = monthly_ret.index[: horizon_years * 12]
    pot = starting_pot
    cum_inflation = 1.0
    rows = [(idx[0] if len(idx) else None, pot, 0.0)]
    for y in range(horizon_years):
        yr_idx = idx[y * 12:(y + 1) * 12]
        if len(yr_idx) == 0:
            break
        growth = float((1 + monthly_ret.loc[yr_idx]).prod())
        cpi_upto = cpi.loc[:yr_idx[-1]]
        infl = float(cpi_upto.iloc[-1]) if len(cpi_upto) else 0.0
        cum_inflation *= (1 + infl)
        nominal_spend_target = spend * cum_inflation
        withdrawal = min(nominal_spend_target, max(pot, 0))
        pot = max(pot - withdrawal, 0) * growth
        rows.append((yr_idx[-1], pot, withdrawal))
    return pd.DataFrame(rows, columns=["Date", "PortfolioValue", "Spend"])


def monte_carlo_prob_ruin(monthly_ret: pd.Series, cpi: pd.Series, starting_pot, spend, horizon_years,
                           n_sims=2000, seed=42, block_mean=12):
    """Reimplements engine.py's run_simulation post-processing loop (guardrails/tax/mortality all
    off - matching historical_walk above) on top of the reused _draw_stationary_block sampler, so
    it works for an arbitrary return series with no portfolio-name registration needed."""
    rng = np.random.default_rng(seed)
    common_idx = monthly_ret.index.intersection(cpi.index)
    r_series = monthly_ret.loc[common_idx].values
    cpi_vals = cpi.loc[common_idx].values
    n_months = horizon_years * 12
    r, c = _draw_stationary_block(r_series, cpi_vals, n_months, n_sims, rng, block_mean=block_mean)

    pot = np.full(n_sims, float(starting_pot))
    cum_inflation = np.ones(n_sims)
    ruin_year = np.full(n_sims, -1)
    legacy = np.empty(n_sims)
    for y in range(horizon_years):
        m0, m1 = y * 12, (y + 1) * 12
        year_growth = np.prod(1 + r[:, m0:m1], axis=1)
        year_infl = c[:, m1 - 1]
        cum_inflation *= (1 + year_infl)
        nominal_spend_target = spend * cum_inflation
        actual_spend = np.minimum(nominal_spend_target, np.maximum(pot, 0))
        pot = np.maximum(pot - actual_spend, 0) * year_growth
        pot = np.maximum(pot, 0)
        newly_ruined = (ruin_year < 0) & (pot <= 0) & (actual_spend < nominal_spend_target - 1e-6)
        ruin_year = np.where(newly_ruined, y, ruin_year)
    legacy[:] = pot
    prob_ruin = float((ruin_year >= 0).mean())
    return prob_ruin, float(np.median(legacy))


def full_stats(monthly_ret: pd.Series, cpi: pd.Series, fee: float, horizon_years: int,
               starting_pot=500_000.0, spend=20_000.0) -> dict:
    m = monthly_ret.dropna()
    hist_df = historical_walk(m, cpi, starting_pot, spend, horizon_years)
    values = hist_df["PortfolioValue"].to_numpy()
    spends = hist_df["Spend"].to_numpy()
    shortfall_years = int(sum(1 for i in range(1, len(spends)) if (spend - spends[i]) > 1e-6))
    ruin_hist = "Y" if values.min() <= 1 else "N"
    legacy_hist = float(values[-1])
    irr = compute_irr(hist_df)
    dd_series_avg = float(np.mean(_running_dd(m.to_numpy())))
    prob_ruin, mc_legacy = monte_carlo_prob_ruin(m, cpi, starting_pot, spend, horizon_years)
    cumulative_pct = (legacy_hist / starting_pot - 1) * 100
    return dict(
        n_months=len(m), start=m.index.min(), end=m.index.max(),
        compound=compound_ret(m), vol=vol(m), cvar_m=cvar(m), cvar_a=cvar(rolling_12m(m)),
        maxdd=max_drawdown(m), avgdd=dd_series_avg, irr=irr, ruin_hist=ruin_hist,
        legacy_hist=legacy_hist, cumulative_pct=cumulative_pct, shortfall=shortfall_years,
        prob_ruin=prob_ruin, mc_median_legacy=mc_legacy, fee=fee,
    )


def _running_dd(monthly: np.ndarray) -> np.ndarray:
    dd = 0.0
    out = np.empty(len(monthly))
    for i, r in enumerate(monthly):
        dd = min(0.0, (1 + dd) * (1 + r) - 1)
        out[i] = dd
    return out


def write_box(ws, top_row, title, names, stats_by_name, colors):
    ws.cell(row=top_row, column=2, value=title).font = TITLE_FONT
    header_row = top_row + 2
    ws.cell(row=header_row - 1, column=3, value="Portfolio Name  >>").font = Font(italic=True, color="808080")
    for k, name in enumerate(names):
        c = ws.cell(row=header_row - 1, column=4 + k, value=name)
        c.font = HEADER_FONT
        c.fill = blue_fill()
        c.border = BOX_BORDER
        c.alignment = Alignment(horizontal="center")

    rows = [
        ("Portfolio Stats", None, None),
        ("Data used (months / window)", "window", None),
        ("Compound ret pa", "compound", "0.00%"),
        ("Volatility pa", "vol", "0.00%"),
        ("Downside stats", None, None),
        ("Max DD (single worst drawdown)", "maxdd", "0.00%"),
        ("Average DD (mean of drawdown series)", "avgdd", "0.00%"),
        ("CVaR 95 Mthly", "cvar_m", "0.00%"),
        ("CVaR 95 Ann", "cvar_a", "0.00%"),
        ("Spending Stats", None, None),
        ("IRR", "irr", "0.00%"),
        ("Ruin? (this one historical path)", "ruin_hist", None),
        ("Probability of ruin (Monte Carlo, 2,000 sims)", "prob_ruin", "0.0%"),
        ("Shortfall years", "shortfall", "0"),
        ("Legacy (historical path)", "legacy_hist", "#,##0"),
        ("Median legacy (Monte Carlo)", "mc_median_legacy", "#,##0"),
        ("Fee (OCF pa)", "fee", "0.000%"),
    ]
    r = header_row
    for label, key, fmt in rows:
        lc = ws.cell(row=r, column=3, value=label)
        if key is None:
            lc.font = HEADER_FONT
            lc.fill = blue_fill()
            for k in range(len(names)):
                ws.cell(row=r, column=4 + k).fill = blue_fill()
        else:
            lc.fill = blue_fill()
            lc.border = BOX_BORDER
            for k, name in enumerate(names):
                cell = ws.cell(row=r, column=4 + k)
                cell.fill = blue_fill()
                cell.border = BOX_BORDER
                s = stats_by_name[name]
                if key == "window":
                    cell.value = f"{s['n_months']} ({s['start'].date()} to {s['end'].date()})"
                elif key == "ruin_hist":
                    cell.value = s["ruin_hist"]
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.value = s[key]
                    if fmt:
                        cell.number_format = fmt
        r += 1
    return r + 2


def _pdf_row(pdf, label, val_a, val_b, col_widths, bold=False):
    pdf.set_font("Helvetica", "B" if bold else "", 9)
    pdf.cell(col_widths[0], 7, label, border=1)
    pdf.cell(col_widths[1], 7, val_a, border=1, align="R")
    pdf.cell(col_widths[2], 7, val_b, border=1, align="R")
    pdf.ln()


def build_pdf(stats_v4, stats_fs, window_years, window_start, window_end):
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 17)
    pdf.set_text_color(20, 20, 20)
    pdf.cell(0, 9, "Mobius Better v4 (candidate) vs Aspen Four Seasons", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, "Prepared for Aspen Advisers UK - decumulation comparison", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(
        0, 5,
        f"Both portfolios evaluated over the same window ({window_start} to {window_end}, "
        f"~{window_years} years) - the common history available across Better v4's holdings. "
        "Scenario: age 65, GBP 500,000 pot, GBP 20,000/yr withdrawal (4%).",
    )
    pdf.ln(3)

    col_widths = [80, 45, 45]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(col_widths[0], 7, "", border=1, fill=True)
    pdf.cell(col_widths[1], 7, "Better v4", border=1, fill=True, align="R")
    pdf.cell(col_widths[2], 7, "Four Seasons", border=1, fill=True, align="R")
    pdf.ln()

    _pdf_row(pdf, "Portfolio Stats", "", "", col_widths, bold=True)
    _pdf_row(pdf, "Compound ret pa", f"{stats_v4['compound']*100:.2f}%", f"{stats_fs['compound']*100:.2f}%", col_widths)
    _pdf_row(pdf, "Volatility pa", f"{stats_v4['vol']*100:.2f}%", f"{stats_fs['vol']*100:.2f}%", col_widths)
    _pdf_row(pdf, "Downside stats", "", "", col_widths, bold=True)
    _pdf_row(pdf, "Max DD (worst drawdown)", f"{stats_v4['maxdd']*100:.2f}%", f"{stats_fs['maxdd']*100:.2f}%", col_widths)
    _pdf_row(pdf, "Average DD", f"{stats_v4['avgdd']*100:.2f}%", f"{stats_fs['avgdd']*100:.2f}%", col_widths)
    _pdf_row(pdf, "CVaR 95 Mthly", f"{stats_v4['cvar_m']*100:.2f}%", f"{stats_fs['cvar_m']*100:.2f}%", col_widths)
    _pdf_row(pdf, "CVaR 95 Ann", f"{stats_v4['cvar_a']*100:.2f}%", f"{stats_fs['cvar_a']*100:.2f}%", col_widths)
    _pdf_row(pdf, "Spending Stats", "", "", col_widths, bold=True)
    _pdf_row(pdf, "IRR", f"{stats_v4['irr']*100:.2f}%", f"{stats_fs['irr']*100:.2f}%", col_widths)
    _pdf_row(pdf, "Ruin? (this historical path)", stats_v4["ruin_hist"], stats_fs["ruin_hist"], col_widths)
    _pdf_row(pdf, "Probability of ruin (Monte Carlo)", f"{stats_v4['prob_ruin']*100:.1f}%",
             f"{stats_fs['prob_ruin']*100:.1f}%", col_widths)
    _pdf_row(pdf, "Shortfall years", str(stats_v4["shortfall"]), str(stats_fs["shortfall"]), col_widths)
    _pdf_row(pdf, "Legacy (historical path)", f"GBP {stats_v4['legacy_hist']:,.0f}",
             f"GBP {stats_fs['legacy_hist']:,.0f}", col_widths)
    _pdf_row(pdf, "Median legacy (Monte Carlo)", f"GBP {stats_v4['mc_median_legacy']:,.0f}",
             f"GBP {stats_fs['mc_median_legacy']:,.0f}", col_widths)
    _pdf_row(pdf, "Fee (OCF pa)", f"{stats_v4['fee']*100:.3f}%", f"{stats_fs['fee']*100:.3f}%", col_widths)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(90, 90, 90)
    pdf.cell(0, 6, "Data source note", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(60, 60, 60)
    pdf.multi_cell(
        0, 4.5,
        "'Berenberg / Protected Equities' (20%) uses the previous model's 'Eq Gbl DM Novum Mgd Vol' "
        "series (a managed-volatility equity strategy). All 11 holdings' return series come from the "
        "PREVIOUS Mobius model workbook's own Asset Returns tab - an older data source than the app's "
        "Bloomberg 14 July 2026 pull. Four Seasons is shown here on this same truncated window for a "
        "fair comparison, so its figures differ from the full-history ones shown elsewhere in this "
        "project.",
    )

    pdf.output(OUT_PDF)
    print(f"Saved {OUT_PDF}")


def build():
    print("Extracting old-model asset-class data...")
    old_df = extract_old_model_returns()
    print(f"  Common window: {old_df.index.min().date()} to {old_df.index.max().date()}, "
          f"{len(old_df)} months")

    weights_v4 = build_better_v4_weights()
    better_v4_monthly = (old_df[weights_v4.index] * weights_v4.values).sum(axis=1) - FLAT_FEE / 12
    better_v4_monthly.name = "Better v4"

    # Aspen Four Seasons, same-window truncation for a fair comparison against the older data source
    main_asset_df = load_asset_returns()
    fs_weights = asset_class_weights("Four Seasons")
    fs_fee = weighted_avg_fee("Four Seasons")
    fs_monthly_full = weighted_monthly_returns(fs_weights, fs_fee, main_asset_df, label="Four Seasons").dropna()
    fs_monthly = fs_monthly_full.loc[(fs_monthly_full.index >= old_df.index.min()) &
                                      (fs_monthly_full.index <= old_df.index.max())]

    cpi = load_cpi(main_asset_df)
    horizon_years = len(old_df) // 12  # ~25 years, matching the shorter common window

    print("Computing stats for Better v4...")
    stats_v4 = full_stats(better_v4_monthly, cpi, FLAT_FEE, horizon_years)
    print("Computing stats for Four Seasons (same truncated window)...")
    stats_fs = full_stats(fs_monthly, cpi, fs_fee, horizon_years)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Mobius Better v4 (candidate) vs Aspen Four Seasons"
    ws["B2"].font = Font(bold=True, size=15)
    ws["B3"] = (
        "Same style as the app's PDF summary, extended with downside-risk stats. Both portfolios "
        f"evaluated over the SAME window ({old_df.index.min().date()} to {old_df.index.max().date()}, "
        f"~{horizon_years} years) since that's the common history available across Better v4's "
        "holdings (incl. Eq Gbl DM Novum Mgd Vol for the Berenberg / Protected Equities slot) - this "
        "trims Four Seasons' most recent ~18 months for a fair comparison. Scenario: age 65, £500,000 "
        "pot, £20,000/yr withdrawal (4%)."
    )
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 56

    names = ["Better v4", "Four Seasons"]
    stats = {"Better v4": stats_v4, "Four Seasons": stats_fs}
    row = write_box(ws, 6, "Decumulation comparison", names, stats, None)

    ws.cell(row=row, column=2, value="Data source note").font = Font(bold=True, size=12)
    row += 1
    warn_cell = ws.cell(row=row, column=2)
    warn_cell.value = (
        "'Berenberg / Protected Equities' (20%) uses the previous model's 'Eq Gbl DM Novum Mgd Vol' "
        "series (a managed-volatility equity strategy). All 11 holdings' return series come from the "
        "PREVIOUS Mobius model workbook's own Asset Returns tab - an OLDER data source than the "
        "current app's Bloomberg 14 July 2026 pull, overlapping for all 11 holdings from 2001-01 to "
        "2025-01 - shorter than this project's usual ~26-year window. Four Seasons is shown here on "
        "this same truncated window for a fair comparison, so its figures differ from the full-history "
        "ones shown elsewhere in this project."
    )
    warn_cell.alignment = Alignment(wrap_text=True, vertical="top")
    warn_cell.fill = blue_fill()
    ws.merge_cells(f"B{row}:F{row+4}")
    ws.row_dimensions[row].height = 110
    for rr in range(row, row + 5):
        for cc in range(2, 7):
            ws.cell(row=rr, column=cc).fill = blue_fill()

    row += 6
    ws.cell(row=row, column=2, value="Better v4 weights used in this sheet").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=3, value="Holding").font = HEADER_FONT
    ws.cell(row=row, column=4, value="Weight").font = HEADER_FONT
    row += 1
    for label, (_col, orig_w) in OLD_MODEL_COLS.items():
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=4, value=orig_w).number_format = "0.0%"
        row += 1

    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 34
    for col in "DEF":
        ws.column_dimensions[col].width = 26

    wb.save(OUT)
    print(f"\nSaved {OUT}")

    build_pdf(stats_v4, stats_fs, horizon_years, old_df.index.min().date(), old_df.index.max().date())


if __name__ == "__main__":
    build()
