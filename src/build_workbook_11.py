"""Stage 11: Mortality. Adds sex/joint-life toggles to Inputs, embeds the S4 pension-scheme mortality
table (Male_Data/Female_Data, ages 20-120), and builds live survival-curve formulas (own life,
partner life, and joint 'at least one alive' life) driven entirely by Inputs toggles - the Excel-native
complement to the Python app's mortality module (src/mortality.py)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.0%"

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
MAX_HORIZON = 30  # must match build_workbook_6.py

wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())

# ---------------------------------------------------------------------
# Inputs: add a Mortality toggle block below the existing rows (15 is the last existing row)
# ---------------------------------------------------------------------
ws = wb["Inputs"]
r0 = 17
ws.cell(row=r0, column=2, value="Mortality (S4 pension-scheme table)").font = BOLD
mortality_rows = [
    ("sex",            "Client's sex (M/F)",                 "M",  None, "Drives the 'own life' survival curve (Mortality tab)"),
    ("joint_life_on",  "Joint life / model a partner too? (Y/N)", "N", None, "Y = money needs to last until the SECOND death"),
    ("partner_sex",    "Partner's sex (M/F)",                 "F",  None, "Only used if joint life = Y"),
    ("partner_age",    "Partner's starting age",               63,  None, "Only used if joint life = Y"),
]
r = r0 + 1
labels = {}
for key, label, val, fmt, note in mortality_rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=val)
    cell.font = BLUE
    cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = SUBTITLE_FONT
    labels[key] = r
    r += 1

for key, row in labels.items():
    NR[key] = f"Inputs!$C${row}"
Path("cellrefs.json").write_text(json.dumps(NR, indent=2))
print("Added mortality inputs:", labels)

# ---------------------------------------------------------------------
# Mortality sheet
# ---------------------------------------------------------------------
if "Mortality" in wb.sheetnames:
    del wb["Mortality"]
ws = wb.create_sheet("Mortality")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
for col in "BCDE":
    ws.column_dimensions[col].width = 13
ws.column_dimensions["G"].width = 3

ws["B2"] = "How long might the client actually live? (Mortality)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: this sheet holds the realistic year-by-year odds of the client being alive, used "
    "to work out 'ruin BEFORE death' (the pot running out while someone's still alive to feel it - the "
    "outcome that actually matters) rather than just the raw 'ruin by the end of the plan' figure.\n\n"
    "Source: the S4 table (CMI, UK pension-scheme member experience) - a better basis for a pension "
    "decumulation model than a general-population table (e.g. ONS National Life Tables), since pension "
    "scheme members tend to live somewhat longer than the population at large. qx = probability a "
    "person alive at that age dies before their next birthday. Sex/partner settings are on the Inputs "
    "tab; everything below recalculates from them."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 75

# --- raw qx table (ages 20-120) ---
mort_csv = Path(__file__).resolve().parent.parent / "data" / "mortality_qx.csv"
qx_df = pd.read_csv(mort_csv, index_col="age")

qx_header_row = 6
ws.cell(row=qx_header_row, column=2, value="Age").font = HEADER_FONT
ws.cell(row=qx_header_row, column=3, value="qx Male").font = HEADER_FONT
ws.cell(row=qx_header_row, column=4, value="qx Female").font = HEADER_FONT
for c in (2, 3, 4):
    ws.cell(row=qx_header_row, column=c).fill = HEADER_FILL
qx_first_row = qx_header_row + 1
for i, age in enumerate(qx_df.index):
    r = qx_first_row + i
    ws.cell(row=r, column=2, value=int(age)).font = BLACK
    ws.cell(row=r, column=3, value=float(qx_df.loc[age, "qx_male"])).number_format = "0.000000"
    ws.cell(row=r, column=4, value=float(qx_df.loc[age, "qx_female"])).number_format = "0.000000"
qx_last_row = qx_first_row + len(qx_df) - 1
age_range = f"$B${qx_first_row}:$B${qx_last_row}"
qxm_range = f"$C${qx_first_row}:$C${qx_last_row}"
qxf_range = f"$D${qx_first_row}:$D${qx_last_row}"

# --- survival curve table (columns F..I): t, S own, S partner, S effective (joint if toggled) ---
sc_header_row = 6
headers = ["t (years)", "S(t) own life", "S(t) partner", "S(t) effective"]
for j, h in enumerate(headers):
    c = 6 + j
    cell = ws.cell(row=sc_header_row, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(c)].width = 14
sc_first_row = sc_header_row + 1
sc_last_row = sc_first_row + MAX_HORIZON  # t = 0..MAX_HORIZON inclusive - the window the MC sheets/chart use
LE_HORIZON = 100  # extends the SAME table further (t up to 100) purely so curtate life expectancy isn't
                   # truncated at the 30yr model horizon - a 65-year-old's LE is ~20yrs, which needs
                   # survival data out to ~age 120, not just to age 95. Matches Python's life_expectancy().
sc_last_row_full = sc_first_row + LE_HORIZON

def qx_lookup_formula(age_expr, sex_expr):
    """qx for a given age (may exceed table max -> clamp to 1, i.e. certain death, matching the S4
    table's own closure convention at age 120) and sex ('M'/'F', from an Inputs cell)."""
    col_range = f'IF({sex_expr}="M",{qxm_range},{qxf_range})'
    return (f'IF({age_expr}>{qx_last_row-qx_first_row+20},1,'
            f'INDEX({col_range},MATCH({age_expr},{age_range},0)))')

for i in range(LE_HORIZON + 1):
    t = i
    r = sc_first_row + i
    ws.cell(row=r, column=6, value=t).font = BLACK
    if t == 0:
        ws.cell(row=r, column=7, value=1.0).number_format = "0.0000"
        ws.cell(row=r, column=8, value=1.0).number_format = "0.0000"
    else:
        prev_r = r - 1
        own_age_expr = f'({NR["age"]}+{t}-1)'
        own_qx = qx_lookup_formula(own_age_expr, NR["sex"])
        ws.cell(row=r, column=7, value=f'=G{prev_r}*(1-({own_qx}))').number_format = "0.0000"
        partner_age_expr = f'({NR["partner_age"]}+{t}-1)'
        partner_qx = qx_lookup_formula(partner_age_expr, NR["partner_sex"])
        ws.cell(row=r, column=8, value=f'=H{prev_r}*(1-({partner_qx}))').number_format = "0.0000"
    eff_formula = f'=IF({NR["joint_life_on"]}="Y",1-(1-G{r})*(1-H{r}),G{r})'
    ws.cell(row=r, column=9, value=eff_formula).number_format = "0.0000"
    if t > MAX_HORIZON:
        # rows beyond the 30yr model horizon are LE-only scratch space - dim them so it's clear they
        # aren't part of the chart/MC-facing curve
        for c in (6, 7, 8, 9):
            ws.cell(row=r, column=c).font = SUBTITLE_FONT

s_eff_range = f"$I${sc_first_row}:$I${sc_last_row}"
s_own_range = f"$G${sc_first_row}:$G${sc_last_row}"

# --- life expectancy (uses the FULL extended table, not just the 30yr model window) ---
le_row = sc_last_row_full + 2
ws.cell(row=le_row, column=6, value="Curtate life expectancy - own life (years)").font = BOLD
ws.cell(row=le_row, column=9, value=f"=SUM(G{sc_first_row+1}:G{sc_last_row_full})").number_format = "0.0"
ws.cell(row=le_row + 1, column=6, value="Curtate life expectancy - partner (years)").font = BOLD
ws.cell(row=le_row + 1, column=9, value=f"=SUM(H{sc_first_row+1}:H{sc_last_row_full})").number_format = "0.0"
ws.cell(row=le_row + 2, column=6, value="Probability of surviving the full 30yr model horizon (effective)").font = BOLD
ws.cell(row=le_row + 2, column=9, value=f"=I{sc_last_row}").number_format = PCT
ws.merge_cells(start_row=le_row, start_column=6, end_row=le_row, end_column=8)
ws.merge_cells(start_row=le_row + 1, start_column=6, end_row=le_row + 1, end_column=8)
ws.merge_cells(start_row=le_row + 2, start_column=6, end_row=le_row + 2, end_column=8)

ws.cell(row=le_row + 4, column=6, value=(
    "'S(t) effective' is the curve everything else in the workbook uses: it's the own-life curve if "
    "Inputs!joint_life_on = N, or the joint 'at least one of the couple still alive' curve "
    "(1-(1-S_own)(1-S_partner), assuming independence between the two lives) if Y. The chart and the "
    "MC sheets' 'ruin before death' calculations only use rows up to t=30 (the model horizon); the "
    "dimmed rows below extend the same table out to t=100 solely so the life expectancy figures above "
    "aren't artificially truncated at 30 years."
)).font = SUBTITLE_FONT
ws.cell(row=le_row + 4, column=6).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=le_row + 4, start_column=6, end_row=le_row + 4, end_column=9)
ws.row_dimensions[le_row + 4].height = 40

# --- chart ---
chart = LineChart()
chart.title = "Survival curve(s)"
chart.y_axis.title = "Probability alive"
chart.x_axis.title = "Years from today"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1
data = Reference(ws, min_col=7, max_col=9, min_row=sc_header_row, max_row=sc_last_row)
cats = Reference(ws, min_col=6, min_row=sc_first_row, max_row=sc_last_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 9
chart.width = 16
ws.add_chart(chart, f"F{le_row + 6}")

wb.save(OUT_PATH)

mortality_range = {
    "s_eff_col": 9, "s_own_col": 7, "s_partner_col": 8,
    "sc_first_row": sc_first_row, "sc_last_row": sc_last_row,
    "s_eff_range": s_eff_range, "s_own_range": s_own_range,
}
Path("mortality_range.json").write_text(json.dumps(mortality_range, indent=2))
print("Saved stage 11 (Mortality). Survival curve rows:", sc_first_row, "-", sc_last_row)
