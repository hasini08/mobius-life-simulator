"""
Builds output/Mobius_Wealth_Blue_Box_Summary.xlsx - a "blue box" summary workbook in the style of
the previous Mobius decumulation model's 'Input and Summary' sheet (Portfolio Stats / Spending Stats
block), extended to our current four portfolios: Aspen Original / Mobius Alternative (accumulation,
no withdrawals) and Aspen Four Seasons / Mobius Better (decumulation, £20,000/yr withdrawals).

Methodology matches the previous model's own formulas exactly, verified against its actual Excel
formulas (not guessed):
  - Compound ret pa : PRODUCT(1+monthly returns)^(12/N) - 1   (geometric, full history)
  - Volatility pa   : STDEV(monthly returns) * SQRT(12)        (full history)
  - CVaR 95 Mthly   : average of monthly returns below their own 5th percentile
  - CVaR 95 Ann     : average of ROLLING 12-month returns below their own 5th percentile
  - Max DD          : worst point of a running drawdown series (resets to 0 at each new peak)
All five of the above are genuine Excel formulas in the workbook, computed from a monthly-returns
sheet that is itself formula-driven (SUMPRODUCT of asset-class returns against portfolio weights,
net of fee) - live and auditable, not pasted values.

IRR / Ruin? / Shortfall years / Legacy depend on the client's specific spending path (guardrails,
inflation-linked withdrawals, tax) - that logic already exists, verified, in src/engine.py
(historical_single_path), so those four are computed via the SAME engine that powers the live app
and written as the underlying cash-flow/value series; IRR itself is still a live Excel IRR() formula
over that cash-flow series, and Ruin?/Shortfall/Legacy are formulas over the value series.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from engine import load_asset_returns, load_cpi, historical_single_path, run_simulation, ClientProfile
from portfolios import AC, PORTFOLIOS, asset_class_weights, weighted_avg_fee

OUT = "../output/Mobius_Wealth_Blue_Box_Summary.xlsx"

ASSET_CLASSES = list(AC.keys())  # fixed order shared by Asset Returns columns and Portfolio Weights columns
PORTFOLIO_ORDER = ["Original", "Alternative", "Four Seasons", "Better"]
DISPLAY = {
    "Original": "Aspen Original", "Alternative": "Mobius Alternative",
    "Four Seasons": "Aspen Four Seasons", "Better": "Mobius Better",
}
SCENARIO = {
    "Original": dict(starting_age=65, horizon_years=30, starting_pot=500_000.0, initial_annual_spend=0.0),
    "Alternative": dict(starting_age=65, horizon_years=30, starting_pot=500_000.0, initial_annual_spend=0.0),
    "Four Seasons": dict(starting_age=65, horizon_years=30, starting_pot=500_000.0, initial_annual_spend=20_000.0),
    "Better": dict(starting_age=65, horizon_years=30, starting_pot=500_000.0, initial_annual_spend=20_000.0),
}

BLUE = "E9F2FB"
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="C9C9C9")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def blue_fill():
    return PatternFill("solid", fgColor=BLUE)


def build():
    asset_df = load_asset_returns()
    cpi = load_cpi(asset_df)

    wb = Workbook()
    wb.remove(wb.active)

    ws_ar = wb.create_sheet("Asset Returns")
    ws_pw = wb.create_sheet("Portfolio Weights")
    ws_mr = wb.create_sheet("Monthly Returns")
    hist_sheets = {}
    for name in PORTFOLIO_ORDER:
        hist_sheets[name] = wb.create_sheet(f"Hist - {name}")
    ws_summary = wb.create_sheet("Summary", 0)

    # ---------------------------------------------------------------- Asset Returns
    ws_ar["A1"] = "Date"
    ws_ar["A1"].font = HEADER_FONT
    for j, ac in enumerate(ASSET_CLASSES):
        c = ws_ar.cell(row=1, column=2 + j, value=ac)
        c.font = HEADER_FONT
    dates = asset_df.index
    for i, dt in enumerate(dates):
        r = i + 2
        ws_ar.cell(row=r, column=1, value=dt.date()).number_format = "yyyy-mm-dd"
        for j, ac in enumerate(ASSET_CLASSES):
            col = AC[ac]
            val = asset_df[col].iloc[i] if col in asset_df.columns else None
            v = None if (val is None or pd.isna(val)) else float(val)
            cell = ws_ar.cell(row=r, column=2 + j, value=v)
            cell.number_format = "0.00%"
    n_rows = len(dates)
    ar_last = n_rows + 1
    ws_ar.freeze_panes = "B2"
    ws_ar.column_dimensions["A"].width = 12

    # ---------------------------------------------------------------- Portfolio Weights
    ws_pw["A1"] = "Portfolio"
    ws_pw["A1"].font = HEADER_FONT
    for j, ac in enumerate(ASSET_CLASSES):
        c = ws_pw.cell(row=1, column=2 + j, value=ac)
        c.font = HEADER_FONT
    fee_col = 2 + len(ASSET_CLASSES)
    ws_pw.cell(row=1, column=fee_col, value="Fee (OCF pa)").font = HEADER_FONT
    for i, name in enumerate(PORTFOLIO_ORDER):
        r = i + 2
        ws_pw.cell(row=r, column=1, value=DISPLAY[name])
        w = asset_class_weights(name)
        for j, ac in enumerate(ASSET_CLASSES):
            v = float(w.get(ac, 0.0))
            cell = ws_pw.cell(row=r, column=2 + j, value=v)
            cell.number_format = "0.00%"
        fee_cell = ws_pw.cell(row=r, column=fee_col, value=float(weighted_avg_fee(name)))
        fee_cell.number_format = "0.000%"
    pw_row = {name: i + 2 for i, name in enumerate(PORTFOLIO_ORDER)}
    ac_first_col = get_column_letter(2)
    ac_last_col = get_column_letter(1 + len(ASSET_CLASSES))
    fee_col_letter = get_column_letter(fee_col)

    # ---------------------------------------------------------------- Monthly Returns (formula-driven)
    ws_mr["A1"] = "Date"
    ws_mr["A1"].font = HEADER_FONT
    triplet_headers = []
    for name in PORTFOLIO_ORDER:
        triplet_headers += [DISPLAY[name], f"{DISPLAY[name]} - Rolling 12m", f"{DISPLAY[name]} - Drawdown"]
    for j, h in enumerate(triplet_headers):
        ws_mr.cell(row=1, column=2 + j, value=h).font = HEADER_FONT

    mr_cols = {}  # name -> (monthly_col_letter, roll_col_letter, dd_col_letter)
    for k, name in enumerate(PORTFOLIO_ORDER):
        m_col = get_column_letter(2 + 3 * k)
        r_col = get_column_letter(3 + 3 * k)
        d_col = get_column_letter(4 + 3 * k)
        mr_cols[name] = (m_col, r_col, d_col)

    for i in range(n_rows):
        r = i + 2
        ar_r = i + 2
        ws_mr.cell(row=r, column=1, value=ws_ar.cell(row=ar_r, column=1).value).number_format = "yyyy-mm-dd"
        for name in PORTFOLIO_ORDER:
            m_col, roll_col, dd_col = mr_cols[name]
            pr = pw_row[name]
            m_cell = ws_mr[f"{m_col}{r}"]
            m_cell.value = (
                f"=SUMPRODUCT('Asset Returns'!${ac_first_col}{ar_r}:${ac_last_col}{ar_r},"
                f"'Portfolio Weights'!${ac_first_col}${pr}:${ac_last_col}${pr})"
                f"-'Portfolio Weights'!${fee_col_letter}${pr}/12"
            )
            m_cell.number_format = "0.00%"

            roll_cell = ws_mr[f"{roll_col}{r}"]
            if i >= 11:
                ref = f"{roll_col}{r}"
                # PRODUCT(1+range) needs a true array formula - without CSE entry, Excel's implicit
                # intersection collapses "1+range" to a single cell (the one in the formula's own
                # row), silently corrupting the result to a single month's return instead of the
                # 12-month compound - confirmed by testing (matches how the previous Mobius model
                # itself stores this exact calculation as an ArrayFormula, not a plain formula).
                roll_cell.value = ArrayFormula(ref=ref, text=f"=PRODUCT(1+{m_col}{r-11}:{m_col}{r})-1")
                roll_cell.number_format = "0.00%"

            dd_cell = ws_mr[f"{dd_col}{r}"]
            if i == 0:
                dd_cell.value = f"=MIN(0,{m_col}{r})"
            else:
                dd_cell.value = f"=MIN(0,(1+{dd_col}{r-1})*(1+{m_col}{r})-1)"
            dd_cell.number_format = "0.00%"
    mr_last = n_rows + 1
    ws_mr.freeze_panes = "B2"
    ws_mr.column_dimensions["A"].width = 12

    # ---------------------------------------------------------------- Hist - <portfolio> sheets
    hist_meta = {}
    for name in PORTFOLIO_ORDER:
        profile = ClientProfile(**SCENARIO[name])
        hdf = historical_single_path(name, asset_df, cpi, profile)
        ws = hist_sheets[name]
        headers = ["Date", "Portfolio Value", "Spend (net received)", "Cash flow (for IRR)", "Shortfall year?"]
        for j, h in enumerate(headers):
            ws.cell(row=1, column=1 + j, value=h).font = HEADER_FONT
        n = len(hdf)
        initial_spend = SCENARIO[name]["initial_annual_spend"]
        for i in range(n):
            r = i + 2
            date_v = hdf["Date"].iloc[i]
            val_v = float(hdf["PortfolioValue"].iloc[i])
            spend_v = float(hdf["Spend"].iloc[i])
            if i == 0:
                cash_flow = -val_v
            elif i == n - 1:
                cash_flow = spend_v + val_v
            else:
                cash_flow = spend_v
            shortfall = 1 if (i > 0 and (initial_spend - spend_v) > 1e-6) else 0
            ws.cell(row=r, column=1, value=date_v.date()).number_format = "yyyy-mm-dd"
            ws.cell(row=r, column=2, value=val_v).number_format = "#,##0"
            ws.cell(row=r, column=3, value=spend_v).number_format = "#,##0"
            ws.cell(row=r, column=4, value=cash_flow).number_format = "#,##0"
            ws.cell(row=r, column=5, value=shortfall)
        ws.column_dimensions["A"].width = 12
        for col in "BCDE":
            ws.column_dimensions[col].width = 16
        hist_meta[name] = dict(sheet=f"'Hist - {name}'", last_row=n + 1, years=profile.horizon_years,
                                start_pot=profile.starting_pot)

    # Monte Carlo probability of ruin - NOT the same thing as the "Ruin?" Y/N row above. "Ruin?" only
    # answers whether THIS ONE actual historical sequence happened to deplete the pot; this is the
    # share of THOUSANDS of simulated alternative futures that do, from the same engine/settings
    # (stationary block bootstrap, 2,000 sims, seed 42) as the live app's default. A portfolio can
    # show "Ruin? N" here and still carry a high probability of ruin - that's not a contradiction,
    # it's the difference between one real historical path and a full Monte Carlo distribution.
    prob_ruin = {}
    for name in PORTFOLIO_ORDER:
        profile = ClientProfile(**SCENARIO[name])
        res = run_simulation(name, asset_df, cpi, profile, method="stationary_block", n_sims=2000, seed=42)
        prob_ruin[name] = res.summary()["Probability of ruin"]

    # ---------------------------------------------------------------- Summary (the blue box)
    ws = ws_summary
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Mobius Wealth — Blue Box Summary"
    ws["B2"].font = Font(bold=True, size=15)
    ws["B3"] = ("Portfolio Stats computed over the full available historical window (1999/2000–2026, "
                "monthly). Spending Stats computed over a 30-year client scenario (age 65, £500,000 pot; "
                "Aspen Original / Mobius Alternative at 0% withdrawal - accumulation; Aspen Four Seasons / "
                "Mobius Better at 4% withdrawal, £20,000/yr - decumulation), via the same engine that "
                "powers the live Streamlit app. Note: 'Ruin?' and 'Probability of ruin' are different "
                "things - 'Ruin?' checks only the ONE actual historical sequence; 'Probability of ruin' "
                "is the share of 2,000 simulated alternative futures that deplete the pot. A portfolio "
                "can show 'Ruin? N' and still carry a high probability of ruin.")
    ws["B3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B3:J3")
    ws.row_dimensions[3].height = 70

    def write_box(top_row, title, names):
        ws.cell(row=top_row, column=2, value=title).font = TITLE_FONT
        header_row = top_row + 2
        ws.cell(row=header_row - 1, column=3, value="Portfolio Name  >>").font = Font(italic=True, color="808080")
        for k, name in enumerate(names):
            c = ws.cell(row=header_row - 1, column=4 + k, value=DISPLAY[name])
            c.font = HEADER_FONT
            c.fill = blue_fill()
            c.border = BOX_BORDER
            c.alignment = Alignment(horizontal="center")

        stat_rows = [
            ("Portfolio Stats", None),
            ("Compound ret pa", "compound"),
            (" ", "vol"),
            ("CVaR 95 Mthly", "cvar_m"),
            ("CVaR 95 Ann", "cvar_a"),
            ("Max DD", "maxdd"),
            ("Spending Stats", None),
            ("IRR", "irr"),
            ("Ruin? (this one historical path)", "ruin"),
            ("Probability of ruin (Monte Carlo, 2,000 sims)", "prob_ruin"),
            ("Shortfall years", "shortfall"),
            ("Legacy", "legacy"),
        ]
        r = header_row
        label_col = 3
        for label, kind in stat_rows:
            lc = ws.cell(row=r, column=label_col, value=label)
            if kind is None:
                lc.font = HEADER_FONT
                lc.fill = blue_fill()
                for k in range(len(names)):
                    ws.cell(row=r, column=4 + k).fill = blue_fill()
            else:
                lc.fill = blue_fill()
                lc.border = BOX_BORDER
                for k, name in enumerate(names):
                    m_col, roll_col, dd_col = mr_cols[name]
                    meta = hist_meta[name]
                    cell = ws.cell(row=r, column=4 + k)
                    cell.fill = blue_fill()
                    cell.border = BOX_BORDER
                    if kind == "compound":
                        ref = cell.coordinate
                        formula = (f"=PRODUCT(1+'Monthly Returns'!${m_col}$2:${m_col}${mr_last})"
                                   f"^(12/COUNT('Monthly Returns'!${m_col}$2:${m_col}${mr_last}))-1")
                        cell.value = ArrayFormula(ref=ref, text=formula)
                        cell.number_format = "0.00%"
                    elif kind == "vol":
                        cell.value = f"=STDEV('Monthly Returns'!${m_col}$2:${m_col}${mr_last})*SQRT(12)"
                        cell.number_format = "0.00%"
                    elif kind == "cvar_m":
                        rng = f"'Monthly Returns'!${m_col}$2:${m_col}${mr_last}"
                        cell.value = f"=AVERAGEIF({rng},\"<\"&PERCENTILE({rng},0.05))"
                        cell.number_format = "0.00%"
                    elif kind == "cvar_a":
                        rng = f"'Monthly Returns'!${roll_col}$2:${roll_col}${mr_last}"
                        cell.value = f"=AVERAGEIF({rng},\"<\"&PERCENTILE({rng},0.05))"
                        cell.number_format = "0.00%"
                    elif kind == "maxdd":
                        cell.value = f"=MIN('Monthly Returns'!${dd_col}$2:${dd_col}${mr_last})"
                        cell.number_format = "0.00%"
                    elif kind == "irr":
                        cell.value = f"=IRR({meta['sheet']}!$D$2:$D${meta['last_row']})"
                        cell.number_format = "0.00%"
                    elif kind == "ruin":
                        cell.value = f"=IF(MIN({meta['sheet']}!$B$2:$B${meta['last_row']})<=1,\"Y\",\"N\")"
                        cell.alignment = Alignment(horizontal="center")
                    elif kind == "prob_ruin":
                        # Not a formula - depends on a Monte Carlo draw (thousands of simulated
                        # futures), computed via the same engine/settings as the live app and
                        # inserted as a value, consistent with how the app's own headline number works.
                        cell.value = float(prob_ruin[name])
                        cell.number_format = "0.0%"
                    elif kind == "shortfall":
                        cell.value = f"=COUNTIF({meta['sheet']}!$E$2:$E${meta['last_row']},1)"
                        cell.number_format = "0"
                    elif kind == "legacy":
                        cell.value = f"={meta['sheet']}!$B${meta['last_row']}"
                        cell.number_format = "#,##0"
            r += 1
        return r + 2

    next_row = write_box(6, "Accumulation — Aspen Original vs Mobius Alternative", ["Original", "Alternative"])
    write_box(next_row, "Decumulation — Aspen Four Seasons vs Mobius Better", ["Four Seasons", "Better"])

    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 22
    for col in "DEFG":
        ws.column_dimensions[col].width = 20

    wb.save(OUT)
    print(f"Saved {OUT}")


if __name__ == "__main__":
    build()
