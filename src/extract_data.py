"""
Extract and clean monthly return data from the Bloomberg_Decumulation_Data file.

Produces tidy CSVs in ../data/:
  - asset_class_returns.csv : monthly TOTAL RETURN for broad asset classes + CPI + cash
  - fund_returns.csv        : monthly TOTAL RETURN for individual named funds

All series indexed by month-end Date. Values are monthly simple returns (fractions,
e.g. 0.0123 = 1.23%), except UK CPI YoY which is a level (annual inflation rate, e.g. 0.032).
"""
import openpyxl
import pandas as pd
from pathlib import Path

SRC = "/root/.claude/uploads/6277f0d3-e8e6-5e30-977d-0317d499a601/6abe440c-Bloomberg_Decumulation_Data_9_July_2026.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data"
OUT.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Bloomberg Direct Returns"]

# Row 1 = block header (name), each block is 3 cols wide: Date, Return, blank
# Row 2 = 'Date' / 'Monthly Return (Bloomberg direct)'
# Data starts row 3
max_col = ws.max_column
max_row = ws.max_row

blocks = []
col = 1
while col <= max_col:
    header = ws.cell(row=1, column=col).value
    if header:
        blocks.append((col, str(header).strip()))
    col += 1

print(f"Found {len(blocks)} data blocks")

series = {}
for start_col, name in blocks:
    dates = []
    vals = []
    for r in range(3, max_row + 1):
        d = ws.cell(row=r, column=start_col).value
        v = ws.cell(row=r, column=start_col + 1).value
        if d is None:
            continue
        dates.append(pd.Timestamp(d))
        vals.append(v)
    # normalise every date to the TRUE calendar month-end so all series align on one monthly index
    norm_dates = pd.DatetimeIndex(dates).to_period("M").to_timestamp("M")
    s = pd.Series(vals, index=norm_dates, name=name)
    s = s[~s.index.duplicated(keep="last")].sort_index()
    series[name] = s

# Classify into asset-class-level series vs individual fund series based on known names
ASSET_CLASS_NAMES = [
    "Global Equities — MSCI World Net TR (NDDUWI Index)",
    "Global Bonds — Bloomberg Global Agg TR (LEGATRUU Index)",
    "UK Gilts All Stocks — FTSE Actuaries UK Conventional Gilts All Stocks (FTFIBGT Index)",
    "UK Gilts Over 15 Years — FTSE Actuaries UK Conventional Gilts Over 15 Years (FTRFBGH Index)",
    "Securitised Credit — Bloomberg US Securitized TR (I05582GB Index)",
    "FTSE EPRA NAREIT Developed Total Return Index USD (RUGL)",
    "Infrastructure Equities — MSCI World Infrastructure Net Total Return USD Index (M1W0OINF)",
    "Commodities — Bloomberg Commodity TR (BCOMTR Index)",
    "UK CPI YoY — Bloomberg PX_LAST (UKRPCJYR Index)",
    "Cash (GBP) — SONIA / short rate proxy (SONIO/N Index)",
    "FTSE Actuaries Govt Securities UK Index Linked TR Over 5 Yr",
    "Blackrock ICS Sterling Liquidity Fund",
    "MSCI Emerging Markets Index",
]

asset_class = {}
funds = {}
for name, s in series.items():
    key = name.strip()
    matched = None
    for ac in ASSET_CLASS_NAMES:
        if key.rstrip() == ac.rstrip() or key.rstrip().rstrip() == ac.strip():
            matched = ac
            break
    if matched:
        asset_class[matched] = s
    else:
        funds[key] = s

print("Asset classes found:", list(asset_class.keys()))
print("\nFund-level series found:")
for k in funds:
    print(" -", k, f"(n={funds[k].count()}, from {funds[k].first_valid_index()})")

ac_df = pd.DataFrame(asset_class)
fund_df = pd.DataFrame(funds)

ac_df.to_csv(OUT / "asset_class_returns.csv")
fund_df.to_csv(OUT / "fund_returns.csv")
print(f"\nSaved asset_class_returns.csv {ac_df.shape} and fund_returns.csv {fund_df.shape} to {OUT}")
