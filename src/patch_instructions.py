"""Patch: rewrite the Instructions sheet only, to document the Equity Sweep, Sensitivity Tables and
glide-path features added after the initial build - without touching any other sheet (avoids a full
rebuild of the ~245k-formula workbook)."""
import openpyxl
from openpyxl.styles import Font, Alignment

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)

del wb["Instructions"]
ws = wb.create_sheet("Instructions", 1)  # keep it second, right after Summary
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 110
ws["B2"] = "Mobius Wealth — Decumulation Simulator (v4, refined)"
ws["B2"].font = TITLE_FONT
notes = [
    "",
    "This workbook refines the previous Mobius decumulation model using the Bloomberg data supplied "
    "9 July 2026, the FNZ 'Growth Passive Plus' holdings, and the Bank of England SONIA series "
    "(IUDSOIA) for the Cash asset class.",
    "",
    "COMPETITIVE FRAMING (for client-facing use): 'Original' stands in for the comparison/benchmark "
    "offering (Aspen Advisers, aspenadvisers.com), 'Alternative' is Mobius's own core offering, and "
    "'Better' is Mobius's enhanced/diversified alternative. The underlying holdings data and "
    "modelling methodology are unchanged either way - this is just how the three columns should be "
    "presented/named in front of a client.",
    "",
    "-------------------------------------------------------------------------------------------",
    "START HERE - PLAIN-ENGLISH GUIDE (no financial-modelling background needed)",
    "-------------------------------------------------------------------------------------------",
    "",
    "WHAT THIS WORKBOOK DOES, IN ONE SENTENCE: it asks 'if this client retires with this pot and "
    "spends this much every year, how likely is the money to last?' - by testing the plan against "
    "thousands of different possible market futures (and, optionally, realistic odds of how long the "
    "client actually lives), rather than assuming one fixed rate of return that may or may not happen.",
    "",
    "THE ONE NUMBER THAT MATTERS MOST: 'Probability of ruin' - the share of tested futures in which "
    "the pot runs out before the plan is meant to end. Lower is safer. Everything else in this "
    "workbook exists to help explain WHY that number is what it is, and which decisions move it.",
    "",
    "HOW TO READ THIS WORKBOOK: start on the Summary tab for the headline comparison across the three "
    "portfolios. Inputs is where every client detail and every on/off toggle lives (yellow cells - "
    "change these to re-run the plan for a different client or scenario). Everything else is either "
    "supporting detail (how a number was built) or an optional 'what if' feature you can switch on.",
    "",
    "GLOSSARY (plain-English meaning of terms used throughout this workbook):",
    "  - Probability of ruin: out of every future tested, the share where the pot hits £0 before the "
    "plan is meant to end.",
    "  - Monte Carlo simulation / simulated path: instead of guessing one future for investment "
    "markets, the model plays out many different possible futures and reports how many go well. Each "
    "individual 'possible future' is called a path.",
    "  - Withdrawal rate: yearly spend divided by the starting pot, as a percentage - a simple gauge "
    "of how hard the pot is being asked to work.",
    "  - Guardrails: an optional rule that trims spending a little after weak markets and allows a "
    "little more after strong ones, instead of spending an identical £ amount regardless of markets.",
    "  - Legacy: whatever is left in the pot at the end of the plan (or at death, in the mortality-"
    "adjusted figures) - what the client would leave behind.",
    "  - Forward-looking blend / CMA (Capital Market Assumptions): by default the model learns "
    "purely from actual market history. This optional setting also factors in what professional "
    "forecasters currently expect for the next 10 years, which tends to be more cautious.",
    "  - Mortality: switches the model from 'assume the client lives to the end of the plan' to 'use "
    "realistic odds of the client being alive at each age' - which changes how the ruin and legacy "
    "figures should be read (see 'ruin BEFORE death' below).",
    "  - Annuity / annuitization: swapping part of the pot, once, for an income guaranteed to be paid "
    "for the rest of the client's life, however long that turns out to be.",
    "  - Ruin BEFORE death: the pot running out while the client (or, for a couple, at least one "
    "partner) is still alive - the outcome that actually matters, as opposed to the raw 'ruin by the "
    "end of the plan' figure, which also counts paths that only run dry after everyone involved has "
    "already died (not really a failure of the plan).",
    "  - Sensitivity analysis: re-testing the plan while changing one setting at a time (spend level, "
    "how much is in shares, guardrail settings) to see how much each one actually matters.",
    "",
    "-------------------------------------------------------------------------------------------",
    "TECHNICAL DETAIL BELOW - for advisers/analysts who want the full methodology, sourcing and "
    "assumptions behind every figure. Skip ahead to the Summary tab if you just want the results.",
    "-------------------------------------------------------------------------------------------",
    "",
    "WHAT'S NEW vs the previous model:",
    "  1. Compares THREE portfolio variants: 'Original' (mainstream retail fund lineup), "
    "'Alternative' (tax/cost-efficient unit-linked lineup - same underlying market exposure, lower fees), "
    "and 'Better' (a more diversified allocation adding REITs, infrastructure, commodities, EM and "
    "index-linked gilts, in the spirit of the previous model's winning portfolio).",
    "  2. Inflation-linked spending using UK CPI (Bloomberg data), not a fixed assumption.",
    "  3. Spending guardrails (Guyton-Klinger style) that cut spend in weak markets and raise it in "
    "strong ones - togglable on/off to see the impact on ruin probability vs shortfall frequency.",
    "  4. An accompanying Python/Streamlit app (see /app) adds an IMPROVED stochastic sampling engine "
    "(stationary block bootstrap + skewed-distribution option) beyond what a spreadsheet can practically "
    "do - this workbook's own Monte Carlo uses a simple annual bootstrap for a direct, auditable "
    "comparison to the previous model's methodology.",
    "  5. Confidence intervals on every probability-of-ruin estimate (standard error + 95% CI), so the "
    "Monte Carlo sampling noise is visible rather than implied to be exact - see the Summary tab and "
    "the app's headline statistics.",
    "  6. Equity allocation sweep: both this workbook (the 'Equity Sweep' tab, historical/formula-driven) "
    "and the Python app (full Monte Carlo) scan total equity weight from 20%-100% per portfolio, "
    "rescaling the rest of the portfolio proportionally, to show where risk/return trades off.",
    "  7. Sensitivity tables: both this workbook (the 'Sensitivity Tables' tab) and the app scan "
    "probability of ruin / final legacy against the initial withdrawal rate and the guardrail band "
    "width independently, one lever at a time.",
    "  8. Equity glide path (dynamic allocation): the Python app can run a de-risking (or up-risking) "
    "glide path - equity weight moving linearly from a starting to an ending level over the horizon - "
    "compared against a fixed allocation. Not implemented in this workbook: a formula-driven annual "
    "Monte Carlo bootstrap can't cleanly vary the weight vector by simulated year within a single "
    "spreadsheet formula without either a full rebuild per year-step or an array/macro approach outside "
    "this workbook's macro-free, LibreOffice-compatible design goal - use the app for this feature.",
    "  9. Asset class correlation: both this workbook (the 'Asset Correlation' tab, a colour-scaled "
    "matrix) and the app show the monthly return correlation across all 11 broad asset classes - useful "
    "for seeing which of the 'Better' portfolio's extra holdings (REITs, infrastructure, etc.) actually "
    "diversify vs just add complexity (several run ~0.75+ correlated with Global Equities).",
    "  10. Spending shortfall heatmap: the app sweeps probability of ruin (or % paths with any "
    "shortfall) across a 2D grid of withdrawal rate x equity weight simultaneously, showing how the two "
    "levers interact (not implemented in this workbook - a 2D live Monte Carlo grid is impractical as a "
    "formula-driven spreadsheet; use the app).",
    "  11. Mortality (S4 pension-scheme table): both this workbook (the 'Mortality' tab plus new columns "
    "on the Summary and each MC tab) and the app now model the client's own mortality, using the S4 "
    "table (CMI, UK self-administered pension scheme experience - a better basis for pension "
    "decumulation than a general-population/ONS table). Supports single life (with a sex toggle) and "
    "joint life (a partner's age/sex, using the 'at least one of the couple still alive' survival curve). "
    "Adds mortality-adjusted metrics alongside the raw ones: 'probability of ruin BEFORE death' (the pot "
    "hits zero while the client is still alive - usually much lower than the raw horizon-end ruin "
    "probability, computed EXACTLY via each simulated path's ruin year against the survival curve, "
    "given the standard assumption that markets and mortality are independent) and 'legacy at death' "
    "(the estate value at a sampled death year rather than at a fixed year-30 cutoff).",
    "  12. Tax + State Pension: both this workbook (the 'Tax' tab, plus new columns on the "
    "Historical Projection and MC tabs) and the app now gross up the client's desired spend for UK "
    "income tax and net it against the State Pension. When Inputs!apply_tax_on = Y, 'Desired annual "
    "spend' is treated as the NET (take-home) figure the client wants, not a pre-tax number - the "
    "model works out how much has to actually leave the pot (taxable) to deliver that, and reduces "
    "the required withdrawal once the State Pension starts. This is often a BIGGER lever on "
    "probability of ruin than portfolio choice: State Pension alone can cover 40-60%+ of a modest "
    "spending need, dramatically de-risking the years it's in payment. SIMPLIFIED BASIS (see the Tax "
    "tab for the full caveat): the whole pot is treated as a taxable pension wrapper (no 25% "
    "tax-free lump sum, ISA or GIA modelling yet - a natural next enhancement), rest-of-UK bands only "
    "(not Scotland), and tax bands/State Pension are held in today's money (assumed to rise with "
    "inflation) rather than literally freezing current nominal thresholds for 30 years.",
    "  13. Forward-looking Capital Market Assumptions (NEW): the pure historical bootstrap (2000-2026) "
    "spans an unusually strong equity run, which can flatter a plan that's only ever been tested "
    "against it. Inputs!cma_blend (the 'CMA' tab) lets you optionally recentre each asset class's "
    "AVERAGE monthly return towards independently published 10-year forward-looking forecasts "
    "(Monevator's compilation of Vanguard/Schroders/JPMorgan/BlackRock and others), while leaving "
    "history's volatility, correlation and worst-case behaviour completely untouched - 0% (default) "
    "uses pure history as everywhere before; 100% recentres fully to the forecasts (mostly LOWER for "
    "equities/REITs, HIGHER for some bonds than this sample's history), which typically raises "
    "probability of ruin - that's the blend correctly showing a plan is more fragile than a single "
    "strong historical window suggested, not a bug. Applied identically in the app (src/cma.py, "
    "shifts asset-class monthly returns before they reach the simulation) and this workbook (CMA tab, "
    "shifts the same three portfolios' monthly returns via one additive term in 'Portfolio Returns', "
    "cascading automatically through Portfolio Annual Returns into both Historical Projection and "
    "every MC tab).",
    "  14. Partial annuitization comparison (NEW): both this workbook (the 'Annuity' tab, plus updated "
    "Historical Projection / MC tabs) and the app can model converting Inputs!annuity_pct of the "
    "starting pot into a guaranteed LIFETIME income at outset, using real, dated UK best-buy annuity "
    "rates (Hargreaves Lansdown, May 2026 - see the Annuity tab for the full sourcing). The guaranteed "
    "income is LEVEL (does not rise with inflation, unlike everything else in this model) and offsets "
    "the withdrawal need from the (now-smaller) remaining pot every year, for life - dramatically "
    "cutting probability of ruin at the cost of giving up upside and, for a level annuity, real-terms "
    "purchasing power over time. The app additionally shows this mortality-adjusted ('before death') "
    "vs raw, and lets you compare joint-life (50%-to-survivor) vs single-life pricing.",
    "",
    "STRUCTURE:",
    "  - Inputs: all client parameters and toggles (yellow cells), including the mortality/sex/joint-"
    "life and tax/State Pension settings.",
    "  - Portfolios: holdings, weights and fees for the three variants, sourced from the FNZ holdings "
    "file, with an asset-class weight roll-up used to drive the return calculations.",
    "  - Asset Class Returns: raw monthly total-return data by asset class (Bloomberg, to 9 July 2026; "
    "Cash from Bank of England SONIA, spliced with a money-market fund proxy pre-2014 - see below).",
    "  - Annual Asset Returns: calendar-year compounded returns derived from the monthly data.",
    "  - Portfolio Returns / Portfolio Annual Returns: each portfolio's net-of-fee return (monthly, "
    "then calendar-year compounded) and the year's inflation - built to match the Python engine's "
    "monthly-compounding method exactly.",
    "  - Tax: the UK income tax bands (editable) and the derived closed-form breakpoints used "
    "everywhere else to gross a desired NET spend up into the actual taxable pot withdrawal needed, "
    "plus an illustrative before/after-State-Pension worked example for the Inputs-tab client.",
    "  - CMA: forward-looking 10-year return forecasts per asset class (editable), each one's "
    "historical monthly mean from this workbook's own data, the derived monthly shift, and each "
    "portfolio's SUMPRODUCT-of-weights-and-shifts (blended into 'Portfolio Returns' via "
    "Inputs!cma_blend).",
    "  - Annuity: the single-life annuity-rate table (editable, by age), the joint-life discount "
    "factor, an interpolated rate for the Inputs-tab client's own age, and the resulting guaranteed "
    "nominal annual income (blended into Historical Projection / MC via Inputs!annuity_pct).",
    "  - Historical Projection: a single deterministic path per portfolio using the ACTUAL historical "
    "sequence of annual returns, for a reasonableness check - now with a 'Gross withdrawal from pot' "
    "column (drives the pot mechanics) and a 'Net received (post-tax)' column (what the client "
    "actually gets - identical to the withdrawal figure when tax is off). The starting pot and every "
    "year's withdrawal target both reflect Inputs!annuity_pct (see Annuity tab) - a no-op at 0%.",
    "  - MC Original / MC Alternative / MC Better: an annual bootstrap Monte Carlo (recalculates live - "
    "press F9 / Ctrl+Shift+F9 to redraw a new set of random paths). Each simulated year now carries a "
    "materialised 'Real spend' (NET, today's-money, guardrail-adjusted) column alongside 'Withdrawal' "
    "(nominal, gross of tax when the toggle is on, net of any annuity income) and 'Pot' (starting "
    "value reduced by Inputs!annuity_pct), plus a 'Ruin year', 'P(alive when ruined)', a random "
    "'Death year' draw and 'Legacy at death' per simulated path.",
    "  - Equity Sweep: rescales each portfolio's weights to a grid of total equity weights (20%-100%) "
    "and reports CAGR, volatility, max drawdown and a historical-path legacy/shortfall check at each.",
    "  - Sensitivity Tables: the same historical-path method, swept against withdrawal rate and "
    "guardrail band width instead of equity weight.",
    "  - Asset Correlation: monthly return correlation matrix across all 11 broad asset classes, colour-"
    "scaled as a heatmap.",
    "  - Mortality: the S4 qx table (ages 20-120, male/female) plus live survival-curve formulas (own "
    "life, partner, and the 'effective' curve everything else uses), life expectancy and a chart.",
    "  - Summary: headline statistics (probability of ruin with SE/95% CI, legacy quantiles, shortfall "
    "years) for all three portfolios, historical and simulated, plus a mortality-adjusted results "
    "block - all of it recalculating live off the guardrails AND tax/State Pension toggles on Inputs "
    "(there's no separate 'tax-adjusted' block here since tax/State Pension change the SAME headline "
    "figures directly, rather than being a parallel what-if; see the Tax tab for an illustrative "
    "before/after-State-Pension breakdown in isolation).",
    "",
    "ASSUMPTIONS TO CONFIRM (flagged in place, see cell comments too):",
    "  - Original-portfolio fund OCFs are NOT given in the source data (only Alternative AMCs are) - "
    "typical published OCFs for this fund type have been assumed. Confirm against real factsheets.",
    "  - The 'Better' portfolio's weights are a judgement-based construction, not sourced from FNZ data.",
    "  - Cash Plus and Four Seasons Fund (the other two FNZ portfolios) are not modelled - by "
    "instruction, since the comparison in scope is Original vs Alternative vs Better.",
    "  - Mortality assumes independence between market returns and the client's mortality (a standard "
    "simplifying assumption), and independence between the two lives of a couple for joint life (in "
    "reality couples' deaths are not fully independent - e.g. shared lifestyle/socioeconomic factors - "
    "but modelling that correlation needs data this exercise doesn't have).",
    "  - Tax is SIMPLIFIED: the whole pot is treated as a taxable pension wrapper (every pound "
    "withdrawn is taxed as income) - no 25% pension-commencement lump sum, ISA (tax-free) or GIA "
    "(capital gains, not income) modelling, which would need to know a real client's actual wrapper "
    "mix. Rest-of-UK bands only (not Scotland). Tax bands and the State Pension are held in TODAY'S "
    "MONEY (assumed to rise with inflation) for the whole horizon, rather than literally freezing "
    "current nominal thresholds for 30 years - flagged clearly since current policy has in fact "
    "frozen the bands in nominal terms since 2021 (repeatedly extended), which would be an "
    "increasingly severe (and, over 30 years, implausible) real-terms tax rise if taken literally.",
    "  - Cash uses the real Bank of England SONIA rate (IUDSOIA, monthly-equivalent of the daily "
    "annualised rate) from 2014 onwards; pre-2014, SONIA wasn't yet the standard reference rate in this "
    "dataset, so a money-market fund proxy (Blackrock ICS Sterling Liquidity Fund) is spliced in, "
    "validated against the overlap period (99.7% correlation, ~0.003% mean difference).",
    "  - Annual model (spec's own suggestion) - the Python app runs monthly for finer sequencing detail. "
    "The Equity Sweep and Sensitivity Tables tabs use a single historical path per grid point (not a "
    "full Monte Carlo replication) for tractability - see the app for the full stochastic version.",
    "  - Forward-looking CMA figures (CMA tab) are a compiled median across published third-party 10-"
    "year forecasts (Monevator), not this firm's own house view - three of the eleven asset classes "
    "have no direct published match and are proxied from the closest available category (flagged on "
    "the CMA tab). Blending shifts each asset class's MEAN return only; it does not otherwise change "
    "the shape, fat tails or cross-asset correlation the historical bootstrap captures.",
    "  - Annuity rates (Annuity tab) are scaled examples from Hargreaves Lansdown's published best-buy "
    "table (14-28 May 2026), single-life/level/no-guarantee-period, cross-checked across two dates for "
    "consistency - NOT a personalised quote (real quotes vary by provider, postcode and health, and "
    "enhanced/impaired-life rates can be materially higher). No 25% tax-free lump sum is modelled "
    "before annuitizing. Joint-life pricing uses one flat discount factor (from a single age-65 data "
    "point), not a full joint-life age curve. Rates for ages outside 55-75 are clipped to the nearest "
    "quoted age rather than extrapolated.",
    "",
    "Data source: Bloomberg (via Ben Alfert), 9 July 2026; Bank of England SONIA (IUDSOIA), supplied by "
    "the user. For illustration purposes only - past performance is not a reliable guide to future returns.",
]
r = 4
for line in notes:
    ws.cell(row=r, column=2, value=line).font = SUBTITLE_FONT if line.startswith("Data source") else BLACK
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    if line.isupper() or line.endswith(":"):
        ws.cell(row=r, column=2).font = BOLD
    ws.row_dimensions[r].height = 14 if len(line) < 90 else 28
    r += 1

wb.save(OUT_PATH)
print("Patched Instructions sheet. Sheet order:", wb.sheetnames)
