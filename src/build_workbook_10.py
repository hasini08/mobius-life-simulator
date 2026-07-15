"""Stage 10: Asset Correlation sheet - monthly return correlation matrix across the 11 broad asset
classes (native CORREL() formulas against 'Asset Class Returns'), with a conditional-formatting
colour scale so it reads as a heatmap - the Excel-native complement to the app's correlation chart."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
ar = json.loads(Path("asset_returns_range.json").read_text())
ASSET_CLASSES = ar["asset_classes"]
afr, alr = ar["first_row"], ar["last_row"]
n_ac = len(ASSET_CLASSES)

if "Asset Correlation" in wb.sheetnames:
    del wb["Asset Correlation"]
ws = wb.create_sheet("Asset Correlation")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22
for i in range(n_ac):
    ws.column_dimensions[get_column_letter(3 + i)].width = 11

ws["B2"] = "How much do these investments actually move together? (Correlation Matrix)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: diversification - spreading money across different investment types to reduce "
    "risk - only really works between investments that DON'T move up and down together. 1.0 (dark "
    "red) = move almost perfectly together; 0 (white) = no relationship; negative (dark blue) = tend "
    "to move in opposite directions (the strongest diversification benefit). || Based on monthly "
    "returns across the 11 broad asset classes over the full history (1999/2000-2026). REITs and "
    "Infrastructure run ~0.75-0.76 correlated with Global Equities here, so they add less true "
    "diversification than their labels might suggest (see the 'Better' portfolio note in the "
    "Instructions tab)."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 60

header_row = 5
ws.cell(row=header_row, column=2, value="").font = HEADER_FONT
ws.cell(row=header_row, column=2).fill = HEADER_FILL
for i, ac in enumerate(ASSET_CLASSES):
    c = 3 + i
    cell = ws.cell(row=header_row, column=c, value=ac)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(text_rotation=45, horizontal="center")

first_data_row = header_row + 1
last_data_row = first_data_row + n_ac - 1
first_data_col = 3
last_data_col = 3 + n_ac - 1

for i, ac_row in enumerate(ASSET_CLASSES):
    r = first_data_row + i
    ws.cell(row=r, column=2, value=ac_row).font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    row_col_letter = get_column_letter(2 + i)  # column letter in Asset Class Returns for this row's asset class
    for j, ac_col in enumerate(ASSET_CLASSES):
        c = 3 + j
        col_col_letter = get_column_letter(2 + j)
        if i == j:
            formula = 1.0
        else:
            formula = (f"=CORREL('Asset Class Returns'!${row_col_letter}${afr}:${row_col_letter}${alr},"
                       f"'Asset Class Returns'!${col_col_letter}${afr}:${col_col_letter}${alr})")
        cell = ws.cell(row=r, column=c, value=formula)
        cell.number_format = "0.00"
        cell.font = BLACK

rule = ColorScaleRule(
    start_type="num", start_value=-1, start_color="D6604D",
    mid_type="num", mid_value=0, mid_color="FFFFFF",
    end_type="num", end_value=1, end_color="4393C3",
)
data_range = f"{get_column_letter(first_data_col)}{first_data_row}:{get_column_letter(last_data_col)}{last_data_row}"
ws.conditional_formatting.add(data_range, rule)

wb.save(OUT_PATH)
print("Saved stage 10 (Asset Correlation). Range:", data_range)
