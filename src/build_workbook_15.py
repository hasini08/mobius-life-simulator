"""Stage 15: patches the Historical Projection sheet in place (surgical edit, not a full sheet
rebuild) to apply the same tax/State Pension gross-up as the MC sheets when Inputs!apply_tax_on = Y -
keeps the deterministic historical check consistent with the stochastic engine rather than silently
ignoring the toggle. Column H ('Gross withdrawal from pot') drives the pot mechanics as before; a new
column J ('Net received, post-tax') reports what the client actually gets in hand - identical to H
when tax is off, but meaningfully less than H when it's on (H is grossed-up to cover the tax bill)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill
from portfolios import PORTFOLIOS

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
GBP = '£#,##0;(£#,##0);"-"'
FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")

wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
tax_deriv = json.loads(Path("tax_range.json").read_text())
blocks = json.loads(Path("historical_projection_blocks.json").read_text())
MAX_HORIZON = 30


def gross_for_net_formula(n_expr):
    d = tax_deriv
    n = f"MAX({n_expr},0)"
    return (f'IF({n}<={d["n1"]},{n},IF({n}<={d["n2"]},({n}-{d["seg2_i"]})/{d["seg2_s"]},'
            f'IF({n}<={d["n3"]},({n}-{d["seg3_i"]})/{d["seg3_s"]},'
            f'IF({n}<={d["n4"]},({n}-{d["seg4_i"]})/{d["seg4_s"]},({n}-{d["seg5_i"]})/{d["seg5_s"]}))))')


def tax_due_formula(x_expr):
    d = tax_deriv
    x = f"({x_expr})"
    return (f'IF({x}<={d["pa"]},0,IF({x}<={d["basic_limit"]},{d["basic_rate"]}*({x}-{d["pa"]}),'
            f'IF({x}<={d["taper_start"]},{d["tax_x2"]}+{d["higher_rate"]}*({x}-{d["basic_limit"]}),'
            f'IF({x}<={d["higher_limit"]},{d["tax_x3"]}+0.6*({x}-{d["taper_start"]}),'
            f'{d["tax_x4"]}+{d["add_rate"]}*({x}-{d["higher_limit"]})))))')


ws = wb["Historical Projection"]
ws.column_dimensions["J"].width = 15
# columns (see build_workbook_5.py): B=year#, C=cal year, D=return, E=inflation, F=cumInfl,
# G=real spend level, H=gross withdrawal from pot, I=pot, J=net received (NEW)
for name, header_row in blocks.items():
    ws.cell(row=header_row, column=8, value="Gross withdrawal from pot").font = HEADER_FONT
    ws.cell(row=header_row, column=8).fill = HEADER_FILL
    ws.cell(row=header_row, column=10, value="Net received (post-tax)").font = HEADER_FONT
    ws.cell(row=header_row, column=10).fill = HEADER_FILL

    first_data_row = header_row + 2  # header_row+1 = year-0 row, +2 = first y_i=1 data row
    for y_i in range(1, MAX_HORIZON + 1):
        row = first_data_row + y_i - 1
        prev = row - 1
        age_this_year = f"({NR['age']}+{y_i}-1)"
        sp_real = f'IF({age_this_year}>={NR["sp_age"]},{NR["sp_annual"]},0)'
        gross_real_target = f'MAX({gross_for_net_formula(f"G{row}")}-({sp_real}),0)'
        target_real = f'IF({NR["apply_tax_on"]}="Y",{gross_real_target},G{row})'
        spend_formula = f'=IF(D{row}="","",MIN(({target_real})*F{row},MAX(I{prev},0)))'
        ws.cell(row=row, column=8, value=spend_formula).number_format = GBP

        sp_nominal = f'({sp_real})*F{row}'
        total_gross_nominal = f'(H{row}+{sp_nominal})'
        net_formula = (f'=IF(D{row}="","",IF({NR["apply_tax_on"]}="Y",'
                        f'{total_gross_nominal}-({tax_due_formula(total_gross_nominal)}),H{row}))')
        ws.cell(row=row, column=10, value=net_formula).number_format = GBP

# Summary's historical "Any shortfall years?" check should look at NET RECEIVED (col J), not the
# gross pot withdrawal (col H) - identical when tax is off, but meaningfully different when it's on.
ws_summary = wb["Summary"]
hist_row_by_name = {}
for r in range(18, 26):
    nm = ws_summary.cell(row=r, column=2).value
    if nm in PORTFOLIOS:
        hist_row_by_name[nm] = r
for name, header_row in blocks.items():
    last_row = header_row + 1 + MAX_HORIZON
    spend_col_range = f"'Historical Projection'!J{header_row+2}:J{last_row}"
    r = hist_row_by_name[name]
    ws_summary.cell(row=r, column=4,
                     value=f'=IF(COUNTIF({spend_col_range},"<"&{NR["spend"]}*0.999)>0,"Yes","No")')

wb.save(OUT_PATH)
print("Saved stage 15 (Historical Projection tax/SP gross-up + net-received column).")
