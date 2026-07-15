"""Stage 16: Forward-looking Capital Market Assumptions (CMA) blending - the Excel-native mirror of
src/cma.py. Adds an Inputs blend-weight toggle, a new 'CMA' sheet holding the per-asset-class
forward-looking forecasts (editable) and the derived monthly shift each implies vs this workbook's
own historical sample, and patches the 'Portfolio Returns' (monthly) sheet in place so every monthly
return used everywhere downstream (Portfolio Annual Returns -> Historical Projection AND the MC
sheets, both of which read from Portfolio Annual Returns) is shifted consistently.

METHOD: portfolio_shift (monthly) = SUMPRODUCT(portfolio weights, per-asset-class monthly shifts).
Since SUMPRODUCT is linear, this is mathematically IDENTICAL to shifting each asset class's monthly
return first and then computing the weighted portfolio return - exactly what src/cma.py does for
the Python engine. Applied as ONE additive term inside 'Portfolio Returns', so no other sheet
(Historical Projection, MC *, Summary) needs to change."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS, AC
import cma as cma_mod

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
PCT1 = "0.0%"

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
ar = json.loads(Path("asset_returns_range.json").read_text())
wr_ranges = json.loads(Path("weight_rows.json").read_text())
pr_ranges = json.loads(Path("portfolio_returns_range.json").read_text())
holding_ranges = json.loads(Path("holding_ranges.json").read_text())

ASSET_CLASSES = ar["asset_classes"]  # fixed order, matches Asset Class Returns / Portfolios columns
n_ac = len(ASSET_CLASSES)
afr, alr = ar["first_row"], ar["last_row"]  # Asset Class Returns monthly data rows

# ---------------------------------------------------------------------
# Inputs: blend-weight toggle
# ---------------------------------------------------------------------
ws_in = wb["Inputs"]
existing_rows = [c.row for c in ws_in["B"] if c.value]
r0 = max(existing_rows) + 2
ws_in.cell(row=r0, column=2, value="Market Return Assumptions").font = BOLD
r = r0 + 1
ws_in.cell(row=r, column=2, value="Forward-looking blend (0% = pure history, 100% = pure forecast)").font = BLACK
cell = ws_in.cell(row=r, column=3, value=0.0)
cell.font = BLUE
cell.fill = INPUT_FILL
cell.number_format = PCT1
ws_in.cell(row=r, column=4, value=(
    "0% uses only the actual 2000-2026 historical returns (as everywhere else in this workbook by "
    "default). 100% recentres each asset class's AVERAGE return to current analyst 10-year forecasts "
    "(see the CMA sheet) while keeping history's volatility and correlation structure intact."
)).font = SUBTITLE_FONT
NR["cma_blend"] = f"Inputs!$C${r}"
Path("cellrefs.json").write_text(json.dumps(NR, indent=2))
print("Added cma_blend input:", NR["cma_blend"])

# ---------------------------------------------------------------------
# CMA sheet
# ---------------------------------------------------------------------
if "CMA" in wb.sheetnames:
    del wb["CMA"]
ws = wb.create_sheet("CMA")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
for i in range(n_ac):
    ws.column_dimensions[get_column_letter(2 + i)].width = 13

ws["B2"] = "How optimistic should our return assumptions be? (Forward-Looking Capital Market Assumptions)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: by default, this workbook assumes the future looks like 2000-2026, which was a "
    "strong run for stock markets. The 'Forward-looking blend' setting on the Inputs tab optionally "
    "makes the model more cautious instead, in line with what professional forecasters currently "
    "expect for the NEXT 10 years - a fairer, less rose-tinted test of the plan. Day-to-day ups and "
    "downs and worst-case scenarios still come from real market history either way; only the AVERAGE "
    "return assumption moves.\n\n"
    "The historical bootstrap used everywhere else in this workbook (2000-2026, ~26 years) is a real, "
    "unadjusted sample - but it is ONE window, and it happens to span an unusually strong run for "
    "global equities. The 'Forward-looking blend' input on the Inputs tab optionally recentres each "
    "asset class's AVERAGE monthly return towards independently published 10-year return forecasts, "
    "leaving month-to-month volatility, correlation and worst-case behaviour exactly as history shows "
    "them - it does not invent a new distribution, it only shifts where the historical one is centred. "
    "Forecasts below: Monevator's compilation of published 10-year GBP nominal forecasts (Vanguard, "
    "Schroders, JPMorgan, BlackRock and others), https://monevator.com/investment-return-forecasts/, "
    "accessed 2026. Three asset classes (marked *) have no direct published match and are proxied from "
    "the closest available category - see the note row below."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 160

header_row = 5
label_col_width = 46
ws.column_dimensions["A"].width = 2
ws.cell(row=header_row, column=1, value="").fill = HEADER_FILL
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    cell = ws.cell(row=header_row, column=c, value=ac + (" *" if ac in cma_mod.PROXIED_ASSET_CLASSES else ""))
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

row_labels = {}
r = header_row + 1
ws.cell(row=r, column=1, value="Forward-looking forecast (10yr), pa").font = BLACK
forecast_row = r
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    cell = ws.cell(row=r, column=c, value=cma_mod.CMA_ANNUAL.get(ac, 0.0))
    cell.font = BLUE
    cell.fill = INPUT_FILL
    cell.number_format = PCT
r += 1

ws.cell(row=r, column=1, value="Historical monthly mean (this workbook's sample)").font = BLACK
hist_monthly_row = r
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    col_letter = get_column_letter(c)
    formula = f"=AVERAGE('Asset Class Returns'!{col_letter}{afr}:{col_letter}{alr})"
    cell = ws.cell(row=r, column=c, value=formula)
    cell.number_format = PCT
    cell.font = GREEN
r += 1

ws.cell(row=r, column=1, value="Historical annual (implied), pa").font = BLACK
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    col_letter = get_column_letter(c)
    formula = f"=(1+{col_letter}{hist_monthly_row})^12-1"
    cell = ws.cell(row=r, column=c, value=formula)
    cell.number_format = PCT
    cell.font = GREEN
r += 1

ws.cell(row=r, column=1, value="CMA-implied monthly mean").font = BLACK
cma_monthly_row = r
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    col_letter = get_column_letter(c)
    formula = f"=(1+{col_letter}{forecast_row})^(1/12)-1"
    cell = ws.cell(row=r, column=c, value=formula)
    cell.number_format = PCT
    cell.font = GREEN
r += 1

ws.cell(row=r, column=1, value="Monthly shift (CMA - historical, full blend)").font = BOLD
shift_row = r
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    col_letter = get_column_letter(c)
    formula = f"={col_letter}{cma_monthly_row}-{col_letter}{hist_monthly_row}"
    cell = ws.cell(row=r, column=c, value=formula)
    cell.number_format = "0.0000"
    cell.font = BOLD
r += 2

ws.cell(row=r, column=1, value=(
    "* No published forward-looking forecast exists for this exact sub-category - proxied with the "
    "closest published category: UK Gilts 15yr+ uses the UK government bonds (all-stocks) figure; "
    "Securitised Credit uses the Global Bonds figure; Infrastructure uses the Global REITs figure "
    "(closest published 'real assets' category). See src/cma.py for the full reasoning."
)).font = SUBTITLE_FONT
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + n_ac - 1)
ws.row_dimensions[r].height = 55
r += 2

# ---------------------------------------------------------------------
# Portfolio-level monthly shift (SUMPRODUCT of weights x per-asset-class shifts) - one per portfolio
# ---------------------------------------------------------------------
ws.cell(row=r, column=1, value="Portfolio-level monthly shift (at 100% blend)").font = BOLD
r += 1
port_shift_header = r
ws.cell(row=r, column=1, value="Portfolio").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.cell(row=r, column=2, value="Monthly shift (full blend)").font = HEADER_FONT
ws.cell(row=r, column=2).fill = HEADER_FILL
r += 1
first_w_col_letter = get_column_letter(3)               # Portfolios weight roll-up starts at col C
last_w_col_letter = get_column_letter(2 + n_ac)          # ... through column M (n_ac=11)
first_shift_col_letter = get_column_letter(2)            # CMA shift row starts at col B
last_shift_col_letter = get_column_letter(1 + n_ac)       # ... through column L

portfolio_shift_cells = {}
for name in PORTFOLIOS:
    ws.cell(row=r, column=1, value=name).font = BOLD
    wrow = wr_ranges["rows"][name]
    formula = (f"=SUMPRODUCT(Portfolios!${first_w_col_letter}${wrow}:${last_w_col_letter}${wrow},"
               f"CMA!${first_shift_col_letter}${shift_row}:${last_shift_col_letter}${shift_row})")
    cell = ws.cell(row=r, column=2, value=formula)
    cell.number_format = "0.0000"
    cell.font = BOLD
    portfolio_shift_cells[name] = f"CMA!$B${r}"
    r += 1

ws.cell(row=r + 1, column=1, value=(
    "This is the SAME SUMPRODUCT-of-weights-and-shifts each portfolio's monthly return uses (Portfolio "
    "Returns sheet), scaled by the Inputs!blend setting - mathematically identical to shifting every "
    "asset class's own monthly return first and then computing the weighted portfolio return, since "
    "SUMPRODUCT is linear."
)).font = SUBTITLE_FONT
ws.cell(row=r + 1, column=1).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
ws.row_dimensions[r + 1].height = 40

cma_wb_range = {
    "asset_classes": ASSET_CLASSES, "forecast_row": forecast_row, "hist_monthly_row": hist_monthly_row,
    "cma_monthly_row": cma_monthly_row, "shift_row": shift_row, "portfolio_shift_cells": portfolio_shift_cells,
}
Path("cma_range.json").write_text(json.dumps(cma_wb_range, indent=2))

# ---------------------------------------------------------------------
# Patch 'Portfolio Returns' (monthly) in place: append the blend-weighted shift term to every cell.
# ---------------------------------------------------------------------
ws_pr = wb["Portfolio Returns"]
pr_first, pr_last = pr_ranges["first_row"], pr_ranges["last_row"]
port_cols = pr_ranges["port_cols"]
blend_ref = NR["cma_blend"]

for name in PORTFOLIOS:
    c = port_cols[name]
    col_letter = get_column_letter(c)
    fr, lr, tot_row = holding_ranges[name]
    wrow = wr_ranges["rows"][name]
    shift_cell = portfolio_shift_cells[name]
    for row_n in range(pr_first, pr_last + 1):
        src_row = afr + (row_n - pr_first)
        base = (f"SUMPRODUCT('Asset Class Returns'!${get_column_letter(2)}${src_row}:"
                f"${get_column_letter(1 + n_ac)}${src_row},"
                f"Portfolios!${first_w_col_letter}${wrow}:${last_w_col_letter}${wrow})"
                f"-Portfolios!$F${tot_row}/12")
        formula = f"={base}+{blend_ref}*{shift_cell}"
        cell = ws_pr.cell(row=row_n, column=c, value=formula)
        cell.number_format = PCT
        cell.font = GREEN

wb.save(OUT_PATH)
print("Saved stage 16 (CMA sheet + Inputs blend + Portfolio Returns patched).")
print("Portfolio shift cells:", portfolio_shift_cells)
