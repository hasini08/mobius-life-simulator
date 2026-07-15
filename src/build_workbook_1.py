"""Stage 1: Inputs, Portfolios (holdings + asset-class weight aggregation), Asset Class Returns
(monthly source data) and Annual Asset Returns (derived via formulas). Test recalculation before
adding the heavier sheets."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from portfolios import PORTFOLIOS, AC

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.0%"
PCT3 = "0.000%"
GBP = '£#,##0;(£#,##0);"-"'

wb = openpyxl.Workbook()
wb.remove(wb.active)

def style_header_row(ws, row, first_col, last_col):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

# ---------------------------------------------------------------------
# Instructions
# ---------------------------------------------------------------------
ws = wb.create_sheet("Instructions")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 110
ws["B2"] = "Mobius Wealth — Decumulation Simulator (v4, refined)"
ws["B2"].font = TITLE_FONT
notes = [
    "",
    "This workbook refines the previous Mobius decumulation model using the Bloomberg data supplied "
    "9 July 2026 and the FNZ 'Growth Passive Plus' holdings.",
    "",
    "WHAT'S NEW vs the previous model:",
    "  1. Compares THREE portfolio variants: 'Original' (mainstream retail fund lineup), "
    "'Alternative' (tax/cost-efficient unit-linked lineup - same underlying market exposure, lower fees), "
    "and 'Better' (a more diversified allocation adding REITs, infrastructure, commodities, EM and "
    "index-linked gilts, in the spirit of the previous model's winning portfolio).",
    "  2. Inflation-linked spending using UK CPI (Bloomberg data), not a fixed assumption.",
    "  3. Spending guardrails (Guyton-Klinger style) that cut spend in weak markets and raise it in "
    "strong ones - togglable on/off to see the impact on ruin probability vs shortfall frequency.",
    "  4. An accompanying Python/Streamlit app (see /app) adds an IMPROVED stochastic sampling engine "
    "(stationary block bootstrap + skewed-distribution option) beyond what a spreadsheet can practically "
    "do - this workbook's own Monte Carlo uses a simple annual bootstrap for a direct, auditable "
    "comparison to the previous model's methodology.",
    "",
    "STRUCTURE:",
    "  - Inputs: all client parameters and toggles (yellow cells).",
    "  - Portfolios: holdings, weights and fees for the three variants, sourced from the FNZ holdings "
    "file, with an asset-class weight roll-up used to drive the return calculations.",
    "  - Asset Class Returns: raw monthly total-return data by asset class (Bloomberg, to 9 July 2026).",
    "  - Annual Asset Returns: calendar-year compounded returns derived from the monthly data.",
    "  - Portfolio Annual Returns: each portfolio's net-of-fee annual return and the year's inflation.",
    "  - Historical Projection: a single deterministic path per portfolio using the ACTUAL historical "
    "sequence of annual returns, for a reasonableness check.",
    "  - MC Original / MC Alternative / MC Better: an annual bootstrap Monte Carlo (recalculates live - "
    "press F9 / Ctrl+Shift+F9 to redraw a new set of random paths).",
    "  - Summary: headline statistics (probability of ruin, legacy quantiles, shortfall years) for all "
    "three portfolios, historical and simulated, with the guardrails toggle applied throughout.",
    "",
    "ASSUMPTIONS TO CONFIRM (flagged in place, see cell comments too):",
    "  - Original-portfolio fund OCFs are NOT given in the source data (only Alternative AMCs are) - "
    "typical published OCFs for this fund type have been assumed. Confirm against real factsheets.",
    "  - The 'Better' portfolio's weights are a judgement-based construction, not sourced from FNZ data.",
    "  - Cash Plus and Four Seasons Fund (the other two FNZ portfolios) are not modelled - by "
    "instruction, since the comparison in scope is Original vs Alternative vs Better.",
    "  - No mortality, tax or state pension modelling - deliberately deferred per the project spec.",
    "  - Annual model (spec's own suggestion) - the Python app runs monthly for finer sequencing detail.",
    "",
    "Data source: Bloomberg (via Ben Alfert), 9 July 2026. For illustration purposes only - past "
    "performance is not a reliable guide to future returns.",
]
r = 4
for line in notes:
    ws.cell(row=r, column=2, value=line).font = SUBTITLE_FONT if line.startswith("Data source") else BLACK
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    if line.isupper() or line.endswith(":"):
        ws.cell(row=r, column=2).font = BOLD
    ws.row_dimensions[r].height = 14 if len(line) < 90 else 28
    r += 1

# ---------------------------------------------------------------------
# Inputs
# ---------------------------------------------------------------------
ws = wb.create_sheet("Inputs")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 60
ws["B2"] = "Client & Model Inputs"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Edit the yellow cells. All formulas throughout the workbook reference these."
ws["B3"].font = SUBTITLE_FONT

rows = [
    ("age",           "Starting age",                              65,    None, "Client's age today"),
    ("horizon",       "Time horizon (years)",                       30,    None, "Grid built for up to 30 years; do not exceed"),
    ("pot",           "Starting pot (£)",                       500000,    GBP,  "Total pension pot at retirement"),
    ("spend",         "Desired annual spend, today's £ (real terms)", 20000, GBP, "Before any guardrail adjustment"),
    ("wr0",           "Initial withdrawal rate",                  "=C8/C7", PCT, "Desired spend / starting pot (calculated)"),
    ("guardrails_on", "Apply spending guardrails? (Y/N)",           "N",    None, "Y = guardrails active in Historical Projection & MC sheets"),
    ("band",          "Guardrail band (± % of initial WR)",         0.20,   PCT, "Breach threshold either side of the initial withdrawal rate"),
    ("cut",           "Guardrail spend cut if above upper band",    0.10,   PCT, "Real-terms spend cut applied"),
    ("raise",         "Guardrail spend rise if below lower band",   0.10,   PCT, "Real-terms spend rise applied"),
    ("start_year",    "Historical projection start year",           2000,   None, "First full calendar year of data available is 2000"),
    ("n_sims",        "Number of Monte Carlo sims (Excel engine)",  600,    None, "Fixed grid size - see MC sheets. Python app allows more."),
]
r = 5
labels = {}
for key, label, val, fmt, note in rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    cell = ws.cell(row=r, column=3, value=val)
    if isinstance(val, str) and val.startswith("="):
        cell.font = BLACK
    else:
        cell.font = BLUE
        cell.fill = INPUT_FILL
    if fmt:
        cell.number_format = fmt
    ws.cell(row=r, column=4, value=note).font = SUBTITLE_FONT
    labels[key] = r
    r += 1

# named single-cell refs for readability elsewhere
NR = {key: f"Inputs!$C${row}" for key, row in labels.items()}
import json
Path("cellrefs.json").write_text(json.dumps(NR, indent=2))
print("Inputs rows:", labels)
print("NR:", NR)

wb.save("/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx")
print("Saved stage 1")
