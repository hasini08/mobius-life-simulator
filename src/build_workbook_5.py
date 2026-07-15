"""Stage 5: Historical Projection - deterministic year-by-year path per portfolio using the actual
historical sequence of annual returns from Inputs!start_year, with guardrails logic driven by the
Inputs sheet toggle. Mirrors the previous model's 'Model Hist' sheet, extended for inflation-linked
spend and guardrails."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
par = json.loads(Path("portfolio_annual_returns_range.json").read_text())
years = par["years"]
par_first, par_last = par["first_row"], par["last_row"]
pcols = par["port_cols"]
infl_col = par["infl_col"]

MAX_HORIZON = 30  # grid built for up to this many years (Inputs!horizon should not exceed this)

ws = wb.create_sheet("Historical Projection")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C4"
ws["A1"] = ("IN PLAIN TERMS: what would ACTUALLY have happened to this plan, replaying the one real "
            "sequence of market returns that occurred starting Inputs!start_year (2000-2026 available) "
            "- a concrete real-world example alongside the thousands of simulated futures elsewhere in "
            "this workbook, not itself a forecast. Guardrails apply if Inputs!guardrails_on = \"Y\".")
ws["A1"].font = SUBTITLE_FONT
ws["A1"].alignment = Alignment(wrap_text=True)
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 10
for col in "CDEFGHI":
    ws.column_dimensions[col].width = 15

row = 3
block_start_rows = {}
for name in PORTFOLIOS:
    ws.cell(row=row, column=2, value=f"{name} portfolio").font = BOLD
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = SUBHEAD_FILL
    row += 1
    headers = ["Year #", "Calendar year", "Return used", "Inflation used", "Cumulative inflation index",
               "Real spend level", "Spend taken", "End-of-year value"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=2 + i, value=h).font = HEADER_FONT
        ws.cell(row=row, column=2 + i).fill = HEADER_FILL
    header_row = row
    row += 1
    block_start_rows[name] = header_row
    # Columns: B=year#, C=calendar year, D=return, E=inflation, F=cumulative inflation index,
    #          G=real spend level (today's money, only moved by guardrails), H=spend taken, I=pot
    c_year_num, c_cal_year, c_ret, c_infl, c_cuminfl, c_realspend, c_spend, c_pot = range(2, 10)

    # Year 0 row: initial pot, no spend yet, cumulative inflation index = 1.0
    ws.cell(row=row, column=c_year_num, value=0)
    ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}-1")
    ws.cell(row=row, column=c_cuminfl, value=1.0).number_format = "0.0000"
    ws.cell(row=row, column=c_realspend, value=f"={NR['spend']}").number_format = GBP
    ws.cell(row=row, column=c_pot, value=f"={NR['pot']}").number_format = GBP
    row += 1
    first_data_row = row

    for y_i in range(1, MAX_HORIZON + 1):
        prev = row - 1
        ws.cell(row=row, column=c_year_num, value=y_i)
        ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}+{y_i}-1")
        # only compute while within horizon AND within available historical years
        # (C{row} holds this row's calendar year; D{row} is the return cell being computed here)
        avail_check = f"AND(B{row}<={NR['horizon']},COUNTIF('Portfolio Annual Returns'!$A${par_first}:$A${par_last},C{row})>0)"
        ret_formula = (f'=IF({avail_check},'
                        f"INDEX('Portfolio Annual Returns'!${get_column_letter(pcols[name])}${par_first}:"
                        f"${get_column_letter(pcols[name])}${par_last},"
                        f"MATCH(C{row},'Portfolio Annual Returns'!$A${par_first}:$A${par_last},0)),\"\")")
        ws.cell(row=row, column=c_ret, value=ret_formula).number_format = PCT
        infl_formula = (f'=IF(D{row}="","",IF({avail_check},'
                         f"INDEX('Portfolio Annual Returns'!${get_column_letter(infl_col)}${par_first}:"
                         f"${get_column_letter(infl_col)}${par_last},"
                         f"MATCH(C{row},'Portfolio Annual Returns'!$A${par_first}:$A${par_last},0)),\"\"))")
        ws.cell(row=row, column=c_infl, value=infl_formula).number_format = PCT

        # cumulative inflation index compounds year over year (matches the Python engine's cum_inflation)
        cuminfl_formula = f'=IF(D{row}="",F{prev},F{prev}*(1+E{row}))'
        ws.cell(row=row, column=c_cuminfl, value=cuminfl_formula).number_format = "0.0000"

        # real spend level (today's money): carries forward, only moved by guardrails when triggered
        if y_i == 1:
            real_spend_formula = f"={NR['spend']}"
        else:
            wr0 = NR["wr0"]
            prior_realspend = f"G{prev}"
            # withdrawal-rate check uses THIS year's nominal target (real_spend x cumulative inflation
            # index, col F) vs prior year-end pot (I{prev})
            target_nominal = f"({prior_realspend}*F{row})"
            wr_now = f"IF(I{prev}>0,{target_nominal}/I{prev},9^9)"
            # capped to +/-50% of the ORIGINAL desired spend so repeated triggers can't compound
            # without bound (matches the Python engine's guardrail cap)
            raw_adj = (
                f'IF({wr_now}>{wr0}*(1+{NR["band"]}),{prior_realspend}*(1-{NR["cut"]}),'
                f'IF({wr_now}<{wr0}*(1-{NR["band"]}),{prior_realspend}*(1+{NR["raise"]}),{prior_realspend}))'
            )
            real_spend_formula = (
                f'=IF(D{row}="",{prior_realspend},'
                f'IF({NR["guardrails_on"]}<>"Y",{prior_realspend},'
                f'MIN(MAX({raw_adj},0.5*{NR["spend"]}),1.5*{NR["spend"]})))'
            )
        ws.cell(row=row, column=c_realspend, value=real_spend_formula).number_format = GBP

        # spend taken this year = min(real spend level x cumulative inflation index, prior pot), blank if beyond horizon
        spend_formula = f'=IF(D{row}="","",MIN(G{row}*F{row},MAX(I{prev},0)))'
        ws.cell(row=row, column=c_spend, value=spend_formula).number_format = GBP

        pot_formula = f'=IF(D{row}="",I{prev},MAX(I{prev}-H{row},0)*(1+D{row}))'
        ws.cell(row=row, column=c_pot, value=pot_formula).number_format = GBP
        row += 1
    last_data_row = row - 1
    row += 2

wb.save(OUT_PATH)
Path("historical_projection_blocks.json").write_text(json.dumps(block_start_rows))
print("Saved stage 5 (Historical Projection). Header rows:", block_start_rows)
