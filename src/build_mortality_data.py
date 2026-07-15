"""Extracts the user-supplied S4 mortality table (Male_Data / Female_Data sheets, Age + q_x, ages
20-120) into a clean data/mortality_qx.csv for the engine and workbook to consume.

S4 is one of the CMI's (Continuous Mortality Investigation) standard tables derived from UK
self-administered pension scheme experience - a better basis for a PENSION decumulation model than a
general-population table (ONS National Life Tables), since pension scheme members' mortality
experience differs systematically from the population at large (this is the standard reason UK
pension actuaries use S-series/PxMA/PxFA tables rather than ONS tables for this kind of work)."""
import openpyxl
import pandas as pd
from pathlib import Path

SRC = "/root/.claude/uploads/6277f0d3-e8e6-5e30-977d-0317d499a601/6ebe1ca5-Python_Decumulation_Mortality_Data.xlsx"
DATA = Path(__file__).resolve().parent.parent / "data"


def load_sheet(ws):
    ages, qx = [], []
    for r in range(4, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        q = ws.cell(row=r, column=2).value
        if a is None or q is None:
            continue
        ages.append(int(a))
        qx.append(float(q))
    return pd.Series(qx, index=pd.Index(ages, name="age"))


def build_and_save():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    male = load_sheet(wb["Male_Data"])
    female = load_sheet(wb["Female_Data"])

    df = pd.DataFrame({"qx_male": male, "qx_female": female})
    df.index.name = "age"
    assert df.index.min() == 20 and df.index.max() == 120, f"unexpected age range: {df.index.min()}-{df.index.max()}"
    assert (df["qx_male"] <= 1).all() and (df["qx_male"] >= 0).all(), "male qx out of [0,1] range"
    assert (df["qx_female"] <= 1).all() and (df["qx_female"] >= 0).all(), "female qx out of [0,1] range"
    assert df.loc[120, "qx_male"] == 1 and df.loc[120, "qx_female"] == 1, "expected qx=1 at closing age 120"

    out_path = DATA / "mortality_qx.csv"
    df.to_csv(out_path)
    print(f"Saved {len(df)} ages ({df.index.min()}-{df.index.max()}) to {out_path}")
    print(df.head())
    print(df.tail())
    return df


if __name__ == "__main__":
    build_and_save()
