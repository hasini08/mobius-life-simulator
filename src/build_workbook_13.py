"""Stage 13: Summary - adds a 'Mortality-adjusted results' block (probability of ruin BEFORE death vs
the raw horizon-end figure, probability of surviving the full horizon, legacy-at-death quantiles, life
expectancy), driven by the Inputs!sex/joint_life_on/partner toggles and the Mortality tab's survival
curve. Appended below the existing content (and the floating bar chart) so nothing already on the
sheet has to be reflowed."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.0%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
mort = json.loads(Path("mortality_range.json").read_text())
ws = wb["Summary"]

r = 40
ws.cell(row=r, column=2, value="Mortality-adjusted results (Excel bootstrap engine)").font = BOLD
r += 1
ws.cell(row=r, column=2, value=(
    '="Life basis: "&IF('+NR["joint_life_on"]+'="Y","Joint life ("&'+NR["sex"]+'&" age "&'+NR["age"]
    + '&" / "&'+NR["partner_sex"]+'&" age "&'+NR["partner_age"]+'&")","Single life ("&'+NR["sex"]
    + '&", age "&'+NR["age"]+'&")")'
))
ws.cell(row=r, column=2).font = SUBTITLE_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 1
ws.cell(row=r, column=2, value=(
    "'Ruin before death' = the pot hits zero while the client (or, for joint life, at least one "
    "partner) is still alive - computed EXACTLY off each simulated path's ruin year and the Mortality "
    "tab's survival curve (no extra sampling noise from mortality itself, only from the market MC). "
    "'Legacy at death' values the estate at a randomly-sampled death year per path rather than at a "
    "fixed year-30 cutoff, so it DOES carry its own sampling noise on top of the market MC."
)).font = SUBTITLE_FONT
ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
ws.row_dimensions[r].height = 40
r += 2

mort_header_row = r
headers = ["Portfolio", "Prob. ruin before death", "Prob. ruin by horizon end (no mortality)",
           "Prob. survive full horizon (exact)", "Median legacy at death", "5th pctl legacy at death",
           "95th pctl legacy at death"]
for i, h in enumerate(headers):
    cell = ws.cell(row=r, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws.column_dimensions[get_column_letter(2 + i)].width = 17
r += 1
mort_data_first_row = r
for name in PORTFOLIOS:
    mc = json.loads(Path(f"mc_range_{name}.json").read_text())
    fr, lr = mc["first_row"], mc["last_row"]
    palive_l = get_column_letter(mc["palive_col"])
    legacyatdeath_l = get_column_letter(mc["legacyatdeath_col"])
    sheet = f"'MC {name}'"
    ws.cell(row=r, column=2, value=name).font = BOLD
    ws.cell(row=r, column=3, value=f"=AVERAGE({sheet}!{palive_l}{fr}:{palive_l}{lr})").number_format = PCT
    # column 4 (prob. ruin by horizon end, no mortality) is filled in below, once we know which
    # existing MC-results row (12-16 block above) corresponds to this portfolio
    ws.cell(row=r, column=4).number_format = PCT
    ws.cell(row=r, column=5, value=f"=Mortality!$I${mort['sc_last_row']}").number_format = PCT
    ws.cell(row=r, column=6, value=f"=MEDIAN({sheet}!{legacyatdeath_l}{fr}:{legacyatdeath_l}{lr})").number_format = GBP
    ws.cell(row=r, column=7, value=f"=PERCENTILE({sheet}!{legacyatdeath_l}{fr}:{legacyatdeath_l}{lr},0.05)").number_format = GBP
    ws.cell(row=r, column=8, value=f"=PERCENTILE({sheet}!{legacyatdeath_l}{fr}:{legacyatdeath_l}{lr},0.95)").number_format = GBP
    r += 1
mort_data_last_row = r - 1

# fix column 4 (prob. ruin by horizon end, no mortality) to reference the EXISTING MC results block
# above (rows 14-16, column C) by portfolio name rather than a fragile row-offset guess
existing_mc_rows = {}
for rr in range(13, 17):
    nm = ws.cell(row=rr, column=2).value
    if nm in PORTFOLIOS:
        existing_mc_rows[nm] = rr
for i, name in enumerate(PORTFOLIOS):
    rr = mort_data_first_row + i
    ws.cell(row=rr, column=4, value=f"=C{existing_mc_rows[name]}").number_format = PCT

r += 1
ws.cell(row=r, column=2, value="Life expectancy (own life / partner, curtate, years - see Mortality tab)").font = BOLD
r += 1
ws.cell(row=r, column=2, value="Own life").font = BLACK
ws.cell(row=r, column=3, value="=Mortality!$I$109").number_format = "0.0"
ws.cell(row=r, column=4, value="Partner").font = BLACK
ws.cell(row=r, column=5, value=f'=IF({NR["joint_life_on"]}="Y",Mortality!$I$110,"n/a")')

wb.save(OUT_PATH)
print("Saved stage 13 (Summary mortality block). Rows", mort_data_first_row, "-", mort_data_last_row)
