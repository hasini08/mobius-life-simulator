"""Stage 6: Monte Carlo sheets (one per portfolio) - simple annual bootstrap (each simulated year
independently samples one historical calendar year's return+inflation pair with replacement), at
Inputs!n_sims x MAX_HORIZON scale. This is the Excel-native complement to the Python app's improved
stochastic methods (stationary block bootstrap / skew-t) - see Instructions.

Column layout is 6 cols/year: RandIdx, CumInfl, RealSpend (NET, today's-money, guardrail-adjusted -
materialised as its own column so (a) the guardrail recursion references the PRIOR year's real spend
directly rather than back-deriving it from nominal spend/cumInfl, which would break once tax grosses
withdrawals up above the net target, and (b) the tax gross-up formula only ever needs to reference a
single cheap cell, not re-embed the whole guardrail expression), Target (NOMINAL, the intended
withdrawal from the pot BEFORE it's capped by whatever's actually left - grossed up for tax net of
State Pension and any annuity income when relevant - ALSO materialised as its own column, critically
so that the Shortfall check (see below) only ever references a cheap cell too), Withdrawal (nominal,
what actually leaves the pot = MIN(Target, available pot)), Pot.

WHY A SEPARATE 'TARGET' COLUMN (added when annuitization was wired in): the Shortfall column used to
recompute each year's intended target INLINE, concatenating 30 years' worth of the tax/annuity gross-
up expression into one cell - each year's expression is a few hundred characters, so the combined
Shortfall formula ballooned to ~12,000-13,000 characters. A single cell that long recalculates fine in
isolation, but multiplied across 600 sims x 3 portfolios x a fully-populated workbook, LibreOffice's
headless recalc pathologically stalled (confirmed empirically: removing even ONE of the 1,800 giant
Shortfall cells was enough to make the whole workbook recalculate in under a minute; with all 1,800
present it never completed even after 30+ minutes) - almost certainly a non-linear blow-up in how the
recalc engine handles very many very-long interdependent formulas together, not any single cell being
individually invalid. Materialising Target as its own column keeps every cell's formula short (a few
hundred characters at most) and turns Shortfall into a cheap sum of cell-to-cell comparisons."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
TEST_MODE = "--test" in sys.argv
N_SIMS = 5 if TEST_MODE else 600
MAX_HORIZON = 5 if TEST_MODE else 30

wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
par = json.loads(Path("portfolio_annual_returns_range.json").read_text())
years = par["years"]
par_first, par_last = par["first_row"], par["last_row"]
pcols = par["port_cols"]
infl_col = par["infl_col"]
n_years_avail = len(years)
tax_deriv = json.loads(Path("tax_range.json").read_text())

# Annuity income offset is OPTIONAL at this stage (stage 17 may not have run yet on a fresh rebuild) -
# defaults to "no annuity" (income=0, remaining pot=full pot) so this stays a no-op when absent,
# exactly like the tax toggle already does when apply_tax_on="N".
ann_path = Path("annuity_range.json")
if ann_path.exists():
    ann_range = json.loads(ann_path.read_text())
    ANNUITY_INCOME = ann_range["income_cell"]
    ANNUITY_REMAINING_POT = ann_range["remaining_pot_cell"]
else:
    ANNUITY_INCOME = "0"
    ANNUITY_REMAINING_POT = NR["pot"]


def gross_for_net_formula(n_expr):
    """Excel formula fragment for tax.gross_for_net() - see build_workbook_14.py / src/tax.py for the
    derivation. n_expr should be a CHEAP expression (ideally a single cell ref) since it's repeated
    ~9x in the nested IF."""
    d = tax_deriv
    n = f"MAX({n_expr},0)"
    return (f'IF({n}<={d["n1"]},{n},IF({n}<={d["n2"]},({n}-{d["seg2_i"]})/{d["seg2_s"]},'
            f'IF({n}<={d["n3"]},({n}-{d["seg3_i"]})/{d["seg3_s"]},'
            f'IF({n}<={d["n4"]},({n}-{d["seg4_i"]})/{d["seg4_s"]},({n}-{d["seg5_i"]})/{d["seg5_s"]}))))')


def target_real_expr(realspend_expr, sp_real_expr, annuity_real_expr):
    """Real-terms (today's-money) withdrawal target still needed from the pot: annuity income offsets
    the target REGARDLESS of the tax toggle (mirrors engine._gross_withdrawal_target() exactly); State
    Pension only enters via the tax gross-up branch."""
    gross_target = f'MAX({gross_for_net_formula(realspend_expr)}-({sp_real_expr})-({annuity_real_expr}),0)'
    no_tax_target = f'MAX(({realspend_expr})-({annuity_real_expr}),0)'
    return f'IF({NR["apply_tax_on"]}="Y",{gross_target},{no_tax_target})'


for name in PORTFOLIOS:
    sheet_name = f"MC {name}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws["A1"] = (f"IN PLAIN TERMS: each row below is one possible future for the '{name}' portfolio - "
                f"a full {MAX_HORIZON}-year run built by repeatedly picking a real historical year's "
                f"market return at random. {N_SIMS:,} rows = {N_SIMS:,} different possible futures "
                f"tested at once. 'Ruined? (1/0)' flags a row where the pot ran out; 'Shortfall years' "
                f"counts years spending had to be cut short of the target. Press Ctrl+Alt+F9 to redraw "
                f"a fresh set of random futures. || Technical detail: annual bootstrap Monte Carlo - "
                f"each simulated year independently samples one of the {n_years_avail} historical "
                f"calendar years' (return, inflation) pair with replacement. Guardrails apply if "
                f"Inputs!guardrails_on = \"Y\"; tax/State Pension apply if Inputs!apply_tax_on = \"Y\"; "
                f"annuitization applies if Inputs!annuity_pct > 0 (see Tax / Annuity tabs).")
    ws["A1"].font = SUBTITLE_FONT
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 60

    ws.cell(row=2, column=1, value="Sim #").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.column_dimensions["A"].width = 8
    col = 2
    year_col_starts = {}
    COLS_PER_YEAR = 6
    for y_i in range(1, MAX_HORIZON + 1):
        year_col_starts[y_i] = col
        headers = [f"Y{y_i} idx", f"Y{y_i} cumInfl", f"Y{y_i} realSpend", f"Y{y_i} target",
                   f"Y{y_i} withdrawal", f"Y{y_i} pot"]
        for j, h in enumerate(headers):
            c = col + j
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(c)].width = 11
        col += COLS_PER_YEAR
    legacy_col = col
    ruin_col = col + 1
    shortfall_col = col + 2
    for c, h in [(legacy_col, "Legacy (final pot)"), (ruin_col, "Ruined? (1/0)"), (shortfall_col, "Shortfall years")]:
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = 14

    ret_col_letter = get_column_letter(pcols[name])
    infl_col_letter = get_column_letter(infl_col)

    for s in range(1, N_SIMS + 1):
        row = 2 + s
        ws.cell(row=row, column=1, value=s)
        for y_i in range(1, MAX_HORIZON + 1):
            c0 = year_col_starts[y_i]
            c_idx, c_cuminfl, c_realspend, c_target, c_withdrawal, c_pot = (
                c0, c0 + 1, c0 + 2, c0 + 3, c0 + 4, c0 + 5)
            idx_letter = get_column_letter(c_idx)
            cuminfl_letter = get_column_letter(c_cuminfl)
            realspend_letter = get_column_letter(c_realspend)
            target_letter = get_column_letter(c_target)
            withdrawal_letter = get_column_letter(c_withdrawal)
            pot_letter = get_column_letter(c_pot)

            in_horizon = f"{y_i}<={NR['horizon']}"
            idx_formula = f'=IF({in_horizon},RANDBETWEEN(1,{n_years_avail}),"")'
            ws.cell(row=row, column=c_idx, value=idx_formula)

            ret_ref = f"INDEX('Portfolio Annual Returns'!${ret_col_letter}${par_first}:${ret_col_letter}${par_last},{idx_letter}{row})"
            infl_ref = f"INDEX('Portfolio Annual Returns'!${infl_col_letter}${par_first}:${infl_col_letter}${par_last},{idx_letter}{row})"

            prior_cuminfl = "1" if y_i == 1 else f"{get_column_letter(year_col_starts[y_i-1]+1)}{row}"
            cuminfl_formula = f'=IF({idx_letter}{row}="",{prior_cuminfl},{prior_cuminfl}*(1+{infl_ref}))'
            ws.cell(row=row, column=c_cuminfl, value=cuminfl_formula).number_format = "0.0000"

            prior_pot_ref = ANNUITY_REMAINING_POT if y_i == 1 else f"{get_column_letter(year_col_starts[y_i-1]+5)}{row}"

            # --- Real spend (NET, today's money) - materialised each year, guardrail-adjusted off the
            # PRIOR year's real spend directly (not back-derived from nominal figures, which tax would
            # otherwise corrupt) ---
            if y_i == 1:
                realspend_formula = f"={NR['spend']}"
            else:
                prior_real_ref = f"{get_column_letter(year_col_starts[y_i-1]+2)}{row}"
                wr_now = f"IF({prior_pot_ref}>0,({prior_real_ref}*{cuminfl_letter}{row})/{prior_pot_ref},9^9)"
                raw_adj = (f'IF({wr_now}>{NR["wr0"]}*(1+{NR["band"]}),({prior_real_ref})*(1-{NR["cut"]}),'
                           f'IF({wr_now}<{NR["wr0"]}*(1-{NR["band"]}),({prior_real_ref})*(1+{NR["raise"]}),{prior_real_ref}))')
                realspend_formula = (f'=IF({NR["guardrails_on"]}<>"Y",{prior_real_ref},'
                                      f'MIN(MAX({raw_adj},0.5*{NR["spend"]}),1.5*{NR["spend"]}))')
            ws.cell(row=row, column=c_realspend, value=realspend_formula).number_format = GBP

            # --- Target (nominal, MATERIALISED): the intended withdrawal from the pot before capping -
            # real target x cumInfl when tax is off; grossed up for tax net of State Pension (real
            # terms, then reflated) when tax is on; annuity income (level, nominal) offsets it either
            # way - see target_real_expr(). This is the ONLY cell that embeds the tax/annuity gross-up
            # expression each year (not repeated again for Withdrawal or Shortfall below). ---
            age_this_year = f"({NR['age']}+{y_i}-1)"
            sp_real = f'IF({age_this_year}>={NR["sp_age"]},{NR["sp_annual"]},0)'
            annuity_real = f'{ANNUITY_INCOME}/{cuminfl_letter}{row}'
            target_real = target_real_expr(f"{realspend_letter}{row}", sp_real, annuity_real)
            target_formula = f'=IF({idx_letter}{row}="","",({target_real})*{cuminfl_letter}{row})'
            ws.cell(row=row, column=c_target, value=target_formula).number_format = GBP

            # --- Withdrawal (nominal, what actually leaves the pot) - CHEAP now: just caps the
            # materialised Target against what's actually available. ---
            withdrawal_formula = f'=IF({idx_letter}{row}="","",MIN({target_letter}{row},MAX({prior_pot_ref},0)))'
            ws.cell(row=row, column=c_withdrawal, value=withdrawal_formula).number_format = GBP

            pot_formula = f'=IF({idx_letter}{row}="",{prior_pot_ref},MAX({prior_pot_ref}-{withdrawal_letter}{row},0)*(1+{ret_ref}))'
            ws.cell(row=row, column=c_pot, value=pot_formula).number_format = GBP

        last_pot_letter = get_column_letter(year_col_starts[MAX_HORIZON] + 5)
        ws.cell(row=row, column=legacy_col, value=f"={last_pot_letter}{row}").number_format = GBP
        pot_cols_letters = [get_column_letter(year_col_starts[y] + 5) for y in range(1, MAX_HORIZON + 1)]
        min_expr = "MIN(" + ",".join(f"{pl}{row}" for pl in pot_cols_letters) + ")"
        ws.cell(row=row, column=ruin_col, value=f"=IF({min_expr}<=0.01,1,0)")
        # Shortfall: withdrawal fell short of the (materialised, cheap-to-reference) Target for that
        # year - a plain sum of 30 cell-to-cell comparisons, NOT a recomputation of the tax/annuity
        # expression (see module docstring for why that mattered).
        withdrawal_cols_letters = [get_column_letter(year_col_starts[y] + 4) for y in range(1, MAX_HORIZON + 1)]
        target_cols_letters = [get_column_letter(year_col_starts[y] + 3) for y in range(1, MAX_HORIZON + 1)]
        shortfall_terms = [f'IF({wl}{row}<{tl}{row}*0.999,1,0)'
                            for wl, tl in zip(withdrawal_cols_letters, target_cols_letters)]
        ws.cell(row=row, column=shortfall_col, value="=" + "+".join(shortfall_terms))

    Path(f"mc_range_{name}.json").write_text(json.dumps({
        "first_row": 3, "last_row": 2 + N_SIMS, "legacy_col": legacy_col, "ruin_col": ruin_col,
        "shortfall_col": shortfall_col, "n_sims": N_SIMS, "max_horizon": MAX_HORIZON,
        "cols_per_year": COLS_PER_YEAR, "year_col_starts": year_col_starts,
    }))
    print(f"Built MC sheet for {name}: {N_SIMS} sims x {MAX_HORIZON} years, legacy_col={legacy_col}")

wb.save(OUT_PATH)
print("Saved stage 6.")
