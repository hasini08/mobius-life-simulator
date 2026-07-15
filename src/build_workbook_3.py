"""Stage 3: add UK CPI YoY as an extra column on 'Asset Class Returns', then build
'Annual Asset Returns' (derived from monthly via formulas)."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
ar = json.loads(Path("asset_returns_range.json").read_text())
ASSET_CLASSES = ar["asset_classes"]
first_row = ar["first_row"]

ac_raw = pd.read_csv("/home/claude/mobius_decumulation/data/asset_class_returns.csv", index_col=0, parse_dates=True)
cpi_col_name = [c for c in ac_raw.columns if "CPI" in c][0]
cpi_series = ac_raw[cpi_col_name]
dates = ac_raw.index

# ---------------------------------------------------------------------
# Add CPI column to 'Asset Class Returns'
# ---------------------------------------------------------------------
ws = wb["Asset Class Returns"]
cpi_col_excel = 2 + len(ASSET_CLASSES)  # next free column
cell = ws.cell(row=2, column=cpi_col_excel, value="UK CPI YoY (level, not a return)")
cell.font = HEADER_FONT
cell.fill = HEADER_FILL
ws.column_dimensions[get_column_letter(cpi_col_excel)].width = 16
for i, dt in enumerate(dates):
    row_n = first_row + i
    v = cpi_series.iloc[i]
    if pd.notna(v):
        c = ws.cell(row=row_n, column=cpi_col_excel, value=float(v))
        c.font = BLUE
        c.number_format = PCT

# ---------------------------------------------------------------------
# Year -> row-range map (asset returns rows) and last-CPI-row-per-year
# ---------------------------------------------------------------------
year_rows = {}
for i, dt in enumerate(dates):
    y = dt.year
    excel_row = first_row + i
    year_rows.setdefault(y, [excel_row, excel_row])
    year_rows[y][1] = excel_row
years = sorted(y for y in year_rows if y >= 2000)

cpi_last_row_by_year = {}
for i, dt in enumerate(dates):
    if pd.notna(cpi_series.iloc[i]):
        cpi_last_row_by_year[dt.year] = first_row + i

# ---------------------------------------------------------------------
# Annual Asset Returns sheet
# ---------------------------------------------------------------------
ws2 = wb.create_sheet("Annual Asset Returns")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B3"
ws2["A1"] = ("Calendar-year compounded returns, derived from monthly data (PRODUCT of (1+monthly)-1). "
             "2026 is partial (Jan-Jun only). UK CPI YoY shown is the latest available month's YoY reading that year.")
ws2["A1"].font = SUBTITLE_FONT
ws2.cell(row=2, column=1, value="Year").font = HEADER_FONT
ws2.cell(row=2, column=1).fill = HEADER_FILL
ws2.column_dimensions["A"].width = 10
for i, ac in enumerate(ASSET_CLASSES):
    c = 2 + i
    cell = ws2.cell(row=2, column=c, value=ac)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws2.column_dimensions[get_column_letter(c)].width = 13
cpi_annual_col = 2 + len(ASSET_CLASSES)
cell = ws2.cell(row=2, column=cpi_annual_col, value="UK CPI YoY (inflation)")
cell.font = HEADER_FONT
cell.fill = HEADER_FILL
ws2.column_dimensions[get_column_letter(cpi_annual_col)].width = 18

for r_i, y in enumerate(years):
    row_n = 3 + r_i
    ws2.cell(row=row_n, column=1, value=y).font = BOLD
    fr, lr = year_rows[y]
    for i, ac in enumerate(ASSET_CLASSES):
        c = 2 + i
        col_letter = get_column_letter(c)
        formula = f"=EXP(SUMPRODUCT(LN(1+'Asset Class Returns'!{col_letter}{fr}:{col_letter}{lr})))-1"
        cell = ws2.cell(row=row_n, column=c, value=formula)
        cell.number_format = PCT
        cell.font = GREEN
    cpi_row = cpi_last_row_by_year.get(y, lr)
    cpi_col_letter = get_column_letter(cpi_col_excel)
    formula = f"='Asset Class Returns'!{cpi_col_letter}{cpi_row}"
    cell = ws2.cell(row=row_n, column=cpi_annual_col, value=formula)
    cell.number_format = PCT
    cell.font = GREEN

annual_first_row, annual_last_row = 3, 3 + len(years) - 1
Path("annual_returns_range.json").write_text(json.dumps({
    "first_row": annual_first_row, "last_row": annual_last_row, "years": years,
    "cpi_annual_col": cpi_annual_col, "asset_classes": ASSET_CLASSES
}))

wb.save(OUT_PATH)
print("Saved stage 3. Years:", years[0], "..", years[-1], f"({len(years)} rows)")
