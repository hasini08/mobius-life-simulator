"""Stage 9: Sensitivity Tables sheet - deterministic historical single-path sensitivity of the final
legacy value / shortfall flag to (a) the initial withdrawal rate and (b) the guardrail band width,
for each of the three portfolios. Equity-weight sensitivity is already covered by the 'Equity Sweep'
sheet, so this sheet completes the "sensitivity to the client's own levers" picture. Same tractability
choice as 'Equity Sweep': one historical path per grid point rather than a full Monte Carlo replication
- see the Python app's 'Sensitivity analysis' section for the full stochastic version."""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portfolios import PORTFOLIOS

FONT = "Arial"
BLACK = Font(name=FONT, color="000000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
PCT = "0.00%"
GBP = '£#,##0;(£#,##0);"-"'

OUT_PATH = "/home/claude/mobius_decumulation/output/Mobius_Wealth_Decumulation_Model_v4.xlsx"
wb = openpyxl.load_workbook(OUT_PATH)
NR = json.loads(Path("cellrefs.json").read_text())
par = json.loads(Path("portfolio_annual_returns_range.json").read_text())
years = par["years"]
par_first, par_last = par["first_row"], par["last_row"]
pcols = par["port_cols"]
infl_col = par["infl_col"]

MAX_HORIZON = 30
WR_GRID = [round(0.02 + 0.005 * i, 4) for i in range(11)]   # 2.0% .. 7.0%, step 0.5%
BAND_GRID = [round(0.05 + 0.05 * i, 4) for i in range(8)]   # 5% .. 40%, step 5%

if "Sensitivity Tables" in wb.sheetnames:
    del wb["Sensitivity Tables"]
ws = wb.create_sheet("Sensitivity Tables")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 34

ws["B2"] = "Which decisions actually move the needle? (Sensitivity Tables)"
ws["B2"].font = TITLE_FONT
ws["B3"] = (
    "IN PLAIN TERMS: tests, one at a time, the two things a client and adviser can actually control "
    "day-to-day - spending level and (if used) how sensitive the guardrails are - to see how much "
    "each one changes the outcome. || Technical detail: deterministic historical single-path "
    "sensitivity (same method as 'Historical Projection'). Equity weight sensitivity is on the "
    "'Equity Sweep' tab. See the Python app for the full Monte Carlo version of these sweeps."
)
ws["B3"].font = SUBTITLE_FONT
ws["B3"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[3].height = 42


def build_projection_block(ws, row, name, ret_col_letter, infl_col_letter, initial_spend_value,
                            guardrails_ref, band_ref):
    """One deterministic year-by-year projection block. initial_spend_value: an Excel formula string
    for year-1 spend (today's £). guardrails_ref / band_ref: Excel refs (or literal) for the
    guardrails-on flag and band width used by this block."""
    headers = ["Year #", "Calendar year", "Return used", "Inflation used", "Cumulative inflation index",
               "Real spend level", "Spend taken", "End-of-year value"]
    for hi, h in enumerate(headers):
        ws.cell(row=row, column=2 + hi, value=h).font = HEADER_FONT
        ws.cell(row=row, column=2 + hi).fill = HEADER_FILL
    header_row = row
    row += 1
    c_year_num, c_cal_year, c_ret, c_infl, c_cuminfl, c_realspend, c_spend, c_pot = range(2, 10)

    ws.cell(row=row, column=c_year_num, value=0)
    ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}-1")
    ws.cell(row=row, column=c_cuminfl, value=1.0).number_format = "0.0000"
    ws.cell(row=row, column=c_realspend, value=initial_spend_value).number_format = GBP
    ws.cell(row=row, column=c_pot, value=f"={NR['pot']}").number_format = GBP
    row += 1

    for y_i in range(1, MAX_HORIZON + 1):
        prev = row - 1
        ws.cell(row=row, column=c_year_num, value=y_i)
        ws.cell(row=row, column=c_cal_year, value=f"={NR['start_year']}+{y_i}-1")
        avail_check = (f"AND(B{row}<={NR['horizon']},"
                        f"COUNTIF('Portfolio Annual Returns'!$A${par_first}:$A${par_last},C{row})>0)")
        ret_formula = (f'=IF({avail_check},'
                        f"INDEX('Portfolio Annual Returns'!${ret_col_letter}${par_first}:"
                        f"${ret_col_letter}${par_last},"
                        f"MATCH(C{row},'Portfolio Annual Returns'!$A${par_first}:$A${par_last},0)),\"\")")
        ws.cell(row=row, column=c_ret, value=ret_formula).number_format = PCT
        infl_formula = (f'=IF(D{row}="","",IF({avail_check},'
                         f"INDEX('Portfolio Annual Returns'!${infl_col_letter}${par_first}:"
                         f"${infl_col_letter}${par_last},"
                         f"MATCH(C{row},'Portfolio Annual Returns'!$A${par_first}:$A${par_last},0)),\"\"))")
        ws.cell(row=row, column=c_infl, value=infl_formula).number_format = PCT

        cuminfl_formula = f'=IF(D{row}="",F{prev},F{prev}*(1+E{row}))'
        ws.cell(row=row, column=c_cuminfl, value=cuminfl_formula).number_format = "0.0000"

        if y_i == 1:
            real_spend_formula = initial_spend_value
        else:
            prior_realspend = f"G{prev}"
            target_nominal = f"({prior_realspend}*F{row})"
            wr_now = f"IF(I{prev}>0,{target_nominal}/I{prev},9^9)"
            wr0_ref = "$C$6"  # this block's own initial withdrawal rate cell (set by caller, col C row6 per block - see below)
            raw_adj = (
                f'IF({wr_now}>{wr0_ref}*(1+{band_ref}),{prior_realspend}*(1-{NR["cut"]}),'
                f'IF({wr_now}<{wr0_ref}*(1-{band_ref}),{prior_realspend}*(1+{NR["raise"]}),{prior_realspend}))'
            )
            real_spend_formula = (
                f'=IF(D{row}="",{prior_realspend},'
                f'IF({guardrails_ref}<>"Y",{prior_realspend},'
                f'MIN(MAX({raw_adj},0.5*{initial_spend_value.lstrip("=")}),1.5*{initial_spend_value.lstrip("=")})))'
            )
        ws.cell(row=row, column=c_realspend, value=real_spend_formula).number_format = GBP

        spend_formula = f'=IF(D{row}="","",MIN(G{row}*F{row},MAX(I{prev},0)))'
        ws.cell(row=row, column=c_spend, value=spend_formula).number_format = GBP

        pot_formula = f'=IF(D{row}="",I{prev},MAX(I{prev}-H{row},0)*(1+D{row}))'
        ws.cell(row=row, column=c_pot, value=pot_formula).number_format = GBP
        row += 1

    last_data_row = row - 1
    first_data_row = header_row + 2
    return header_row, first_data_row, last_data_row


row = 5

# =======================================================================
# Section A: sensitivity to withdrawal rate (guardrails per Inputs toggle)
# =======================================================================
ws.cell(row=row, column=2, value="A. Sensitivity to initial withdrawal rate "
                                   "(guardrails apply if Inputs!guardrails_on = \"Y\")").font = BOLD
row += 2

for name in PORTFOLIOS:
    ret_col_letter = get_column_letter(pcols[name])
    infl_col_letter = get_column_letter(infl_col)

    ws.cell(row=row, column=2, value=f"{name} portfolio").font = BOLD
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = SUBHEAD_FILL
    row += 1
    summary_header_row = row
    for i, h in enumerate(["Withdrawal rate", "Initial spend (£)", "Historical final legacy",
                            "Any shortfall on hist. path?"]):
        cell = ws.cell(row=row, column=2 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    summary_rows_for_blocks = []
    for wr in WR_GRID:
        ws.cell(row=row, column=2, value=wr).number_format = PCT
        ws.cell(row=row, column=3, value=f"={wr}*{NR['pot']}").number_format = GBP
        summary_rows_for_blocks.append(row)
        row += 1
    row += 2

    for i, wr in enumerate(WR_GRID):
        srow = summary_rows_for_blocks[i]
        ws.cell(row=row, column=2, value=f"Withdrawal rate = {wr:.1%}").font = BOLD
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = SUBHEAD_FILL
        row += 1
        # a small "wr0" helper cell for this block (referenced as $C$header_row+... via wr0_ref pattern
        # below - simplest is to just hardcode the numeric wr since it's a known grid constant)
        block_header_row = row
        initial_spend_expr = f"=C{srow}"
        # patch build_projection_block's wr0_ref usage: pass the literal wr value directly since it's
        # a fixed grid constant for this block (avoids needing an extra helper cell)
        header_row, first_data_row, last_data_row = build_projection_block(
            ws, row, name, ret_col_letter, infl_col_letter, initial_spend_expr,
            guardrails_ref=NR["guardrails_on"], band_ref=NR["band"],
        )
        # the block above used a placeholder wr0 ref of $C$6 (wrong) - fix it now the block exists:
        for rr in range(first_data_row, last_data_row + 1):
            cell = ws.cell(row=rr, column=7)  # Real spend level column (G)
            if cell.value and isinstance(cell.value, str) and "$C$6" in cell.value:
                cell.value = cell.value.replace("$C$6", str(wr))
        row = last_data_row + 2

        ws.cell(row=srow, column=4, value=f"=I{last_data_row}").number_format = GBP
        spend_range = f"H{first_data_row}:H{last_data_row}"
        ws.cell(row=srow, column=5,
                value=f'=IF(COUNTIF({spend_range},"<"&C{srow}*0.999)>0,"Yes","No")')
    row += 1

row += 2

# =======================================================================
# Section B: sensitivity to guardrail band (guardrails forced ON; spend = Inputs!spend)
# =======================================================================
ws.cell(row=row, column=2, value="B. Sensitivity to guardrail band width "
                                   "(guardrails forced ON here, independent of the Inputs toggle; "
                                   "spend = Inputs!spend)").font = BOLD
row += 1
ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
row += 1

for name in PORTFOLIOS:
    ret_col_letter = get_column_letter(pcols[name])
    infl_col_letter = get_column_letter(infl_col)

    ws.cell(row=row, column=2, value=f"{name} portfolio").font = BOLD
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = SUBHEAD_FILL
    row += 1
    summary_header_row = row
    for i, h in enumerate(["Guardrail band (±)", "Historical final legacy", "Any shortfall on hist. path?",
                            "Years with a spend cut/raise"]):
        cell = ws.cell(row=row, column=2 + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    summary_rows_for_blocks = []
    for band in BAND_GRID:
        ws.cell(row=row, column=2, value=band).number_format = PCT
        summary_rows_for_blocks.append(row)
        row += 1
    row += 2

    for i, band in enumerate(BAND_GRID):
        srow = summary_rows_for_blocks[i]
        ws.cell(row=row, column=2, value=f"Guardrail band = ±{band:.0%}").font = BOLD
        for c in range(2, 10):
            ws.cell(row=row, column=c).fill = SUBHEAD_FILL
        row += 1
        initial_spend_expr = f"={NR['spend']}"
        header_row, first_data_row, last_data_row = build_projection_block(
            ws, row, name, ret_col_letter, infl_col_letter, initial_spend_expr,
            guardrails_ref='"Y"', band_ref=str(band),
        )
        # this block's wr0 is the model's normal initial withdrawal rate (Inputs!wr0), not a swept
        # value - fix the placeholder $C$6 reference accordingly
        for rr in range(first_data_row, last_data_row + 1):
            cell = ws.cell(row=rr, column=7)  # Real spend level column (G)
            if cell.value and isinstance(cell.value, str) and "$C$6" in cell.value:
                cell.value = cell.value.replace("$C$6", NR["wr0"])
        row = last_data_row + 2

        ws.cell(row=srow, column=3, value=f"=I{last_data_row}").number_format = GBP
        spend_range = f"H{first_data_row}:H{last_data_row}"
        ws.cell(row=srow, column=4,
                value=f'=IF(COUNTIF({spend_range},"<"&{NR["spend"]}*0.999)>0,"Yes","No")')
        realspend_range = f"G{first_data_row}:G{last_data_row}"
        ws.cell(row=srow, column=5,
                value=f'=SUMPRODUCT((({realspend_range}<>{NR["spend"]})*1))')
    row += 1

wb.save(OUT_PATH)
print("Saved stage 9 (Sensitivity Tables). WR grid:", WR_GRID, "Band grid:", BAND_GRID)
