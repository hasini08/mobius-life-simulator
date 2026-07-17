"""
One-off migration: wires the validated "Better v4" construction (see build_better_v4_summary.py)
into the ACTUAL app data - not just a standalone script - by:

  1. Extracting the 11 asset-class return series Better v4 uses from the previous Mobius model's
     own 'Asset Returns' tab (same source/columns as build_better_v4_summary.py) and merging them
     into data/asset_class_returns.csv (outer join on date - these series cover 2001-2025, shorter
     than the main 1999/2000-2026 window, so Better's own usable window will naturally shrink to
     match, exactly like Four Seasons already does with its own shorter-history holdings - no
     engine.py changes needed, dropna() already handles this per-portfolio).
  2. Adding those 11 labels to data/asset_class_map.csv.
  3. Replacing data/portfolio_holdings.csv's "Better" rows with the Better v4 weights (Berenberg /
     Protected Equities = Eq Gbl DM Novum Mgd Vol, confirmed - not a proxy), at the same flat 7bps
     fee used for Mobius's other portfolios.

Run once: `python migrate_better_v4_into_main_data.py`. Safe to re-run (idempotent - overwrites the
same rows each time rather than duplicating).
"""
import numbers
import openpyxl
import pandas as pd

V2_MODEL_PATH = r"C:\Users\YahampathH\OneDrive - Mobius Life Limited\Documents\Copy of Mobius decumulation model v2 (004).xlsm"

DATA_DIR_STR = "../data"
ASSET_RETURNS_CSV = f"{DATA_DIR_STR}/asset_class_returns.csv"
ASSET_MAP_CSV = f"{DATA_DIR_STR}/asset_class_map.csv"
HOLDINGS_CSV = f"{DATA_DIR_STR}/portfolio_holdings.csv"

# label -> column index in the old model's 'Asset Returns' sheet (same mapping as
# build_better_v4_summary.py's OLD_MODEL_COLS, values reused as the Better v4 weights)
BETTER_V4_HOLDINGS = {
    "Global Agg Bonds":                        (25, 0.325),
    "Eq Gbl DM Quality Gross":                 (8,  0.125),
    "Eq Gbl DM Novum Mgd Vol":                 (14, 0.200),  # Berenberg / Protected Equities
    "Eq EM Net":                               (6,  0.025),
    "US HY Corp Bond":                         (22, 0.025),
    "US ABS":                                  (21, 0.050),
    "EM Corp Bond":                            (23, 0.025),
    "US Prop REITS":                           (19, 0.025),
    "Commod":                                  (18, 0.050),
    "Hedge Fund Credit Suisse":                (15, 0.075),
    "HF Trend":                                (16, 0.075),
}
FLAT_MOBIUS_FEE = 0.0007


def extract_old_model_returns() -> pd.DataFrame:
    wb = openpyxl.load_workbook(V2_MODEL_PATH, data_only=True, keep_vba=True)
    ws = wb["Asset Returns"]
    dates = [ws.cell(row=r, column=3).value for r in range(9, ws.max_row + 1)]
    series = {}
    for label, (col, _w) in BETTER_V4_HOLDINGS.items():
        vals = [ws.cell(row=r, column=col).value for r in range(9, ws.max_row + 1)]
        pairs = {d: v for d, v in zip(dates, vals) if d is not None and isinstance(v, numbers.Number)}
        series[label] = pd.Series(pairs)
    df = pd.DataFrame(series)
    df.index = pd.DatetimeIndex(df.index).to_period("M").to_timestamp("M")
    return df.sort_index()


def main():
    print("Extracting Better v4 asset-class data from the previous model...")
    new_returns = extract_old_model_returns()
    print(f"  {len(new_returns)} rows, {new_returns.index.min().date()} to {new_returns.index.max().date()}")

    print("Merging into data/asset_class_returns.csv...")
    asset_df = pd.read_csv(ASSET_RETURNS_CSV, index_col=0, parse_dates=True)
    already_present = [c for c in new_returns.columns if c in asset_df.columns]
    if already_present:
        asset_df = asset_df.drop(columns=already_present)  # idempotent re-run: replace, don't duplicate
    merged = asset_df.join(new_returns, how="outer")
    merged.to_csv(ASSET_RETURNS_CSV)
    print(f"  New shape: {merged.shape} (was {asset_df.shape})")

    print("Updating data/asset_class_map.csv...")
    ac_df = pd.read_csv(ASSET_MAP_CSV)
    ac_df = ac_df[~ac_df["Label"].isin(BETTER_V4_HOLDINGS.keys())]  # idempotent
    new_ac_rows = pd.DataFrame({"Label": list(BETTER_V4_HOLDINGS.keys()),
                                "BloombergColumn": list(BETTER_V4_HOLDINGS.keys())})
    pd.concat([ac_df, new_ac_rows], ignore_index=True).to_csv(ASSET_MAP_CSV, index=False)

    print("Updating data/portfolio_holdings.csv 'Better' rows to the v4 construction...")
    holdings_df = pd.read_csv(HOLDINGS_CSV)
    holdings_df = holdings_df[holdings_df["Portfolio"] != "Better"]  # drop old Better rows
    new_better_rows = pd.DataFrame([
        {"Portfolio": "Better", "Holding": label, "AssetClass": label, "Weight": weight,
         "OCF": FLAT_MOBIUS_FEE}
        for label, (_col, weight) in BETTER_V4_HOLDINGS.items()
    ])
    pd.concat([holdings_df, new_better_rows], ignore_index=True).to_csv(HOLDINGS_CSV, index=False)

    print("\nDone. 'Better' now uses the Better v4 construction in the live app.")


if __name__ == "__main__":
    main()
